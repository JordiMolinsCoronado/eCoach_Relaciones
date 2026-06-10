import os
import time
import urllib.request
import urllib.error
import json
import asyncio
import re
import csv
import shutil

from contextvars import ContextVar
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path

from dotenv import load_dotenv
from google import genai
from google.genai import types, errors
from openai import OpenAI
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

load_dotenv()

# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash-lite")
GEMINI_INPUT_PRICE_PER_1M = float(os.getenv("GEMINI_INPUT_PRICE_PER_1M", "0.10"))
GEMINI_OUTPUT_PRICE_PER_1M = float(os.getenv("GEMINI_OUTPUT_PRICE_PER_1M", "0.40"))

DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_MODEL = os.getenv("DEEPSEEK_MODEL", "deepseek-v4-pro")
LLM_MAX_OUTPUT_TOKENS = int(os.getenv("LLM_MAX_OUTPUT_TOKENS", "5000"))
SHOW_DEBUG_FOOTER = os.getenv("SHOW_DEBUG_FOOTER", "true").strip().lower() in {"1", "true", "yes", "on"}
LLM_PROVIDER_RETRIES = int(os.getenv("LLM_PROVIDER_RETRIES", "1"))
LLM_PROVIDER_RETRY_SLEEP_SECONDS = float(os.getenv("LLM_PROVIDER_RETRY_SLEEP_SECONDS", "2"))
DEEPSEEK_BASE_URL = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
DEEPSEEK_INPUT_PRICE_PER_1M = float(os.getenv("DEEPSEEK_INPUT_PRICE_PER_1M", "0.435"))
DEEPSEEK_OUTPUT_PRICE_PER_1M = float(os.getenv("DEEPSEEK_OUTPUT_PRICE_PER_1M", "0.87"))

SHOW_USAGE_NOTE = os.getenv(
    "SHOW_USAGE_NOTE",
    os.getenv("SHOW_GEMINI_USAGE_NOTE", "true"),
).lower() == "true"

LLM_PROVIDER_ORDER = [
    provider.strip().lower()
    for provider in os.getenv("LLM_PROVIDER_ORDER", "deepseek,gemini").split(",")
    if provider.strip()
]

ROUTER_ENABLED = os.getenv("ROUTER_ENABLED", "true").lower() == "true"
ROUTER_PROVIDER_ORDER = [
    provider.strip().lower()
    for provider in os.getenv("ROUTER_PROVIDER_ORDER", "deepseek,gemini").split(",")
    if provider.strip()
]
ROUTER_DEEPSEEK_MODEL = os.getenv("ROUTER_DEEPSEEK_MODEL", "deepseek-v4-flash")
ROUTER_GEMINI_MODEL = os.getenv("ROUTER_GEMINI_MODEL", GEMINI_MODEL)

ROUTER_DIRECT_REPLY_ROUTES = {"smalltalk", "help", "thanks_ack"}
ROUTER_ALLOWED_ROUTES = {
    "smalltalk",
    "help",
    "thanks_ack",
    "simple_response",
    "memory_update",
    "patrimonial_analysis",
    "deep_review",
    "followup_management",
    "unknown",
}
ROUTER_ALLOWED_DEPTHS = {"none", "light", "normal", "deep"}
ROUTER_ALLOWED_CONFIDENCE = {"low", "medium", "high"}
DEFAULT_ROUTER_CLIENT_FILES = [
    "quien_soy",
    "que_quiero",
    "que_tengo_que_hacer",
    "estilo_respuesta",
]

WEALTH_KNOWLEDGE_DIR = Path("Wealth_Knowledge")
AGENT_CONFIG_DIR = Path("AgentConfig")
INTERVENTION_PATTERNS_DIR = AGENT_CONFIG_DIR / "InterventionPatterns"
INITIAL_CLIENT_FILES_DIR = Path("InitialClientFiles")
APP_TEXTS_DIR = Path("AppTexts")

CLIENTS_DIR = Path("ClientData")
DEFAULT_CLIENT_NAME = os.getenv("ACTIVE_CLIENT_NAME", "Cliente1")

CURRENT_CLIENT_NAME: ContextVar[str] = ContextVar(
    "CURRENT_CLIENT_NAME",
    default=DEFAULT_CLIENT_NAME,
)

CLIENT_LOCKS: dict[str, asyncio.Lock] = {}


def get_client_lock(client_name: str) -> asyncio.Lock:
    """Return one lock per client folder.

    Different Telegram users can run concurrently, while each user's
    file reads/writes stay sequential and safe.
    """
    if client_name not in CLIENT_LOCKS:
        CLIENT_LOCKS[client_name] = asyncio.Lock()

    return CLIENT_LOCKS[client_name]


def client_locked_handler(handler):
    """Wrap a Telegram handler so it runs under the current client's lock."""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        activate_client_from_update(update)
        client_name = current_client_name()

        async with get_client_lock(client_name):
            await handler(update, context)

    return wrapper


def current_client_name() -> str:
    """Return the client currently active in this async context."""
    return CURRENT_CLIENT_NAME.get()


def set_current_client_name(client_name: str) -> None:
    """Set the client currently active in this async context."""
    CURRENT_CLIENT_NAME.set(client_name)


def active_client_dir() -> Path:
    """Return the active client's folder."""
    return CLIENTS_DIR / current_client_name()


def client_files() -> dict[str, Path]:
    """Return live Markdown file paths for the active client."""
    active_dir = active_client_dir()

    return {
        "quien_soy": active_dir / "quien_soy.md",
        "que_quiero": active_dir / "que_quiero.md",
        "que_tengo_que_hacer": active_dir / "que_tengo_que_hacer.md",
        "estilo_respuesta": active_dir / "estilo_respuesta.md",
        "agent_config": active_dir / "agent_config.md",
        "agent_observations": active_dir / "agent_observations.md",
        "proactivity_log": active_dir / "proactivity_log.md",
        "historial_interacciones": active_dir / "historial_interacciones.md",
    }



# Client files allowed in normal answer/orchestrator context.
# Raw logs, observations, buffers, backups and histories can contaminate future answers
# with old wording. They remain available for their own workflows, but not as normal context.
NORMAL_ORCHESTRATOR_CLIENT_KEYS = {
    "quien_soy",
    "que_quiero",
    "que_tengo_que_hacer",
    "estilo_respuesta",
}

BLOCKED_NORMAL_CONTEXT_CLIENT_KEYS = {
    "historial_interacciones",
    "agent_observations",
    "proactivity_log",
    "agent_config",
}

def normal_orchestrator_client_keys() -> list[str]:
    available = client_files()
    return [key for key in NORMAL_ORCHESTRATOR_CLIENT_KEYS if key in available]


def filter_normal_orchestrator_client_keys(keys) -> list[str]:
    available = set(client_files().keys())
    clean: list[str] = []

    for key in keys or []:
        key = str(key)
        if key not in available:
            continue
        if key in BLOCKED_NORMAL_CONTEXT_CLIENT_KEYS:
            continue
        if key not in NORMAL_ORCHESTRATOR_CLIENT_KEYS:
            continue
        if key not in clean:
            clean.append(key)

    return clean


def wealth_log_file() -> Path:
    return active_client_dir() / "wealth_logs.md"


def followup_triggers_file() -> Path:
    return active_client_dir() / "followup_triggers.json"


def followup_archive_file() -> Path:
    return active_client_dir() / "followup_archive.json"


def app_state_file() -> Path:
    return active_client_dir() / "app_state.json"


# ---------------------------------------------------------------------
# eCoach_Relaciones v3 project board / supported sovereignty layer
# ---------------------------------------------------------------------
# This layer keeps the existing Telegram + LLM engine, but adds the new
# project-based architecture: workflow board, path choice, dirty/stale flags,
# and Saturday pilot story.
#
# It is intentionally file-based and backward-compatible with the old
# ClientData/<telegram_id>/ layout.

ECOACH_PROJECT_ID = "saturday_demo_relationships_guided_agency"
ECOACH_PROJECT_TITLE = "Piloto sábado - agencia relacional guiada"

def ecoach_projects_dir(client_dir: Path | None = None) -> Path:
    base = client_dir if client_dir is not None else active_client_dir()
    return base / "projects"

def ecoach_project_dir(client_dir: Path | None = None) -> Path:
    return ecoach_projects_dir(client_dir) / f"project_{ECOACH_PROJECT_ID}"

def ecoach_workflow_board_file(client_dir: Path | None = None) -> Path:
    return ecoach_project_dir(client_dir) / "workflow_board.json"

def ecoach_project_file(client_dir: Path | None = None) -> Path:
    return ecoach_project_dir(client_dir) / "project.json"

def ecoach_now_iso() -> str:
    return datetime.now(ZoneInfo(APP_TIMEZONE)).isoformat(timespec="seconds")

def ecoach_initial_workflow_board() -> dict:
    return {
        "project_id": ECOACH_PROJECT_ID,
        "title": ECOACH_PROJECT_TITLE,
        "sections": {
            "quien_soy": {"status": "active", "notes": []},
            "que_quiero": {"status": "not_started", "notes": []},
            "documents": {"status": "not_started", "artifacts": []},
            "diagnosis": {"status": "not_started", "artifacts": []},
            "path_choice": {"status": "not_started", "selected": None},
            "provider_choice": {"status": "not_started", "selected": None},
            "mi_plan": {"status": "not_started", "artifacts": []},
            "onboarding": {"status": "not_started", "artifacts": []},
            "seguimientos": {"status": "not_started", "items": []},
            "side_threads": {"status": "not_started", "items": []},
        },
        "dirty_flags": [],
        "stale_flags": [],
        "needs_rebuild_flags": [],
        "history": [],
    }

def ecoach_initial_project() -> dict:
    return {
        "project_id": ECOACH_PROJECT_ID,
        "title": ECOACH_PROJECT_TITLE,
        "domain": "relaciones",
        "active_skill": "initial_discovery",
        "active_section": "quien_soy",
        "status": "active",
        "created_at": ecoach_now_iso(),
        "updated_at": ecoach_now_iso(),
    }

def ecoach_load_json(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return default

def ecoach_save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def ensure_ecoach_project_files(client_dir: Path | None = None) -> None:
    project_path = ecoach_project_file(client_dir)
    board_path = ecoach_workflow_board_file(client_dir)
    if not project_path.exists():
        ecoach_save_json(project_path, ecoach_initial_project())
    if not board_path.exists():
        ecoach_save_json(board_path, ecoach_initial_workflow_board())

def load_ecoach_board(client_dir: Path | None = None) -> dict:
    ensure_ecoach_project_files(client_dir)
    return ecoach_load_json(ecoach_workflow_board_file(client_dir), ecoach_initial_workflow_board())

def save_ecoach_board(board: dict, client_dir: Path | None = None) -> None:
    board["updated_at"] = ecoach_now_iso()
    ecoach_save_json(ecoach_workflow_board_file(client_dir), board)

def ecoach_set_section(board: dict, section: str, status: str | None = None, **fields) -> None:
    board.setdefault("sections", {})
    board["sections"].setdefault(section, {})
    if status is not None:
        board["sections"][section]["status"] = status
    board["sections"][section]["last_updated"] = ecoach_now_iso()
    for key, value in fields.items():
        board["sections"][section][key] = value

def ecoach_add_flag(board: dict, list_name: str, value: str) -> None:
    board.setdefault(list_name, [])
    if value not in board[list_name]:
        board[list_name].append(value)

def ecoach_add_history(board: dict, action: str, detail: dict | None = None) -> None:
    board.setdefault("history", [])
    board["history"].append(
        {
            "at": ecoach_now_iso(),
            "action": action,
            "detail": detail or {},
        }
    )

def ecoach_mark_documents_received(client_dir: Path | None = None, filename: str | None = None) -> None:
    board = load_ecoach_board(client_dir)
    artifacts = board.setdefault("sections", {}).setdefault("documents", {}).setdefault("artifacts", [])
    if filename:
        artifacts.append({"filename": filename, "received_at": ecoach_now_iso()})
    ecoach_set_section(board, "documents", "complete")
    ecoach_set_section(board, "diagnosis", "needs_rebuild")
    ecoach_add_flag(board, "dirty_flags", "diagnosis")
    ecoach_add_flag(board, "needs_rebuild_flags", "diagnosis")
    ecoach_add_history(board, "documents_received", {"filename": filename})
    save_ecoach_board(board, client_dir)

def ecoach_mark_documents_analyzed(client_dir: Path | None = None) -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "diagnosis", "active")
    ecoach_add_flag(board, "dirty_flags", "diagnosis")
    ecoach_add_history(board, "document_analysis_started")
    save_ecoach_board(board, client_dir)


def ecoach_mark_diagnosis_complete(client_dir: Path | None = None, source: str = "uploaded_documents") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "documents", "complete")
    ecoach_set_section(board, "diagnosis", "complete")
    ecoach_set_section(
        board,
        "path_choice",
        "waiting_user",
        label="Esperando elección de camino tras diagnóstico",
    )

    for flag_name in ["dirty_flags", "needs_rebuild_flags"]:
        flags = board.setdefault(flag_name, [])
        board[flag_name] = [flag for flag in flags if flag != "diagnosis"]

    ecoach_add_history(board, "diagnosis_completed", {"source": source})
    save_ecoach_board(board, client_dir)

def ecoach_select_path(path_key: str, path_label: str, client_dir: Path | None = None) -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "que_quiero", "complete")
    ecoach_set_section(board, "path_choice", "complete", selected=path_key, label=path_label)
    ecoach_set_section(board, "provider_choice", "waiting_user")
    ecoach_add_history(board, "path_selected", {"path": path_key, "label": path_label})
    save_ecoach_board(board, client_dir)

def ecoach_select_provider(provider_key: str, provider_label: str, client_dir: Path | None = None) -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "provider_choice", "complete", selected=provider_key, label=provider_label)
    ecoach_set_section(board, "mi_plan", "complete")
    ecoach_set_section(board, "onboarding", "waiting_user")
    ecoach_set_section(board, "seguimientos", "active")
    ecoach_add_history(board, "provider_selected", {"provider": provider_key, "label": provider_label})
    save_ecoach_board(board, client_dir)

def ecoach_revise_to_guided(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "que_quiero", "complete")
    ecoach_set_section(board, "path_choice", "complete", selected="guided", label="Agencia relacional guiada")
    ecoach_set_section(board, "provider_choice", "waiting_user")
    ecoach_set_section(board, "mi_plan", "stale")
    ecoach_set_section(board, "onboarding", "stale")
    ecoach_set_section(board, "seguimientos", "stale")
    for flag in ["mi_plan", "onboarding", "seguimientos"]:
        ecoach_add_flag(board, "stale_flags", flag)
    ecoach_add_history(board, "upstream_revision_to_guided", {"source": source})
    save_ecoach_board(board, client_dir)

def ecoach_mark_material_new_data(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(board, "diagnosis", "needs_rebuild")
    ecoach_set_section(board, "mi_plan", "stale")
    ecoach_set_section(board, "onboarding", "stale")
    ecoach_set_section(board, "seguimientos", "stale")
    ecoach_add_flag(board, "dirty_flags", "diagnosis")
    ecoach_add_flag(board, "needs_rebuild_flags", "diagnosis")
    for flag in ["proposal", "mi_plan", "onboarding", "seguimientos"]:
        ecoach_add_flag(board, "stale_flags", flag)
    ecoach_add_history(board, "material_new_financial_data", {"source": source})
    save_ecoach_board(board, client_dir)

def ecoach_open_side_thread(topic: str, attached_to: str = "onboarding", client_dir: Path | None = None) -> None:
    board = load_ecoach_board(client_dir)
    side = board.setdefault("sections", {}).setdefault("side_threads", {})
    side.setdefault("items", [])
    side["items"].append(
        {
            "topic": topic,
            "attached_to": attached_to,
            "opened_at": ecoach_now_iso(),
            "status": "active",
        }
    )
    ecoach_set_section(board, "side_threads", "active")
    ecoach_add_history(board, "side_thread_opened", {"topic": topic, "attached_to": attached_to})
    save_ecoach_board(board, client_dir)

def ecoach_normalize_text(text: str) -> str:
    raw = (text or "").lower()
    try:
        decomposed = unicodedata.normalize("NFKD", raw)
        raw = "".join(char for char in decomposed if not unicodedata.combining(char))
    except Exception:
        pass
    return raw


def ecoach_start_initial_portfolio_discovery(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "que_quiero",
        "complete",
        selected="understand_before_moving_money",
        label="Entender la dinámica antes de actuar",
    )
    ecoach_set_section(
        board,
        "documents",
        "waiting_user",
        label="Esperando relato, mensajes o hechos concretos",
    )
    ecoach_set_section(board, "diagnosis", "not_started")
    ecoach_set_section(board, "path_choice", "not_started")
    ecoach_set_section(board, "provider_choice", "not_started")
    ecoach_set_section(board, "mi_plan", "not_started")
    ecoach_set_section(board, "onboarding", "not_started")
    ecoach_set_section(board, "seguimientos", "not_started")
    ecoach_add_history(board, "initial_portfolio_discovery_started", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_choose_lower_cost_path(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "path_choice",
        "complete",
        selected="lower_cost_without_moving_money",
        label="Explorar alternativas de menor coste sin mover dinero todavía",
    )
    ecoach_set_section(
        board,
        "provider_choice",
        "waiting_user",
        label="Esperando comparación de proveedores/opciones",
    )
    ecoach_add_history(board, "path_choice_lower_cost", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_mark_options_comparison_requested(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "provider_choice",
        "complete",
        selected="compare_options_before_provider",
        label="Comparar opciones antes de elegir proveedor",
    )
    ecoach_set_section(
        board,
        "mi_plan",
        "waiting_user",
        label="Esperando elección de opción para crear Mi Plan",
    )
    ecoach_add_history(board, "options_comparison_requested", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_mark_guided_ecoach_requested(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "provider_choice",
        "complete",
        selected="guided_with_ecoach",
        label="Agencia relacional guiada con eCoach",
    )
    ecoach_set_section(
        board,
        "mi_plan",
        "waiting_user",
        label="Esperando datos de riesgo/proveedor para crear Mi Plan",
    )
    ecoach_add_history(board, "guided_ecoach_requested", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_mark_guided_mi_plan_prepared(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "mi_plan",
        "complete",
        selected="provisional_guided_conservative_plan",
        label="Mi Plan guiado provisional conservador",
    )
    ecoach_set_section(
        board,
        "onboarding",
        "waiting_user",
        label="Esperando decisión sobre plataforma/datos fiscales antes de ejecutar",
    )
    ecoach_add_history(board, "guided_mi_plan_prepared", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_mark_bank_message_sent(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)
    ecoach_set_section(
        board,
        "onboarding",
        "complete",
        selected="bank_message_sent",
        label="Mensaje al banco enviado",
    )
    ecoach_set_section(
        board,
        "seguimientos",
        "waiting_user",
        label="Esperando respuesta del banco",
    )
    ecoach_add_history(board, "bank_message_sent", {"source": source})
    save_ecoach_board(board, client_dir)


def ecoach_mark_bank_response_incomplete(client_dir: Path | None = None, source: str = "free_text") -> None:
    board = load_ecoach_board(client_dir)

    # Repair/confirm upstream project state.
    ecoach_set_section(
        board,
        "documents",
        "complete",
        label="Documentos iniciales analizados",
    )
    ecoach_set_section(
        board,
        "diagnosis",
        "complete",
        label="Diagnóstico inicial completado",
    )
    ecoach_set_section(
        board,
        "path_choice",
        "complete",
        selected="lower_cost_without_moving_money",
        label="Explorar alternativas de menor coste sin mover dinero",
    )
    ecoach_set_section(
        board,
        "provider_choice",
        "complete",
        selected="guided_with_ecoach",
        label="Agencia relacional guiada con eCoach",
    )
    ecoach_set_section(
        board,
        "mi_plan",
        "complete",
        selected="provisional_guided_conservative_plan",
        label="Mi Plan guiado provisional conservador",
    )
    ecoach_set_section(
        board,
        "onboarding",
        "complete",
        selected="bank_message_sent",
        label="Mensaje al banco enviado",
    )
    ecoach_set_section(
        board,
        "seguimientos",
        "waiting_user",
        selected="bank_response_incomplete",
        label="Respuesta del banco incompleta; esperando datos faltantes",
    )

    board["dirty"] = []
    board["stale"] = []
    board["needs_rebuild"] = []

    ecoach_add_history(board, "bank_response_incomplete", {"source": source})
    save_ecoach_board(board, client_dir)

def ecoach_detect_free_text_control(text: str) -> str | None:
    normalized = ecoach_normalize_text(text)

    bank_response_incomplete_signal = (
        (
            "banco me ha respondido" in normalized
            or "el banco me ha respondido" in normalized
            or "respuesta del banco" in normalized
            or "han respondido" in normalized
            or "me han enviado" in normalized
        )
        and (
            "solo me ha enviado" in normalized
            or "sólo me ha enviado" in normalized
            or "solo han enviado" in normalized
            or "sólo han enviado" in normalized
            or "lista de fondos" in normalized
            or "con importes" in normalized
            or "no veo plusvalias" in normalized
            or "no veo plusvalías" in normalized
            or "no veo costes" in normalized
            or "costes ex-post" in normalized
            or "incompleta" in normalized
            or "faltan datos" in normalized
        )
    )

    if bank_response_incomplete_signal:
        return "bank_response_incomplete"


    bank_message_sent_signal = (
        (
            "he enviado el mensaje al banco" in normalized
            or "ya he enviado el mensaje al banco" in normalized
            or "mensaje al banco enviado" in normalized
            or "lo he enviado al banco" in normalized
            or "ya lo he enviado al banco" in normalized
            or "he escrito al banco" in normalized
            or "ya he escrito al banco" in normalized
        )
        and (
            "banco" in normalized
            or "respondan" in normalized
            or "respuesta" in normalized
        )
    )

    if bank_message_sent_signal:
        return "bank_message_sent"


    guided_mi_plan_signal = (
        (
            "mi plan" in normalized
            or "plan guiado" in normalized
            or "primer mi plan" in normalized
            or "primer plan" in normalized
            or "propuesta conservadora" in normalized
            or "propuesta conservadora de menor coste" in normalized
        )
        and (
            "guiado" in normalized
            or "ecoach" in normalized
            or "provisional" in normalized
            or "sin ejecucion" in normalized
            or "sin ejecución" in normalized
            or "menor coste" in normalized
        )
    )

    if guided_mi_plan_signal:
        return "prepare_guided_mi_plan"


    guided_ecoach_signal = (
        (
            "cartera guiada" in normalized
            or "guiada con ecoach" in normalized
            or "con vuestra ayuda" in normalized
            or "con tu ayuda" in normalized
            or "hacerlo con ecoach" in normalized
            or "hacerlo con vuestra ayuda" in normalized
            or "profundizar primero en la cartera guiada" in normalized
        )
        and (
            "ecoach" in normalized
            or "vuestra ayuda" in normalized
            or "tu ayuda" in normalized
            or "guiada" in normalized
        )
    )

    if guided_ecoach_signal:
        return "guided_ecoach_path"


    comparison_table_signal = (
        (
            "tabla comparativa" in normalized
            or "primera tabla" in normalized
            or "prepara una tabla" in normalized
            or "preparar una tabla" in normalized
            or "comparativa" in normalized
            or "comparar opciones" in normalized
            or "comparar alternativas" in normalized
        )
        and (
            "sin mover dinero" in normalized
            or "no mover dinero" in normalized
            or "sin ejecutar" in normalized
            or "solo comparar" in normalized
            or "sólo comparar" in normalized
            or "primera" in normalized
        )
    )

    if comparison_table_signal:
        return "prepare_options_comparison"


    lower_cost_signal = (
        ("alternativas" in normalized or "opciones" in normalized or "comparar" in normalized)
        and (
            "menor coste" in normalized
            or "menos coste" in normalized
            or "bajo coste" in normalized
            or "coste bajo" in normalized
            or "costes bajos" in normalized
            or "más barato" in normalized
            or "mas barato" in normalized
            or "baratas" in normalized
            or "baratos" in normalized
        )
    )

    if lower_cost_signal:
        return "lower_cost_alternatives"


    guided_signal = (
        "guiada" in normalized
        or "cartera guiada" in normalized
        or "con vuestra ayuda" in normalized
        or "con tu ayuda" in normalized
        or "soberania acompanada" in normalized
    )

    explicit_revision_signal = (
        "espera" in normalized
        or "cambio de idea" in normalized
        or "he cambiado de idea" in normalized
        or "lo he pensado" in normalized
        or "vuelve a ser" in normalized
        or "no quiero deleg" in normalized
        or "prefiero" in normalized
        or "quiero cambiar" in normalized
        or "quiero cartera guiada" in normalized
        or "quiero una cartera guiada" in normalized
        or "creo que quiero cartera guiada" in normalized
    )

    prior_or_rejected_delegation_signal = (
        "delegar" in normalized
        or "delegada" in normalized
        or "gestion delegada" in normalized
        or "delegacion" in normalized
    )

    exploratory_uncertainty_signal = (
        "no se si" in normalized
        or "no sé si" in normalized
        or "dudo" in normalized
        or "dudando" in normalized
        or "estoy dudando" in normalized
        or "deberia" in normalized
        or "debería" in normalized
        or "opciones" in normalized
        or "comparar" in normalized
        or "entender" in normalized
    )

    if (
        guided_signal
        and explicit_revision_signal
        and prior_or_rejected_delegation_signal
        and not exploratory_uncertainty_signal
    ):
        return "revise_to_guided"

    if (
        ("olvid" in normalized or "falt" in normalized or "me deje" in normalized)
        and (
            "50.000" in normalized
            or "50000" in normalized
            or "caixabank" in normalized
            or "otro fondo" in normalized
            or "otros fondos" in normalized
        )
    ):
        return "material_new_data"

    initial_portfolio_discovery_signal = (
        ("fondos" in normalized or "cartera" in normalized)
        and (
            "banco" in normalized
            or "pagando demasiado" in normalized
            or "costes" in normalized
            or "comisiones" in normalized
            or "no quiero mover dinero" in normalized
            or "entender que tengo" in normalized
            or "entender qué tengo" in normalized
            or "opciones reales" in normalized
            or "que necesitas" in normalized
            or "qué necesitas" in normalized
        )
    )

    if initial_portfolio_discovery_signal:
        return "initial_portfolio_discovery"

    if (
        ("miedo" in normalized or "temor" in normalized or "me asusta" in normalized or "me da cosa" in normalized)
        and ("myinvestor" in normalized or "abrir" in normalized or "cuenta" in normalized)
    ):
        return "fear_opening_account"

    if ("indexa" in normalized) and ("tranquilidad" in normalized or "prefiero" in normalized or "elijo" in normalized):
        return "provider_indexa"

    return None

def ecoach_board_summary_text(client_dir: Path | None = None) -> str:
    board = load_ecoach_board(client_dir)
    sections = board.get("sections", {})
    lines = [
        "Estado del proyecto eCoach Relaciones",
        "",
        f"Proyecto: {board.get('title', ECOACH_PROJECT_TITLE)}",
        "",
        "Secciones:",
    ]
    for key in [
        "quien_soy",
        "que_quiero",
        "documents",
        "diagnosis",
        "path_choice",
        "provider_choice",
        "mi_plan",
        "onboarding",
        "seguimientos",
        "side_threads",
    ]:
        item = sections.get(key, {})
        extra = ""
        if item.get("selected"):
            extra = f" → {item.get('selected')}"
        lines.append(f"- {key}: {item.get('status', 'unknown')}{extra}")

    lines.extend(
        [
            "",
            f"Dirty: {', '.join(board.get('dirty_flags', [])) or '—'}",
            f"Stale: {', '.join(board.get('stale_flags', [])) or '—'}",
            f"Needs rebuild: {', '.join(board.get('needs_rebuild_flags', [])) or '—'}",
        ]
    )
    return "\n".join(lines)

async def show_workflow_board(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()
    ensure_ecoach_project_files()
    await update.effective_message.reply_text(
        ecoach_board_summary_text(),
        reply_markup=MAIN_KEYBOARD,
    )


async def reset_ecoach_project(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    client_dir = active_client_dir()
    project_dir = ecoach_project_dir(client_dir)
    project_dir.mkdir(parents=True, exist_ok=True)

    ecoach_save_json(ecoach_project_file(client_dir), ecoach_initial_project())
    ecoach_save_json(ecoach_workflow_board_file(client_dir), ecoach_initial_workflow_board())

    await update.effective_message.reply_text(
        "Proyecto eCoach reiniciado. La demo vuelve al estado inicial.\n\n"
        + ecoach_board_summary_text(client_dir),
        reply_markup=MAIN_KEYBOARD,
    )


def load_skill_text(skill_name: str) -> str:
    skill_path = Path(__file__).resolve().parent / "skills" / f"{skill_name}.md"
    if not skill_path.exists():
        raise RuntimeError(f"Missing skill file: {skill_path}")
    return skill_path.read_text(encoding="utf-8").strip()


def generate_initial_discovery_reply(user_text: str) -> str:
    skill = load_skill_text("initial_discovery")

    prompt = f"""USER MESSAGE:
{user_text}

TASK:
Write the first eCoach Relaciones response following the skill instructions.

Return only the client-facing Spanish answer.
Do not return JSON.
Do not mention that you are following a skill.
"""

    answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=skill,
    )

    try:
        save_wealth_log(
            user_input=user_text,
            ai_answer=answer,
            usage=usage,
        )
    except Exception:
        pass

    return answer.strip()



def generate_skill_client_reply(skill_name: str, task: str, facts: str) -> str:
    skill = load_skill_text(skill_name)

    prompt = f"""TASK:
{task}

FACTS / STRUCTURED CONTENT PROVIDED BY CODE:
{facts}

Write the client-facing answer following the skill instructions.

Return only the Spanish answer.
Do not return JSON.
Do not mention the skill.
"""

    answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=skill,
    )

    try:
        save_wealth_log(
            user_input=f"[skill:{skill_name}] {task}\n\n{facts}",
            ai_answer=answer,
            usage=usage,
        )
    except Exception:
        pass

    return answer.strip()


async def answer_callback_with_skill(
    query,
    skill_name: str,
    task: str,
    facts: str,
    reply_markup=None,
) -> None:
    thinking_message = await query.message.reply_text("Pensando...")

    try:
        answer = await asyncio.to_thread(
            generate_skill_client_reply,
            skill_name,
            task,
            facts,
        )
    except Exception as error:
        try:
            await thinking_message.delete()
        except Exception:
            pass

        await query.message.reply_text(
            f"No he podido preparar la respuesta: {error}",
            reply_markup=reply_markup,
        )
        return

    try:
        await thinking_message.delete()
    except Exception:
        pass

    await query.message.reply_text(answer, reply_markup=reply_markup)


async def reply_initial_discovery_with_llm(update: Update, user_text: str) -> None:
    thinking_message = await update.message.reply_text("Pensando...")

    try:
        answer = await asyncio.to_thread(generate_initial_discovery_reply, user_text)
    except Exception as error:
        try:
            await thinking_message.delete()
        except Exception:
            pass

        await update.message.reply_text(
            f"No he podido preparar la respuesta inicial: {error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        await thinking_message.delete()
    except Exception:
        pass

    await update.message.reply_text(answer, reply_markup=alternatives_path_keyboard())


async def try_handle_ecoach_control_message(update: Update, context: ContextTypes.DEFAULT_TYPE, user_text: str) -> bool:
    """Handle key project-state messages deterministically before the generic LLM."""
    action = ecoach_detect_free_text_control(user_text)
    if action is None:
        return False

    client_dir = active_client_dir()

    if action == "revise_to_guided":
        ecoach_revise_to_guided(client_dir, source="free_text")
        await update.message.reply_text(
            "Esto cambia una decisión importante, y está bien. No es una contradicción: es más claridad.\n\n"
            "El plan anterior con gestión delegada queda como versión antigua. Ahora actualizo `Qué quiero`: cartera guiada.\n\n"
            "Marco como desactualizados `Mi Plan`, onboarding y seguimientos, porque dependían de la decisión anterior.\n\n"
            "Nueva dirección: no autogestionada sola, no delegada a ciegas. Cartera guiada: soberanía acompañada.",
            reply_markup=provider_comparison_keyboard(),
        )
        return True

    if action == "material_new_data":
        ecoach_mark_material_new_data(client_dir, source="free_text")
        await update.message.reply_text(
            "Esto sí cambia el análisis. No pasa nada: precisamente por eso trabajamos con versiones.\n\n"
            "Si aparecen más fondos o más importe —por ejemplo otros 50.000 € en CaixaBank— el diagnóstico debe reconstruirse.\n\n"
            "Marco diagnóstico como `needs_rebuild` y dejo propuesta, Mi Plan, onboarding y seguimientos como `stale`.\n\n"
            "Siguiente paso: integramos ese dato antes de volver a proponer caminos.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "initial_portfolio_discovery":
        ecoach_start_initial_portfolio_discovery(client_dir, source="free_text")
        await reply_initial_discovery_with_llm(update, user_text)
        return True

    if action == "lower_cost_alternatives":
        ecoach_choose_lower_cost_path(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Perfecto. Exploramos alternativas de menor coste, pero sin mover dinero todavía.\n\n"
            "Ahora el objetivo no es ejecutar nada ni elegir proveedor a ciegas. El objetivo es comparar con calma.\n\n"
            "Prepararé la comparación en cuatro bloques:\n"
            "1. cartera actual en el banco;\n"
            "2. alternativas de menor coste con fondos indexados o RTO;\n"
            "3. gestión delegada tipo roboadvisor;\n"
            "4. cartera guiada con eCoach, donde usted mantiene el control y yo le ayudo a entender y ejecutar si decide hacerlo.\n\n"
            "Para hacerlo bien, compararemos coste total, riesgo aproximado, simplicidad, fiscalidad, liquidez y grado de control.\n\n"
            "Siguiente paso: puedo preparar una primera tabla comparativa sin mover dinero.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "bank_response_incomplete":
        ecoach_mark_bank_response_incomplete(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Entendido. La respuesta del banco es útil, pero incompleta.\n\n"
            "De momento podemos usar la lista de fondos e importes como confirmación parcial de posiciones, pero todavía no deberíamos ejecutar ningún cambio.\n\n"
            "Faltan datos críticos:\n"
            "- plusvalías o minusvalías latentes por fondo;\n"
            "- informe ex-post de costes;\n"
            "- costes reales de gestión, depósito, custodia, asesoramiento o retrocesiones;\n"
            "- confirmación de si los fondos son traspasables sin impacto fiscal;\n"
            "- fechas y precios de adquisición.\n\n"
            "Mi Plan guiado sigue siendo válido como versión provisional, pero no pasa todavía a versión ejecutable.\n\n"
            "Siguiente paso: preparar una segunda petición al banco, más precisa y más difícil de responder de forma incompleta.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "bank_message_sent":
        ecoach_mark_bank_message_sent(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Perfecto. Queda registrado: el mensaje al banco ya está enviado.\n\n"
            "Ahora no hay que hacer más fuerza. El siguiente paso depende de la respuesta del banco.\n\n"
            "Cuando respondan, puede subir aquí el email, PDF, captura o texto. Yo lo convertiré en una tabla clara con:\n"
            "- posiciones actuales;\n"
            "- plusvalías/minusvalías;\n"
            "- costes reales;\n"
            "- traspasabilidad;\n"
            "- datos que falten;\n"
            "- y cómo afecta eso al Mi Plan guiado.\n\n"
            "Si la respuesta del banco es incompleta, también lo detectaremos y prepararé una segunda petición más precisa.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "prepare_guided_mi_plan":
        ecoach_mark_guided_mi_plan_prepared(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Perfecto. Preparo un primer Mi Plan guiado, provisional y sin ejecución.\n\n"
            "Mi Plan guiado — versión provisional\n\n"
            "1. Punto de partida\n"
            "- Cartera actual aproximada: 20.888 €.\n"
            "- Cuatro fondos Ruralvía/Gescooperativo.\n"
            "- Coste visible aproximado: 1,41% anual.\n"
            "- Perfil conservador, con alrededor de un 20% de renta variable.\n\n"
            "2. Objetivo\n"
            "- Reducir costes.\n"
            "- Mantener un perfil conservador.\n"
            "- No mover dinero todavía.\n"
            "- Entender antes de actuar.\n\n"
            "3. Propuesta inicial\n"
            "- Construir una cartera conservadora de menor coste.\n"
            "- Usar una plataforma posible como MyInvestor, Renta 4, Openbank u otra equivalente.\n"
            "- Priorizar fondos indexados, clases limpias o instrumentos con costes transparentes.\n"
            "- Evitar ejecutar nada hasta revisar fiscalidad, traspasabilidad y costes reales de salida.\n\n"
            "4. Diseño provisional orientativo\n"
            "- Mantener una parte defensiva amplia en renta fija monetaria/corto plazo o fondos conservadores de bajo coste.\n"
            "- Mantener una parte limitada de renta variable global, coherente con un perfil conservador.\n"
            "- Evitar concentración innecesaria en fondos mixtos caros si se puede separar renta fija y renta variable con más claridad.\n\n"
            "5. Datos necesarios antes de ejecutar\n"
            "- Plusvalías o minusvalías latentes de cada fondo.\n"
            "- Si los fondos son traspasables sin impacto fiscal.\n"
            "- Costes reales ex-post del banco.\n"
            "- Plataforma candidata.\n"
            "- Confirmación del nivel de riesgo deseado.\n\n"
            "6. Próxima acción\n"
            "Antes de ejecutar, lo más prudente es pedir al banco dos cosas: informe fiscal de posiciones y desglose de costes reales ex-post. Con eso, Mi Plan puede pasar de provisional a ejecutable.\n\n"
            "Puedo preparar ahora el mensaje exacto para pedir esos datos al banco.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "guided_ecoach_path":
        board = load_ecoach_board(active_client_dir())
        documents_state = board.get("sections", {}).get("documents", {}).get("status")
        diagnosis_state = board.get("sections", {}).get("diagnosis", {}).get("status")

        if documents_state != "complete" or diagnosis_state != "complete":
            board = load_ecoach_board(active_client_dir())
            ecoach_set_section(
                board,
                "que_quiero",
                "complete",
                selected="understand_before_moving_money",
                label="Entender antes de decidir o mover dinero",
            )
            ecoach_set_section(
                board,
                "documents",
                "waiting_user",
                label="Esperando relato o hechos de la situación",
            )
            save_ecoach_board(board, active_client_dir())
            await reply_initial_discovery_with_llm(update, user_text)
            return True

        ecoach_mark_guided_ecoach_requested(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Perfecto. Entonces profundizamos primero en la cartera guiada con eCoach.\n\n"
            "La idea es sencilla: usted mantiene el control. No delega a ciegas, no mueve dinero todavía, y usamos la información de la cartera actual para construir una propuesta paso a paso.\n\n"
            "Antes de crear Mi Plan, necesito concretar tres decisiones:\n"
            "1. riesgo deseado: mantener perfil conservador, subir un poco riesgo, o reducirlo;\n"
            "2. plataforma posible: por ejemplo MyInvestor, Renta 4, Openbank u otra;\n"
            "3. rediseño: mantener parte de la cartera actual o construir una cartera nueva de menor coste.\n\n"
            "Siguiente paso: puedo preparar un primer Mi Plan guiado, provisional y sin ejecución, con una propuesta conservadora de menor coste.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "prepare_options_comparison":
        ecoach_mark_options_comparison_requested(active_client_dir(), source="free_text")
        await update.message.reply_text(
            "Perfecto. Preparo una primera tabla comparativa sin mover dinero.\n\n"
            "Partimos de la cartera ya leída: unos 20.888 €, cuatro fondos Ruralvía/Gescooperativo, coste visible aproximado de 1,41% anual y perfil conservador con alrededor de un 20% de renta variable.\n\n"
            "| Camino | Qué significa | Coste esperado | Control | Ventaja | Riesgo / límite |\n"
            "|---|---|---:|---|---|---|\n"
            "| Mantener cartera actual | Seguir en Ruralvía, pero pidiendo más transparencia | Alto: visible ~1,41% + posibles costes extra | Bajo-medio | No hay cambios ni fricción | Coste alto y dependencia del banco |\n"
            "| Cartera indexada / RTO | Usar fondos indexados o clases limpias en una plataforma tipo MyInvestor/Renta 4/Openbank | Bajo-medio: aprox. 0,2%-0,6% según fondos/plataforma | Alto | Menor coste y más control | Requiere entender y ejecutar órdenes |\n"
            "| Gestión delegada / roboadvisor | Delegar la cartera en un proveedor tipo Indexa u otro roboadvisor | Medio: normalmente menor que banco tradicional | Bajo | Simplicidad y disciplina | Sigue siendo delegación; menos control fino |\n"
            "| Cartera guiada con eCoach | Usted mantiene la cuenta/proveedor; eCoach ayuda a comparar, decidir y ejecutar paso a paso | Variable: depende del proveedor elegido | Alto | Soberanía acompañada: entiende antes de actuar | Requiere implicación del cliente |\n\n"
            "Lectura inicial: parece razonable explorar alternativas de menor coste, pero no movería nada todavía hasta tener datos fiscales y costes ex-post del banco.\n\n"
            "Siguiente paso: elegir qué comparación quiere profundizar primero: RTO/indexada, gestión delegada/roboadvisor, o cartera guiada con eCoach.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    if action == "fear_opening_account":
        ecoach_open_side_thread("fear_opening_provider_account", "onboarding", client_dir)
        await update.message.reply_text(
            "Tiene sentido que abrir una cuenta nueva dé respeto. No significa mover dinero hoy.\n\n"
            "Lo bajamos a pasos pequeños:\n"
            "1. mirar requisitos de apertura;\n"
            "2. comprobar seguridad y funcionamiento;\n"
            "3. abrir sin traspasar todavía, si encaja;\n"
            "4. decidir después con calma.\n\n"
            "Este miedo abre un hilo lateral. No invalida el plan principal.",
            reply_markup=provider_selected_keyboard(),
        )
        return True

    if action == "provider_indexa":
        ecoach_select_provider("indexa", "Indexa", client_dir)
        await update.message.reply_text(
            build_provider_selected_message("Indexa", "indexa"),
            reply_markup=provider_selected_keyboard(),
        )
        return True

    return False



def client_name_from_update(update: Update) -> str:
    """Map one Telegram chat to one isolated client folder."""
    if update.effective_chat is None:
        return DEFAULT_CLIENT_NAME

    return f"telegram_{update.effective_chat.id}"


def activate_client_from_update(update: Update) -> None:
    """Activate the correct client folder for this Telegram update."""
    set_current_client_name(client_name_from_update(update))


def iter_telegram_client_names() -> list[str]:
    """Return all Telegram-based client folder names."""
    if not CLIENTS_DIR.exists():
        return []

    return sorted(
        path.name
        for path in CLIENTS_DIR.iterdir()
        if path.is_dir() and path.name.startswith("telegram_")
    )


APP_TIMEZONE = os.getenv("APP_TIMEZONE", "Europe/Madrid")
PROACTIVE_SCHEDULER_HOUR = int(os.getenv("PROACTIVE_SCHEDULER_HOUR", "10"))
PROACTIVE_SCHEDULER_MINUTE = int(os.getenv("PROACTIVE_SCHEDULER_MINUTE", "40"))

gemini_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

deepseek_client = (
    OpenAI(
        api_key=DEEPSEEK_API_KEY,
        base_url=DEEPSEEK_BASE_URL,
    )
    if DEEPSEEK_API_KEY
    else None
)

MAIN_KEYBOARD = ReplyKeyboardRemove()

MEMORY_CONFIRM_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["✅ Confirmar cambio memoria", "❌ Cancelar cambio memoria"],
        ["👤 Quién soy", "🎯 Qué quiero"],
        ["✅ Plan de acción", "⏰ Seguimientos"],
    ],
    resize_keyboard=True,
    one_time_keyboard=False,
)

PENDING_MEMORY_EDIT: dict[str, dict] = {}
PENDING_MEMORY_PATCH: dict[str, dict] = {}

SESSION_CONSOLIDATION_CONFIRM_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["✅ Confirmar guardado sesión", "❌ Cancelar guardado sesión"],
        ["👤 Quién soy", "🎯 Qué quiero"],
        ["✅ Plan de acción", "⏰ Seguimientos"],
        ["💾 Guardar sesión"],
    ],
    resize_keyboard=True,
    one_time_keyboard=False,
)

PENDING_SESSION_CONSOLIDATION: dict[str, dict] = {}

# ---------------------------------------------------------------------
# Initial client files
# ---------------------------------------------------------------------

def ensure_client_files() -> None:
    """Create the active client folder and starter Markdown files if needed."""
    active_client_dir().mkdir(parents=True, exist_ok=True)

    for key, path in client_files().items():
        if path.exists():
            continue

        template_path = INITIAL_CLIENT_FILES_DIR / f"{key}.md"

        if template_path.exists():
            content = template_path.read_text(encoding="utf-8").strip()
        else:
            content = f"# {key}\n\nPendiente."

        path.write_text(content + "\n", encoding="utf-8")

    # eCoach_Relaciones v3: create project/board files beside the legacy memory files.
    try:
        ensure_ecoach_project_files(active_client_dir())
    except NameError:
        # During early imports in unusual tooling, the v3 helpers may not yet be defined.
        pass

# ---------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------

def now_app() -> datetime:
    """Current datetime in the app timezone."""
    return datetime.now(ZoneInfo(APP_TIMEZONE))


def today_app() -> date:
    """Current date in the app timezone."""
    return now_app().date()


def format_app_datetime() -> str:
    """Standard timestamp in the app timezone."""
    return now_app().strftime("%Y-%m-%d %H:%M:%S")


def format_app_minute() -> str:
    """Standard minute-level timestamp in the app timezone."""
    return now_app().strftime("%Y-%m-%d %H:%M")


def read_text_file(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8").strip()


def write_text_file(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def append_text_file(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(text.rstrip() + "\n")


def compact_log_preview(text: str, max_chars: int = 1200) -> str:
    """Return a short preview for large or malformed LLM outputs in logs."""
    value = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    if len(value) <= max_chars:
        return value
    omitted = len(value) - max_chars
    return value[:max_chars] + f"\n\n[TRUNCATED: {omitted} chars omitted]"

def load_app_state() -> dict:
    """Load small persistent app state, such as the Telegram chat ID."""
    path = app_state_file()

    if not path.exists():
        return {}

    try:
        content = path.read_text(encoding="utf-8").strip()

        if not content:
            return {}

        data = json.loads(content)

        if isinstance(data, dict):
            return data

        return {}

    except json.JSONDecodeError:
        return {}


def save_app_state(state: dict) -> None:
    """Save small persistent app state."""
    path = app_state_file()
    path.parent.mkdir(parents=True, exist_ok=True)

    path.write_text(
        json.dumps(state, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def remember_chat_id(update: Update) -> None:
    """Remember the current Telegram chat ID so scheduled jobs can send proactive messages."""
    if update.effective_chat is None:
        return

    state = load_app_state()
    state["telegram_chat_id"] = update.effective_chat.id
    state["telegram_chat_id_updated_at"] = format_app_datetime()
    save_app_state(state)


def get_remembered_chat_id() -> int | None:
    """Return the remembered Telegram chat ID, if available."""
    state = load_app_state()

    chat_id = state.get("telegram_chat_id")

    if chat_id is None:
        return None

    try:
        return int(chat_id)
    except (TypeError, ValueError):
        return None
    
def backup_client_file(path: Path) -> None:
    """
    Create a timestamped backup of a client Markdown file before overwriting it.
    """
    if not path.exists():
        return

    timestamp = now_app().strftime("%Y%m%d_%H%M%S")

    backups_dir = path.parent / "backups"
    backups_dir.mkdir(parents=True, exist_ok=True)

    backup_path = backups_dir / f"{path.stem}_{timestamp}{path.suffix}"

    backup_path.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")

def write_client_file_with_backup(path: Path, text: str) -> None:
    """
    Backup a client Markdown file, then overwrite it.
    """
    backup_client_file(path)
    write_text_file(path, text)

def ensure_followup_triggers_file() -> None:
    """Create followup_triggers.json if it does not exist."""
    path = followup_triggers_file()
    path.parent.mkdir(parents=True, exist_ok=True)

    if not path.exists():
        path.write_text("[]\n", encoding="utf-8")


def get_pending_followup_triggers() -> list[dict]:
    """Return pending seguimiento triggers."""
    triggers = load_followup_triggers()

    return [
        trigger for trigger in triggers
        if str(trigger.get("status", "pending")).lower() == "pending"
    ]


def format_pending_followups_for_prompt() -> str:
    """Return a compact summary of pending seguimientos for the orchestrator prompt."""
    pending = get_pending_followup_triggers()

    if not pending:
        return "No hay seguimientos pendientes."

    lines: list[str] = []

    for index, trigger in enumerate(pending, start=1):
        lines.append(
            "\n".join(
                [
                    f"{index}.",
                    f"- id: {trigger.get('id', '')}",
                    f"- date: {trigger.get('date', '')}",
                    f"- time: {trigger.get('time', '')}",
                    f"- type: {trigger.get('type', '')}",
                    f"- message_template: {trigger.get('message_template', '')}",
                    f"- reason: {trigger.get('reason', '')}",
                    f"- status: {trigger.get('status', '')}",
                ]
            )
        )

    return "\n\n".join(lines)


def parse_followup_date(date_text: str) -> date | None:
    """Parse a seguimiento date in YYYY-MM-DD format."""
    if not date_text:
        return None

    try:
        return datetime.strptime(date_text.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_followup_time(time_text: str) -> time | None:
    """Parse a seguimiento time in HH:MM format."""
    if not time_text:
        return None

    try:
        return datetime.strptime(time_text.strip(), "%H:%M").time()
    except ValueError:
        return None


def followup_has_time(trigger: dict) -> bool:
    return parse_followup_time(str(trigger.get("time", "") or "")) is not None


def followup_due_datetime(trigger: dict) -> datetime | None:
    trigger_date = parse_followup_date(str(trigger.get("date", "") or ""))
    trigger_time = parse_followup_time(str(trigger.get("time", "") or ""))

    if trigger_date is None or trigger_time is None:
        return None

    return datetime.combine(
        trigger_date,
        trigger_time,
        tzinfo=ZoneInfo(APP_TIMEZONE),
    )


def parse_snooze_date(date_text: str) -> date | None:
    """Parse a snooze date using the fast LLM router, not hardcoded language rules."""
    system_prompt = load_required_app_text("snooze_date_parser_system_prompt.md")

    prompt = f"""
Fecha actual en la zona horaria de la app: {today_app().strftime('%Y-%m-%d')}
Zona horaria: {APP_TIMEZONE}

Instrucción del usuario para posponer un seguimiento:
{date_text}
""".strip()

    try:
        raw_answer, usage = llm_generate(
            prompt=prompt,
            system_prompt=system_prompt,
            provider_order=ROUTER_PROVIDER_ORDER,
            deepseek_model=ROUTER_DEEPSEEK_MODEL,
            gemini_model=ROUTER_GEMINI_MODEL,
        )

        parsed = extract_json_from_text(raw_answer)
        date_value = parsed.get("date")

        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### LLM snooze date parse",
                    "",
                    f"Input: {date_text}",
                    f"Parsed date: {date_value}",
                    f"Confidence: {parsed.get('confidence', '')}",
                    f"Provider: {usage.get('provider', 'unknown')}",
                    f"Model: {usage.get('model', 'unknown')}",
                    f"Tokens: {usage.get('total_tokens', 0)}",
                    f"Cost USD: {usage.get('estimated_cost_usd', 0):.6f}",
                    f"Reason: {parsed.get('reason', '')}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        if date_value is None or str(date_value).strip().lower() in {"", "null", "none"}:
            return None

        return parse_followup_date(str(date_value))

    except Exception as error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### LLM snooze date parse failed",
                    "",
                    f"Input: {date_text}",
                    f"Error: {str(error)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )
        return None


def validate_followup_type(raw_type: str) -> str:
    """Validate the LLM-generated seguimiento type against the internal enum.

    This function does not interpret natural language. The LLM must choose
    the semantic type. Python only accepts a known schema value or falls back
    safely.
    """
    allowed_types = {
        "document_expected",
        "document_review",
        "client_action_reminder",
        "decision_deadline",
        "meeting_preparation",
        "advisor_followup",
        "general_checkin",
    }

    raw_type_clean = raw_type.strip().lower()

    if raw_type_clean in allowed_types:
        return raw_type_clean

    return "general_checkin"


def build_followup_duplicate_prompt(new_trigger: dict, candidate_triggers: list[dict]) -> tuple[str, str]:
    """Build a tiny LLM prompt to decide whether a seguimiento is duplicate."""
    system_prompt = load_required_app_text("followup_duplicate_detector_system_prompt.md")

    prompt = "\n".join(
        [
            "Seguimiento nuevo:",
            json.dumps(new_trigger, ensure_ascii=False, indent=2),
            "",
            "Seguimientos existentes candidatos:",
            json.dumps(candidate_triggers, ensure_ascii=False, indent=2),
        ]
    )

    return system_prompt, prompt


def llm_bool(value, default: bool = False) -> bool:
    """Accept only real JSON booleans. Do not interpret language strings."""
    if isinstance(value, bool):
        return value

    return default


def is_duplicate_followup_with_llm(new_trigger: dict, existing_triggers: list[dict]) -> bool:
    """Use a small LLM call to decide whether a seguimiento is semantically duplicate.

    Structural pre-filtering is allowed: only pending seguimientos with the same
    ISO date are candidates. Semantic comparison is done only by the LLM.
    """
    new_date = str(new_trigger.get("date", "")).strip()

    candidates = []

    for existing in existing_triggers:
        existing_status = str(existing.get("status", "pending")).lower()

        if existing_status != "pending":
            continue

        existing_date = str(existing.get("date", "")).strip()

        if new_date and existing_date and existing_date != new_date:
            continue

        candidates.append(
            {
                "id": existing.get("id", ""),
                "date": existing.get("date", ""),
                "type": existing.get("type", ""),
                "message_template": existing.get("message_template", ""),
                "reason": existing.get("reason", ""),
            }
        )

    if not candidates:
        return False

    try:
        system_prompt, prompt = build_followup_duplicate_prompt(
            new_trigger={
                "id": new_trigger.get("id", ""),
                "date": new_trigger.get("date", ""),
                "type": new_trigger.get("type", ""),
                "message_template": new_trigger.get("message_template", ""),
                "reason": new_trigger.get("reason", ""),
            },
            candidate_triggers=candidates,
        )

        raw_answer, _usage = llm_generate(
            prompt=prompt,
            system_prompt=system_prompt,
            provider_order=ROUTER_PROVIDER_ORDER,
            deepseek_model=ROUTER_DEEPSEEK_MODEL,
            gemini_model=ROUTER_GEMINI_MODEL,
        )

        decision = extract_json_from_text(raw_answer)

        confidence = str(decision.get("confidence", "low")).strip().lower()

        if confidence not in {"medium", "high"}:
            return False

        return llm_bool(decision.get("is_duplicate", False), default=False)

    except Exception as error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### LLM duplicate seguimiento check failed",
                    "",
                    f"Error: {str(error)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        return False


def load_followup_triggers() -> list[dict]:
    """Load proactive seguimiento triggers from JSON."""
    ensure_followup_triggers_file()

    try:
        content = followup_triggers_file().read_text(encoding="utf-8").strip()

        if not content:
            return []

        data = json.loads(content)

        if isinstance(data, list):
            return data

        return []

    except json.JSONDecodeError:
        return []


def save_followup_triggers(triggers: list[dict]) -> None:
    """Save proactive seguimiento triggers to JSON."""
    path = followup_triggers_file()
    path.parent.mkdir(parents=True, exist_ok=True)

    path.write_text(
        json.dumps(triggers, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def load_followup_archive() -> list[dict]:
    """Load archived seguimientos from JSON."""
    path = followup_archive_file()
    path.parent.mkdir(parents=True, exist_ok=True)

    if not path.exists():
        path.write_text("[]\n", encoding="utf-8")
        return []

    try:
        content = path.read_text(encoding="utf-8").strip()
        if not content:
            return []
        data = json.loads(content)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def save_followup_archive(items: list[dict]) -> None:
    """Save archived seguimientos to JSON."""
    path = followup_archive_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(items, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def archive_followup_trigger(trigger: dict, archive_reason: str) -> None:
    """Append a seguimiento to the archive with metadata."""
    archive = load_followup_archive()

    archived_trigger = dict(trigger)
    archived_trigger["archived_at"] = format_app_datetime()
    archived_trigger["archive_reason"] = archive_reason

    archive.append(archived_trigger)
    save_followup_archive(archive)
    

def append_agent_observation(
    source: str,
    user_input: str,
    observations: dict[str, str],
) -> None:
    """Append internal agent observations to agent_observations.md."""
    ensure_client_files()

    timestamp = format_app_minute()

    sections: list[str] = []

    for agent_name, observation in observations.items():
        if not observation:
            continue

        clean_agent_name = agent_name.replace("_", " ").title()

        sections.append(
            f"### {clean_agent_name}\n\n{observation.strip()}"
        )

    if not sections:
        return

    template = load_required_app_text("agent_observation_entry.md")

    entry = render_template(
        template,
        {
            "timestamp": timestamp,
            "source": source,
            "user_input": user_input,
            "agent_sections": "\n\n".join(sections),
        },
    )

    append_text_file(client_files()["agent_observations"], entry)

def available_wealth_knowledge_files() -> list[str]:
    """Return Markdown filenames available for semantic routing."""
    if not WEALTH_KNOWLEDGE_DIR.exists():
        return []

    return sorted(path.name for path in WEALTH_KNOWLEDGE_DIR.glob("*.md"))


def load_wealth_knowledge() -> str:
    """Load all Markdown files from Wealth_Knowledge."""
    return load_selected_wealth_knowledge(selected_files=None)


def load_selected_wealth_knowledge(selected_files: list[str] | None = None) -> str:
    """Load selected Markdown files from Wealth_Knowledge.

    selected_files=None means full legacy behavior: load everything.
    selected_files=[] means intentionally load no wealth knowledge.
    Unknown filenames are ignored safely.
    """
    if not WEALTH_KNOWLEDGE_DIR.exists():
        return ""

    if selected_files is None:
        paths = sorted(WEALTH_KNOWLEDGE_DIR.glob("*.md"))
    else:
        allowed = set(available_wealth_knowledge_files())
        paths = [
            WEALTH_KNOWLEDGE_DIR / filename
            for filename in selected_files
            if filename in allowed
        ]

    parts: list[str] = []

    for path in paths:
        content = path.read_text(encoding="utf-8").strip()
        if content:
            parts.append(f"# Archivo fuente: {path.name}\n\n{content}")

    return "\n\n---\n\n".join(parts)

def load_agent_config_files() -> str:
    """Load global agent configuration Markdown files."""
    if not AGENT_CONFIG_DIR.exists():
        return ""

    parts: list[str] = []

    for path in sorted(AGENT_CONFIG_DIR.glob("*.md")):
        content = path.read_text(encoding="utf-8").strip()
        if content:
            parts.append(f"# Agent config file: {path.name}\n\n{content}")

    return "\n\n---\n\n".join(parts)

def load_client_agent_config() -> str:
    """Load the active client's agent behavior configuration."""
    ensure_client_files()
    return read_text_file(client_files()["agent_config"])

def load_client_context() -> dict[str, str]:
    """Load all live Markdown files for the active client."""
    ensure_client_files()
    return {key: read_text_file(path) for key, path in client_files().items()}


def load_selected_client_context(selected_keys: list[str] | None = None) -> dict[str, str]:
    """Load only selected live Markdown files for the active client.

    Missing non-selected keys are returned as empty strings so old templates still work.
    """
    ensure_client_files()
    files = client_files()
    selected_keys = filter_normal_orchestrator_client_keys(selected_keys)

    if selected_keys is None:
        selected_keys = normal_orchestrator_client_keys()

    selected = set(selected_keys)
    context: dict[str, str] = {}

    for key, path in files.items():
        if key in selected:
            context[key] = read_text_file(path)
        else:
            context[key] = ""

    return context

def load_app_text(name: str) -> str:
    """Load an app text template from AppTexts."""
    path = APP_TEXTS_DIR / name

    if not path.exists():
        return ""

    return path.read_text(encoding="utf-8").strip()

def load_required_app_text(name: str) -> str:
    """Load an app text template and fail clearly if it is missing."""
    text = load_app_text(name)

    if text:
        return text

    fallback = load_app_text("missing_app_text_error.md")

    if fallback:
        message = render_template(
            fallback,
            {
                "filename": str(APP_TEXTS_DIR / name),
            },
        )
    else:
        message = f"Error de configuración: falta {APP_TEXTS_DIR / name}"

    raise RuntimeError(message)

def render_template(template: str, values: dict[str, str]) -> str:
    """Very simple {{placeholder}} replacement."""
    rendered = template

    for key, value in values.items():
        rendered = rendered.replace("{{" + key + "}}", value)

    return rendered

def save_wealth_log(user_input: str, ai_answer: str, usage: dict) -> None:
    """Save a trace of a model-generated answer."""
    timestamp = format_app_minute()

    template = load_required_app_text("wealth_log_entry.md")

    entry = render_template(
        template,
        {
            "timestamp": timestamp,
            "active_client_name": current_client_name(),
            "user_input": user_input,
            "model": str(usage.get("model", GEMINI_MODEL)),
            "input_tokens": str(usage.get("input_tokens", 0)),
            "output_tokens": str(usage.get("output_tokens", 0)),
            "total_tokens": str(usage.get("total_tokens", 0)),
            "estimated_cost_usd": f"{usage.get('estimated_cost_usd', 0):.6f}",
            "ai_answer": ai_answer,
        },
    )

    append_text_file(wealth_log_file(), entry)

def process_with_orchestrator(
    text_to_process: str,
    source: str,
    selected_client_keys: list[str] | None = None,
    selected_knowledge_files: list[str] | None = None,
    routing_decision: dict | None = None,
) -> str:
    """
    Main LLM orchestrator.

    It updates:
    - quien_soy.md
    - que_quiero.md
    - que_tengo_que_hacer.md

    It also saves:
    - agent observations
    - seguimiento triggers

    And returns the client-facing answer.
    """
    ensure_client_files()

    client_context = load_selected_client_context(selected_client_keys)
    wealth_knowledge = load_selected_wealth_knowledge(selected_knowledge_files)
    global_agent_config = load_agent_config_files()
    client_agent_config = load_client_agent_config()

    pending_followups_context = format_pending_followups_for_prompt()
    workflow_board_context = ecoach_board_summary_text(active_client_dir())

    current_date = today_app().strftime("%Y-%m-%d")

    system_prompt = load_required_app_text("orchestrator_system_prompt.md")
    prompt_template = load_required_app_text("orchestrator_user_prompt.md")

    prompt = render_template(
        prompt_template,
        {
            "text_to_process": text_to_process,
            "source": source,
            "current_date": current_date,
            "quien_soy": client_context["quien_soy"],
            "que_quiero": client_context["que_quiero"],
            "que_tengo_que_hacer": client_context["que_tengo_que_hacer"],
            "estilo_respuesta": client_context["estilo_respuesta"],
            "agent_config": client_agent_config,
            "global_agent_config": global_agent_config,
            "wealth_knowledge": wealth_knowledge,
            "pending_followups": pending_followups_context,
            "workflow_board": workflow_board_context,
        },
    )

    try:
        raw_answer, usage = llm_generate_for_route(
            prompt=prompt,
            system_prompt=system_prompt,
            routing_decision=routing_decision,
        )

    except errors.ServerError as error:
        timestamp = format_app_minute()

        log_entry = render_template(
            load_required_app_text("orchestrator_server_error_log.md"),
            {
                "timestamp": timestamp,
                "error": str(error),
                "text_to_process": text_to_process,
            },
        )

        append_text_file(wealth_log_file(), log_entry)

        return load_required_app_text("orchestrator_server_error_reply.md")

    except Exception as error:
        timestamp = format_app_minute()

        log_entry = render_template(
            load_required_app_text("orchestrator_unexpected_error_log.md"),
            {
                "timestamp": timestamp,
                "error": str(error),
                "text_to_process": text_to_process,
            },
        )

        append_text_file(wealth_log_file(), log_entry)

        return load_required_app_text("orchestrator_unexpected_error_reply.md")

    try:
        parsed = extract_json_from_text(raw_answer)

    except Exception as first_error:
        timestamp = format_app_minute()

        warning_log_entry = render_template(
            load_required_app_text("orchestrator_json_parse_warning_log.md"),
            {
                "timestamp": timestamp,
                "first_error": str(first_error),
                "raw_answer": compact_log_preview(raw_answer),
            },
        )

        append_text_file(wealth_log_file(), warning_log_entry)

        try:
            parsed = repair_json_with_llm(raw_answer)

            success_log_entry = render_template(
                load_required_app_text("orchestrator_json_repair_success_log.md"),
                {
                    "timestamp": timestamp,
                },
            )

            append_text_file(wealth_log_file(), success_log_entry)

        except Exception as repair_error:
            failure_log_entry = render_template(
                load_required_app_text("orchestrator_json_failure_log.md"),
                {
                    "timestamp": timestamp,
                    "first_error": str(first_error),
                    "repair_error": str(repair_error),
                    "raw_answer": compact_log_preview(raw_answer),
                },
            )

            append_text_file(wealth_log_file(), failure_log_entry)

            return load_required_app_text("orchestrator_json_failure_reply.md")

    required_keys = [
        "quien_soy",
        "que_quiero",
        "que_tengo_que_hacer",
        "respuesta_cliente",
        "proxima_accion",
        "agent_observations",
        "followup_triggers",
        "notas_auditoria",
    ]

    missing_keys = [key for key in required_keys if key not in parsed]

    if missing_keys:
        missing_keys_text = "\n".join(f"- {key}" for key in missing_keys)

        return render_template(
            load_required_app_text("orchestrator_missing_keys_reply.md"),
            {
                "missing_keys": missing_keys_text,
            },
        )

    write_client_file_with_backup(client_files()["quien_soy"], parsed["quien_soy"])
    write_client_file_with_backup(client_files()["que_quiero"], parsed["que_quiero"])
    write_client_file_with_backup(
        client_files()["que_tengo_que_hacer"],
        parsed["que_tengo_que_hacer"],
    )

    agent_observations = parsed.get("agent_observations", {})

    if isinstance(agent_observations, dict):
        append_agent_observation(
            source=source,
            user_input=text_to_process,
            observations=agent_observations,
        )

    new_followup_triggers = parsed.get("followup_triggers", [])

    if isinstance(new_followup_triggers, list) and new_followup_triggers:
        existing_triggers = load_followup_triggers()

        timestamp = format_app_datetime()
        id_timestamp = now_app().strftime("%Y%m%d_%H%M%S")

        normalized_triggers: list[dict] = []

        # Seguimiento dates are interpreted by the LLM orchestrator.
        # Do not override natural language dates with hardcoded local rules.
        date_override = None
        
        for index, trigger in enumerate(new_followup_triggers, start=1):
            if not isinstance(trigger, dict):
                continue

            normalized_trigger = {
                "id": f"followup_{id_timestamp}_{index}",
                "client_id": current_client_name(),
                "created_at": str(trigger.get("created_at") or timestamp),
                "date": str(date_override or trigger.get("date") or ""),
                "time": str(trigger.get("time") or ""),
                "type": validate_followup_type(str(trigger.get("type") or "general_checkin")),
                "message_template": str(trigger.get("message_template") or ""),
                "reason": str(trigger.get("reason") or ""),
                "sensitivity": str(trigger.get("sensitivity") or "low"),
                "requires_private_context": bool(trigger.get("requires_private_context", True)),
                "status": str(trigger.get("status") or "pending").lower(),
                "source": source,
            }

            if normalized_trigger["message_template"]:
                if is_duplicate_followup_with_llm(normalized_trigger, existing_triggers + normalized_triggers):
                    append_proactivity_log(
                        event_type="followup_duplicate_skipped",
                        trigger=normalized_trigger,
                        notes="Skipped because a similar pending seguimiento already exists.",
                    )
                    continue
                
                normalized_triggers.append(normalized_trigger)

        if normalized_triggers:
            save_followup_triggers(existing_triggers + normalized_triggers)

            for trigger in normalized_triggers:
                append_proactivity_log(
                    event_type="followup_created",
                    trigger=trigger,
                    notes="Created by LLM orchestrator from client interaction.",
                )

    timestamp = format_app_minute()

    historial_template = load_required_app_text("historial_interaction_entry.md")

    historial_entry = render_template(
        historial_template,
        {
            "timestamp": timestamp,
            "source": source,
            "text_to_process": text_to_process,
            "respuesta_cliente": parsed["respuesta_cliente"],
            "proxima_accion": parsed["proxima_accion"],
            "notas_auditoria": parsed["notas_auditoria"],
            "model": str(usage.get("model", GEMINI_MODEL)),
            "input_tokens": str(usage.get("input_tokens", 0)),
            "output_tokens": str(usage.get("output_tokens", 0)),
            "total_tokens": str(usage.get("total_tokens", 0)),
            "estimated_cost_usd": f"{usage.get('estimated_cost_usd', 0):.6f}",
        },
    )

    append_text_file(
        client_files()["historial_interacciones"],
        historial_entry,
    )

    save_wealth_log(
        user_input=text_to_process,
        ai_answer=parsed["respuesta_cliente"],
        usage=usage,
    )

    final_reply_template = load_required_app_text("orchestrator_final_reply.md")

    return render_template(
        final_reply_template,
        {
            "respuesta_cliente": parsed["respuesta_cliente"],
            "proxima_accion": parsed["proxima_accion"],
            "gemini_usage_note": format_usage_note(usage),
        },
    )

# ---------------------------------------------------------------------
# LLM helpers
# ---------------------------------------------------------------------

def gemini_llm(
    prompt: str,
    system_prompt: str | None = None,
    model: str | None = None,
) -> tuple[str, dict]:
    """Send a prompt to Gemini and return answer text plus usage metadata."""
    if gemini_client is None:
        raise RuntimeError(load_required_app_text("gemini_not_configured_error.md"))

    config = None

    if system_prompt:
        config = types.GenerateContentConfig(system_instruction=system_prompt)

    selected_model = model or GEMINI_MODEL

    response = gemini_client.models.generate_content(
        model=selected_model,
        contents=prompt,
        config=config,
    )

    usage = getattr(response, "usage_metadata", None)

    input_tokens = getattr(usage, "prompt_token_count", 0) if usage else 0
    output_tokens = getattr(usage, "candidates_token_count", 0) if usage else 0
    total_tokens = getattr(usage, "total_token_count", 0) if usage else 0

    estimated_cost_usd = (
        (input_tokens / 1_000_000) * GEMINI_INPUT_PRICE_PER_1M
        + (output_tokens / 1_000_000) * GEMINI_OUTPUT_PRICE_PER_1M
    )

    metadata = {
        "model": selected_model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "estimated_cost_usd": estimated_cost_usd,
    }

    return response.text or "", metadata

def deepseek_llm(
    prompt: str,
    system_prompt: str | None = None,
    model: str | None = None,
) -> tuple[str, dict]:
    """Send a prompt to DeepSeek and return answer text plus usage metadata."""
    if deepseek_client is None:
        raise RuntimeError("DeepSeek API key is not configured.")

    messages = []

    if system_prompt:
        messages.append(
            {
                "role": "system",
                "content": system_prompt,
            }
        )

    messages.append(
        {
            "role": "user",
            "content": prompt,
        }
    )

    selected_model = model or DEEPSEEK_MODEL

    response = deepseek_client.chat.completions.create(
        model=selected_model,
        messages=messages,
        temperature=0,
        max_tokens=LLM_MAX_OUTPUT_TOKENS,
    )

    answer = response.choices[0].message.content or ""

    usage = getattr(response, "usage", None)

    input_tokens = getattr(usage, "prompt_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "completion_tokens", 0) if usage else 0
    total_tokens = getattr(usage, "total_tokens", input_tokens + output_tokens) if usage else 0

    estimated_cost_usd = (
        (input_tokens / 1_000_000) * DEEPSEEK_INPUT_PRICE_PER_1M
        + (output_tokens / 1_000_000) * DEEPSEEK_OUTPUT_PRICE_PER_1M
    )

    metadata = {
        "provider": "deepseek",
        "model": selected_model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "estimated_cost_usd": estimated_cost_usd,
    }

    return answer, metadata


def llm_generate(
    prompt: str,
    system_prompt: str | None = None,
    provider_order: list[str] | None = None,
    deepseek_model: str | None = None,
    gemini_model: str | None = None,
) -> tuple[str, dict]:
    """
    Generate with the configured provider order.

    Default:
    LLM_PROVIDER_ORDER=deepseek,gemini
    """
    last_error: Exception | None = None
    providers = provider_order or LLM_PROVIDER_ORDER

    for provider in providers:
        for attempt in range(LLM_PROVIDER_RETRIES + 1):
            try:
                if provider == "deepseek":
                    return deepseek_llm(
                        prompt=prompt,
                        system_prompt=system_prompt,
                        model=deepseek_model,
                    )

                if provider == "gemini":
                    raw_answer, usage = gemini_llm(
                        prompt=prompt,
                        system_prompt=system_prompt,
                        model=gemini_model,
                    )
                    usage["provider"] = "gemini"
                    return raw_answer, usage

                raise RuntimeError(f"Unknown LLM provider: {provider}")

            except Exception as error:
                last_error = error

                will_retry = attempt < LLM_PROVIDER_RETRIES

                append_text_file(
                    wealth_log_file(),
                    "\n".join(
                        [
                            f"## {format_app_minute()}",
                            "",
                            "### LLM provider failed",
                            "",
                            f"Provider: {provider}",
                            f"Attempt: {attempt + 1}/{LLM_PROVIDER_RETRIES + 1}",
                            f"Will retry same provider: {will_retry}",
                            "",
                            f"Error: {str(error)}",
                            "",
                            "---",
                            "",
                        ]
                    ),
                )

                if will_retry:
                    time.sleep(LLM_PROVIDER_RETRY_SLEEP_SECONDS)
                    continue

                break

    raise RuntimeError(f"All LLM providers failed. Last error: {last_error}")


def llm_generate_for_route(
    prompt: str,
    system_prompt: str | None,
    routing_decision: dict | None,
) -> tuple[str, dict]:
    """Generate the main orchestrator answer using configured detailed-call defaults.

    The router should decide semantic depth and context, not concrete provider/model names.
    Concrete AI providers/models belong in configuration, not in routing logic.
    """
    return llm_generate(prompt=prompt, system_prompt=system_prompt)


def extract_json_from_text(text: str) -> dict:
    """
    Extract JSON from an LLM response.

    The model may return pure JSON, or JSON wrapped in ```json ... ```.
    This helper makes parsing a bit more robust.
    """
    cleaned = text.strip()

    if cleaned.startswith("```json"):
        cleaned = cleaned.removeprefix("```json").strip()

    if cleaned.startswith("```"):
        cleaned = cleaned.removeprefix("```").strip()

    if cleaned.endswith("```"):
        cleaned = cleaned.removesuffix("```").strip()

    return json.loads(cleaned)

def repair_json_with_llm(raw_text: str) -> dict:
    """
    Ask the configured LLM provider to repair an invalid JSON-like response into valid JSON.

    This is used only when the first JSON parse fails.
    """
    repair_system_prompt = load_required_app_text("json_repair_system_prompt.md")
    repair_prompt_template = load_required_app_text("json_repair_user_prompt.md")

    repair_prompt = render_template(
        repair_prompt_template,
        {
            "raw_text": raw_text,
        },
    )

    repaired_answer, _usage = llm_generate(
        prompt=repair_prompt,
        system_prompt=repair_system_prompt,
    )

    return extract_json_from_text(repaired_answer)

def format_usage_note(usage: dict) -> str:
    """Small cost note appended to Telegram answers during development."""
    if not SHOW_USAGE_NOTE:
        return ""

    template = load_required_app_text("usage_note.md")

    return render_template(
        template,
        {
            "provider": str(usage.get("provider", "unknown")),
            "model": str(usage.get("model", GEMINI_MODEL)),
            "input_tokens": str(usage.get("input_tokens", 0)),
            "output_tokens": str(usage.get("output_tokens", 0)),
            "total_tokens": str(usage.get("total_tokens", 0)),
            "estimated_cost_usd": f"{usage.get('estimated_cost_usd', 0):.6f}",
        },
    )

def build_router_prompt(user_text: str) -> tuple[str, str]:
    """Build the tiny semantic routing prompt."""
    available_knowledge = available_wealth_knowledge_files()
    knowledge_list = "\n".join(f"- {name}" for name in available_knowledge) or "- No hay archivos Wealth_Knowledge disponibles."
    client_files_list = "\n".join(f"- {name}" for name in normal_orchestrator_client_keys())

    system_prompt = load_required_app_text("router_system_prompt.md")

    prompt = f"""
Fecha actual:
{now_app().date().isoformat()}

Mensaje del usuario:
{user_text}

Archivos de cliente disponibles:
{client_files_list}

Archivos de conocimiento patrimonial disponibles:
{knowledge_list}

Criterios semánticos:
- Decide por significado, no por palabras exactas. Tolera faltas, conjugaciones, lenguaje incompleto y mezcla de idiomas.
- Para saludos simples, agradecimientos, confirmaciones breves o preguntas de ayuda general, NO uses orquestador.
- En esos casos simples debes devolver exactamente: needs_orchestrator=false, confidence="high", depth="none", client_files=[], knowledge_files=[], should_update_memory=false, should_extract_followups=false y direct_response con una respuesta breve y útil en español.
- Para saludos simples usa route="smalltalk". Para agradecimientos usa route="thanks_ack". Para ayuda general usa route="help".
- Si el mensaje solo requiere una respuesta social o explicativa breve, no uses orquestador.
- Si el mensaje contiene información nueva del cliente, pide actualización de memoria.
- Si hay una intención de recordar, posponer, preparar, esperar documentos o actuar en el futuro, pide extracción de seguimientos.
- Si hay una decisión patrimonial, financiera, fiscal, bancaria, familiar o documental, selecciona únicamente los archivos de conocimiento realmente relevantes.
- Si el caso parece complejo, sensible o de alto impacto, usa profundidad deep.
- Si no estás seguro, usa confidence="low" y needs_orchestrator=true para que el sistema cargue más contexto.
""".strip()

    return system_prompt, prompt


def parse_router_bool(value, default: bool = False) -> bool:
    """Accept only real JSON booleans from the router.

    Do not interpret natural-language strings such as yes/no/sí/no.
    The router prompt requires JSON booleans; malformed values fall back safely.
    """
    if isinstance(value, bool):
        return value

    return default


def normalize_router_decision(raw_decision: dict) -> dict:
    """Normalize and validate the router decision with safe fallbacks."""
    available_client_keys = set(client_files().keys())
    available_knowledge = set(available_wealth_knowledge_files())

    raw_route = str(raw_decision.get("route", "unknown")).strip().lower()
    route = raw_route
    if route not in ROUTER_ALLOWED_ROUTES:
        route = "unknown"

    confidence = str(raw_decision.get("confidence", "low")).strip().lower()
    if confidence not in ROUTER_ALLOWED_CONFIDENCE:
        confidence = "low"

    depth = str(raw_decision.get("depth", "normal")).strip().lower()
    if depth not in ROUTER_ALLOWED_DEPTHS:
        depth = "normal"

    needs_orchestrator = parse_router_bool(raw_decision.get("needs_orchestrator", True), default=True)

    client_files_raw = raw_decision.get("client_files", DEFAULT_ROUTER_CLIENT_FILES)
    if not isinstance(client_files_raw, list):
        client_files_raw = DEFAULT_ROUTER_CLIENT_FILES

    selected_client_files = [
        str(name)
        for name in client_files_raw
        if str(name) in available_client_keys
    ]

    selected_client_files = filter_normal_orchestrator_client_keys(selected_client_files)

    if needs_orchestrator and not selected_client_files:
        selected_client_files = normal_orchestrator_client_keys()

    knowledge_files_raw = raw_decision.get("knowledge_files", [])
    if not isinstance(knowledge_files_raw, list):
        knowledge_files_raw = []

    selected_knowledge_files = [
        str(name)
        for name in knowledge_files_raw
        if str(name) in available_knowledge
    ]

    direct_response = str(raw_decision.get("direct_response", "") or "").strip()

    # Safety: allow direct replies when the LLM router explicitly says no
    # orchestrator is needed, confidence is high, and it provides the response.
    # Do not infer meaning from the route label; the route is only metadata.
    if not needs_orchestrator:
        if confidence != "high" or not direct_response:
            # No hardcoded language fallback. If the router cannot provide
            # the direct response confidently, let the orchestrator answer.
            needs_orchestrator = True
            route = "unknown"
            depth = "normal"

    # Safety: if the router says low confidence, do not under-load.
    if needs_orchestrator and confidence == "low":
        selected_client_files = normal_orchestrator_client_keys()
        selected_knowledge_files = None  # None = full legacy Wealth_Knowledge load.
        depth = "normal"

    answer_model = str(raw_decision.get("answer_model", "default") or "default").strip()

    return {
        "route": route,
        "raw_route": raw_route,
        "confidence": confidence,
        "needs_orchestrator": needs_orchestrator,
        "answer_model": answer_model,
        "depth": depth,
        "client_files": selected_client_files,
        "knowledge_files": selected_knowledge_files,
        "should_update_memory": parse_router_bool(
            raw_decision.get("should_update_memory", needs_orchestrator),
            default=needs_orchestrator,
        ),
        "should_extract_followups": parse_router_bool(
            raw_decision.get("should_extract_followups", needs_orchestrator),
            default=needs_orchestrator,
        ),
        "user_visible_delay_hint": parse_router_bool(
            raw_decision.get("user_visible_delay_hint", False),
            default=False,
        ),
        "direct_response": direct_response,
        "reason": str(raw_decision.get("reason", "") or "").strip(),
        "followup_triggers": raw_decision.get("followup_triggers", []),
    }



# --- DETERMINISTIC MEMORY GUARD v2 ---
# Stable user facts must enter session-memory even if the LLM router
# classifies the message as simple_response.
def looks_like_stable_memory_fact(user_text: str) -> bool:
    t = (user_text or "").strip().lower()
    if not t:
        return False

    patterns = [
        # Personal profile
        r"\btengo\s+\d{1,3}\s+años\b",
        r"\btengo\s+\d{1,3}\s+anos\b",
        r"\bvivo\s+en\s+[a-záéíóúàèìòùäëïöüñç\s\-]+",
        r"\bresido\s+en\s+[a-záéíóúàèìòùäëïöüñç\s\-]+",
        r"\bsoy\s+de\s+[a-záéíóúàèìòùäëïöüñç\s\-]+",
        r"\bestoy casad[oa]\b",
        r"\bestoy divorciad[oa]\b",
        r"\btengo\s+\d+\s+hij[oa]s?\b",

        # Work / income
        r"\btrabajo\s+(como|en|para)\b",
        r"\bsoy\s+(autónomo|autonomo|empleado|funcionario|empresario|jubilado|pensionista)\b",
        r"\bgano\s+[\d\.\,]+\s*(€|eur|euros)?\b",
        r"\bmis ingresos\b",
        r"\bmi sueldo\b",
        r"\bmi salario\b",

        # Assets / debts / liquidity
        r"\btengo\s+[\d\.\,]+\s*(€|eur|euros)\b",
        r"\btengo\s+(un|una)\s+(piso|casa|vivienda|hipoteca|deuda|préstamo|prestamo)\b",
        r"\bmi hipoteca\b",
        r"\bmi deuda\b",
        r"\bmi relaciones\b",
        r"\btengo invertid[oa]s?\b",
        r"\btengo ahorrad[oa]s?\b",

        # Goals / preferences
        r"\bquiero\s+(jubilarme|invertir|comprar|vender|ahorrar|amortizar)\b",
        r"\bmi objetivo\b",
        r"\bmi prioridad\b",
        r"\bmi perfil de riesgo\b",
        r"\bno quiero\s+(arriesgar|perder|invertir|comprar|vender)\b",
    ]

    return any(re.search(pattern, t, flags=re.IGNORECASE) for pattern in patterns)


def apply_deterministic_memory_guard(user_text: str, decision: dict) -> dict:
    if not isinstance(decision, dict):
        return decision

    if looks_like_stable_memory_fact(user_text):
        decision["should_update_memory"] = True

        if decision.get("route") in {None, "", "smalltalk"}:
            decision["route"] = "memory_update"

        reason = decision.get("reason") or ""
        if "deterministic memory guard" not in reason.lower():
            decision["reason"] = (reason + " Deterministic memory guard: stable user fact detected.").strip()

    return decision
# --- END DETERMINISTIC MEMORY GUARD v2 ---


def route_user_message(user_text: str) -> dict:
    """Use a tiny fast LLM call to decide the processing route."""
    system_prompt, prompt = build_router_prompt(user_text)

    raw_answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=system_prompt,
        provider_order=ROUTER_PROVIDER_ORDER,
        deepseek_model=ROUTER_DEEPSEEK_MODEL,
        gemini_model=ROUTER_GEMINI_MODEL,
    )

    raw_decision = extract_json_from_text(raw_answer)
    decision = normalize_router_decision(raw_decision)

    append_text_file(
        wealth_log_file(),
        "\n".join(
            [
                f"## {format_app_minute()}",
                "",
                "### Router decision",
                "",
                f"Route: {decision['route']}",
                f"Raw route: {decision.get('raw_route', '')}",
                f"Confidence: {decision['confidence']}",
                f"Depth: {decision['depth']}",
                f"Needs orchestrator: {decision['needs_orchestrator']}",
                f"Client files: {decision['client_files']}",
                f"Knowledge files: {decision['knowledge_files']}",
                f"Router provider: {usage.get('provider', 'unknown')}",
                f"Router model: {usage.get('model', 'unknown')}",
                f"Router tokens: {usage.get('total_tokens', 0)}",
                f"Router cost USD: {usage.get('estimated_cost_usd', 0):.6f}",
                f"Direct response present: {bool(decision.get('direct_response'))}",
                f"Reason: {decision['reason']}",
                "",
                "---",
                "",
            ]
        ),
    )

    return apply_deterministic_memory_guard(user_text, decision)
def save_immediate_followup_triggers(new_followup_triggers: list, source: str) -> list[dict]:
    """Save explicit user-requested seguimientos immediately."""
    if not isinstance(new_followup_triggers, list) or not new_followup_triggers:
        return []

    existing_triggers = load_followup_triggers()

    timestamp = format_app_datetime()
    id_timestamp = now_app().strftime("%Y%m%d_%H%M%S")

    normalized_triggers: list[dict] = []

    for index, trigger in enumerate(new_followup_triggers, start=1):
        if not isinstance(trigger, dict):
            continue

        normalized_trigger = {
            "id": f"followup_{id_timestamp}_{index}",
            "client_id": current_client_name(),
            "created_at": str(trigger.get("created_at") or timestamp),
            "date": str(trigger.get("date") or ""),
            "time": str(trigger.get("time") or ""),
            "type": validate_followup_type(str(trigger.get("type") or "general_checkin")),
            "message_template": str(trigger.get("message_template") or ""),
            "reason": str(trigger.get("reason") or ""),
            "sensitivity": str(trigger.get("sensitivity") or "low"),
            "requires_private_context": bool(trigger.get("requires_private_context", True)),
            "status": str(trigger.get("status") or "pending").lower(),
            "source": source,
        }

        if not normalized_trigger["date"] or not parse_followup_date(normalized_trigger["date"]):
            continue

        if normalized_trigger.get("time") and not parse_followup_time(str(normalized_trigger.get("time"))):
            continue

        if not normalized_trigger["message_template"]:
            continue

        if is_duplicate_followup_with_llm(normalized_trigger, existing_triggers + normalized_triggers):
            append_proactivity_log(
                event_type="followup_duplicate_skipped",
                trigger=normalized_trigger,
                notes="Skipped because a similar pending seguimiento already exists.",
            )
            continue

        normalized_triggers.append(normalized_trigger)

    if normalized_triggers:
        save_followup_triggers(existing_triggers + normalized_triggers)

        for trigger in normalized_triggers:
            append_proactivity_log(
                event_type="followup_created",
                trigger=trigger,
                notes="Created immediately from explicit user seguimiento request.",
            )

    return normalized_triggers


def process_with_router(text_to_process: str, source: str) -> str:
    """Route a user message using the LLM router, then answer directly or call the orchestrator."""
    if not ROUTER_ENABLED:
        return process_with_orchestrator(
            text_to_process=text_to_process,
            source=source,
        )

    try:
        decision = route_user_message(text_to_process)
    except Exception as error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Router failed",
                    "",
                    f"Error: {str(error)}",
                    "Fallback: full orchestrator.",
                    "",
                    "---",
                    "",
                ]
            ),
        )
        return process_with_orchestrator(
            text_to_process=text_to_process,
            source=source,
        )

    if not decision["needs_orchestrator"]:
        if decision["route"] == "followup_management":
            requested_followups = decision.get("followup_triggers", [])

            saved_followups = save_immediate_followup_triggers(
                requested_followups,
                source=source,
            )

            append_text_file(
                wealth_log_file(),
                "\n".join(
                    [
                        f"## {format_app_minute()}",
                        "",
                        "### Immediate seguimiento route",
                        "",
                        f"Requested seguimientos: {requested_followups}",
                        f"Saved seguimientos: {saved_followups}",
                        "",
                        "---",
                        "",
                    ]
                ),
            )

            if requested_followups and not saved_followups:
                return (
                    "Ya tenía un seguimiento parecido pendiente, así que no he creado otro duplicado. "
                    "Puedes verlo en ⏰ Seguimientos."
                )

        return decision["direct_response"]

    return process_with_orchestrator(
        text_to_process=text_to_process,
        source=source,
        selected_client_keys=filter_normal_orchestrator_client_keys(decision["client_files"]),
        selected_knowledge_files=decision["knowledge_files"],
        routing_decision=decision,
    )


def run_orchestrator_for_client(
    client_name: str,
    text_to_process: str,
    source: str,
) -> str:
    """Run the synchronous orchestrator inside the correct client context.

    Intended for asyncio.to_thread(), so blocking LLM calls do not freeze
    the Telegram event loop.
    """
    set_current_client_name(client_name)

    return process_with_router(
        text_to_process=text_to_process,
        source=source,
    )


async def send_long_message(
    update: Update,
    text: str,
    reply_markup=None,
    chunk_size: int = 3500,
) -> None:
    """Send long Telegram messages in safe chunks."""
    if not text:
        empty_response_message = load_required_app_text("empty_response_message.md")

        await update.message.reply_text(
            empty_response_message,
            reply_markup=reply_markup,
        )
        return

    chunks = [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    for index, chunk in enumerate(chunks):
        if index == len(chunks) - 1:
            await update.message.reply_text(
                chunk,
                reply_markup=reply_markup,
            )
        else:
            await update.message.reply_text(chunk)


# ---------------------------------------------------------------------
# Memory editing helpers
# ---------------------------------------------------------------------

def current_chat_key(update: Update) -> str:
    if update.effective_chat:
        return str(update.effective_chat.id)
    return current_client_name()


def memory_file_label(target_key: str) -> str:
    labels = {
        "quien_soy": "Quién soy",
        "que_quiero": "Qué quiero",
        "que_tengo_que_hacer": "Plan de acción",
        "estilo_respuesta": "Estilo de respuesta",
    }
    return labels.get(target_key, target_key)


def build_memory_patch_prompt(target_key: str, old_content: str, user_instruction: str) -> tuple[str, str]:
    system_prompt = """Eres un editor de memoria patrimonial del usuario.

Tu tarea:
- Recibir el contenido actual de un archivo markdown.
- Recibir una instrucción natural del usuario.
- Proponer una versión actualizada del archivo.
- Mantener la estructura del archivo.
- No inventar datos.
- No borrar información salvo que el usuario lo pida.
- Integrar el cambio en la sección más adecuada.
- Devolver SOLO JSON válido.

Formato JSON:
{
  "summary": "resumen corto del cambio propuesto",
  "updated_content": "contenido markdown completo actualizado"
}
"""

    prompt = f"""Archivo objetivo: {target_key}

CONTENIDO ACTUAL:

{old_content}

INSTRUCCIÓN DEL USUARIO:
{user_instruction}

Devuelve SOLO JSON válido con summary y updated_content.
"""
    return system_prompt, prompt


def propose_memory_patch(target_key: str, old_content: str, user_instruction: str) -> dict:
    system_prompt, prompt = build_memory_patch_prompt(target_key, old_content, user_instruction)

    raw_answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=system_prompt,
        provider_order=ROUTER_PROVIDER_ORDER,
        deepseek_model=ROUTER_DEEPSEEK_MODEL,
        gemini_model=ROUTER_GEMINI_MODEL,
    )

    data = extract_json_from_text(raw_answer)
    updated_content = str(data.get("updated_content", "")).strip()
    summary = str(data.get("summary", "")).strip()

    if not updated_content:
        raise RuntimeError("Memory patch did not include updated_content.")

    if not summary:
        summary = "Actualizar memoria."

    data["_usage"] = usage
    data["updated_content"] = updated_content
    data["summary"] = summary
    return data


async def begin_memory_edit(update: Update, target_key: str) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    chat_key = current_chat_key(update)
    PENDING_MEMORY_EDIT[chat_key] = {
        "client_name": current_client_name(),
        "target_key": target_key,
    }

    content = read_text_file(client_files()[target_key])
    label = memory_file_label(target_key)

    await send_long_message(
        update,
        f"""Aquí está tu resumen actual de **{label}**:

{content}

Para cambiarlo, responde en lenguaje natural.

Ejemplos:
- Añade que...
- Cambia X por Y...
- Borra lo de...
- Hazlo más corto...
""",
        reply_markup=MAIN_KEYBOARD,
    )


async def handle_memory_edit_instruction(update: Update, user_text: str) -> bool:
    chat_key = current_chat_key(update)

    if chat_key not in PENDING_MEMORY_EDIT:
        return False

    edit_state = PENDING_MEMORY_EDIT[chat_key]
    target_key = edit_state["target_key"]

    activate_client_from_update(update)
    ensure_client_files()

    old_content = read_text_file(client_files()[target_key])

    thinking_message = await update.message.reply_text(
        "Estoy preparando una propuesta de cambio para tu memoria…",
        reply_markup=MAIN_KEYBOARD,
    )

    try:
        patch = propose_memory_patch(target_key, old_content, user_text)
    except Exception as error:
        await thinking_message.edit_text(
            f"No he podido preparar el cambio de memoria: {error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    PENDING_MEMORY_PATCH[chat_key] = {
        "client_name": current_client_name(),
        "target_key": target_key,
        "old_content": old_content,
        "updated_content": patch["updated_content"],
        "summary": patch["summary"],
    }

    label = memory_file_label(target_key)

    preview = f"""Propongo este cambio en **{label}**:

Resumen:
{patch["summary"]}

Nuevo contenido completo:

{patch["updated_content"]}

¿Confirmas que guarde este cambio?
"""

    await thinking_message.delete()
    await send_long_message(
        update,
        preview,
        reply_markup=MEMORY_CONFIRM_KEYBOARD,
    )
    return True


async def confirm_memory_change(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    chat_key = current_chat_key(update)

    if chat_key not in PENDING_MEMORY_PATCH:
        await update.message.reply_text(
            "No hay ningún cambio de memoria pendiente de confirmar.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    patch = PENDING_MEMORY_PATCH.pop(chat_key)
    PENDING_MEMORY_EDIT.pop(chat_key, None)

    target_key = patch["target_key"]
    target_path = client_files()[target_key]

    timestamp = datetime.now(ZoneInfo(APP_TIMEZONE)).strftime("%Y%m%d_%H%M%S")
    backup_dir = target_path.parent / "backups" / f"memory_edit_{timestamp}"
    backup_dir.mkdir(parents=True, exist_ok=True)

    if target_path.exists():
        shutil.copy2(target_path, backup_dir / target_path.name)

    write_text_file(target_path, patch["updated_content"])

    append_text_file(
        wealth_log_file(),
        "\n".join(
            [
                f"## {format_app_minute()}",
                "",
                "### Memory edit confirmed",
                "",
                f"Target: {target_key}",
                f"Summary: {patch['summary']}",
                f"Backup folder: {backup_dir}",
                "",
                "---",
                "",
            ]
        ),
    )

    await update.message.reply_text(
        f"Memoria actualizada: {memory_file_label(target_key)}.\n\nHe guardado copia de seguridad.",
        reply_markup=MAIN_KEYBOARD,
    )


async def cancel_memory_change(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_key = current_chat_key(update)
    PENDING_MEMORY_EDIT.pop(chat_key, None)
    PENDING_MEMORY_PATCH.pop(chat_key, None)

    await update.message.reply_text(
        "Cambio de memoria cancelado.",
        reply_markup=MAIN_KEYBOARD,
    )



# ---------------------------------------------------------------------
# Session buffer helpers
# ---------------------------------------------------------------------


# --- Uploaded document ingestion v0.1 ---
SUPPORTED_UPLOAD_EXTENSIONS = {".pdf", ".xlsx", ".csv", ".txt", ".md"}
UPLOAD_SESSION_STALE_MINUTES = 5
UPLOAD_NOTICE_TASKS: dict[str, asyncio.Task] = {}


def uploaded_docs_root_dir() -> Path:
    return active_client_dir() / "UploadedDocs"


def active_upload_session_marker_file() -> Path:
    return uploaded_docs_root_dir() / "_active_upload_session.txt"


def safe_uploaded_filename(filename: str) -> str:
    base = (filename or "uploaded_file").strip()
    base = re.sub(r"[^\w\-.() áéíóúàèìòùäëïöüÁÉÍÓÚÀÈÌÒÙÄËÏÖÜñÑçÇ]+", "_", base)
    base = base.strip("._ ")
    return base or "uploaded_file"


def create_upload_session_id() -> str:
    return now_app().strftime("%Y%m%d_%H%M%S")


def get_or_create_upload_session_dir() -> Path:
    root = uploaded_docs_root_dir()
    root.mkdir(parents=True, exist_ok=True)

    marker = active_upload_session_marker_file()
    session_id = ""

    if marker.exists():
        session_id = marker.read_text(encoding="utf-8").strip()

    if not session_id:
        session_id = create_upload_session_id()
        marker.write_text(session_id, encoding="utf-8")

    session_dir = root / session_id
    (session_dir / "raw").mkdir(parents=True, exist_ok=True)
    (session_dir / "extracted").mkdir(parents=True, exist_ok=True)
    (session_dir / "output").mkdir(parents=True, exist_ok=True)

    return session_dir


def latest_upload_session_dir() -> Path | None:
    root = uploaded_docs_root_dir()
    marker = active_upload_session_marker_file()

    if marker.exists():
        session_id = marker.read_text(encoding="utf-8").strip()
        if session_id and (root / session_id).exists():
            return root / session_id

    if not root.exists():
        return None

    sessions = [
        p for p in root.iterdir()
        if p.is_dir() and not p.name.startswith("_")
    ]

    if not sessions:
        return None

    return sorted(sessions, key=lambda p: p.name)[-1]


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    parent = path.parent

    for i in range(2, 1000):
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate

    raise RuntimeError(f"Could not create unique path for {path}")


def extract_pdf_text_and_tables(path: Path) -> str:
    try:
        import pdfplumber
    except Exception as exc:
        return f"[ERROR] pdfplumber is not installed or failed to import: {exc}"

    parts: list[str] = [f"# Extracted PDF: {path.name}"]

    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                parts.append(f"\n\n## Page {page_index}")

                text = page.extract_text() or ""
                if text.strip():
                    parts.append("\n### Text\n")
                    parts.append(text.strip())

                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []

                for table_index, table in enumerate(tables, start=1):
                    parts.append(f"\n### Table {table_index}\n")
                    for row in table:
                        cleaned = [
                            (cell or "").replace("\n", " ").strip()
                            for cell in row
                        ]
                        parts.append(" | ".join(cleaned))

        return "\n".join(parts).strip()

    except Exception as exc:
        return f"[ERROR] Could not extract PDF {path.name}: {exc}"


def extract_xlsx_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return f"[ERROR] openpyxl is not installed or failed to import: {exc}"

    parts: list[str] = [f"# Extracted Excel: {path.name}"]

    try:
        wb = load_workbook(path, data_only=True, read_only=True)

        for ws in wb.worksheets:
            parts.append(f"\n\n## Sheet: {ws.title}")
            max_rows = min(ws.max_row or 0, 500)
            max_cols = min(ws.max_column or 0, 40)

            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    parts.append(" | ".join(values))

        return "\n".join(parts).strip()

    except Exception as exc:
        return f"[ERROR] Could not extract Excel {path.name}: {exc}"


def extract_csv_text(path: Path) -> str:
    parts: list[str] = [f"# Extracted CSV: {path.name}"]

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 500:
                    parts.append("[TRUNCATED AFTER 500 ROWS]")
                    break
                parts.append(" | ".join(str(cell).strip() for cell in row))
        return "\n".join(parts).strip()

    except UnicodeDecodeError:
        try:
            with path.open("r", encoding="latin-1", newline="") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 500:
                        parts.append("[TRUNCATED AFTER 500 ROWS]")
                        break
                    parts.append(" | ".join(str(cell).strip() for cell in row))
            return "\n".join(parts).strip()
        except Exception as exc:
            return f"[ERROR] Could not extract CSV {path.name}: {exc}"

    except Exception as exc:
        return f"[ERROR] Could not extract CSV {path.name}: {exc}"


def extract_plain_text_file(path: Path) -> str:
    try:
        return f"# Extracted text file: {path.name}\n\n" + path.read_text(encoding="utf-8", errors="replace").strip()
    except Exception as exc:
        return f"[ERROR] Could not read text file {path.name}: {exc}"


def extract_uploaded_file(path: Path) -> str:
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return extract_pdf_text_and_tables(path)

    if suffix == ".xlsx":
        return extract_xlsx_text(path)

    if suffix == ".csv":
        return extract_csv_text(path)

    if suffix in {".txt", ".md"}:
        return extract_plain_text_file(path)

    return f"[UNSUPPORTED] {path.name}"


def extract_latest_upload_session() -> tuple[Path | None, str]:
    session_dir = latest_upload_session_dir()
    if session_dir is None:
        return None, "No hay documentos subidos para analizar."

    raw_dir = session_dir / "raw"
    extracted_dir = session_dir / "extracted"
    output_dir = session_dir / "output"
    extracted_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        p for p in raw_dir.iterdir()
        if p.is_file() and p.suffix.lower() in SUPPORTED_UPLOAD_EXTENSIONS
    )

    if not files:
        return session_dir, "No hay documentos compatibles en la última sesión de subida."

    extracted_parts: list[str] = [
        "# Uploaded document extraction",
        "",
        f"Client: {current_client_name()}",
        f"Session: {session_dir.name}",
        "",
        "The following content was extracted from user-uploaded documents. No OCR was used. If data is missing, ask for clearer files or bank confirmation.",
    ]

    for file_path in files:
        extracted = extract_uploaded_file(file_path)
        extracted_path = extracted_dir / f"{file_path.stem}.txt"
        extracted_path.write_text(extracted, encoding="utf-8")
        extracted_parts.append("\n\n---\n\n")
        extracted_parts.append(extracted)

    combined = "\n".join(extracted_parts).strip()
    (output_dir / "extracted_context.md").write_text(combined, encoding="utf-8")

    return session_dir, combined


def build_uploaded_docs_orchestrator_message(extracted_context: str) -> str:
    return f"""
El cliente ha subido documentos patrimoniales por Telegram. Analiza el contenido extraído y crea una primera estructura de cartera.

Reglas esenciales:
- El cliente NO tiene que crear tablas, calcular costes ni comparar fondos.
- El cliente aporta documentos o respuestas; el eCoach extrae, ordena, calcula, estima y detecta datos faltantes.
- No hagas recomendaciones de producto concretas.
- No inventes datos que no aparecen.
- Si falta información, márcala como "No encontrado", "Pendiente" o "Estimación inicial".
- No prepares todavía mensajes para el banco en esta fase. Si falta información privada, enumérala como "datos pendientes", pero no conviertas eso en la próxima acción todavía.
- No uses una tabla Markdown ancha: en Telegram se lee mal.
- Usa formato móvil-friendly: secciones claras, bullets y fichas por fondo.

Estructura obligatoria de la respuesta:

1. Confirmación breve:
   "He leído los documentos subidos y he creado una primera estructura."

2. Fuentes usadas:
   Explica qué datos vienen del Excel y qué datos vienen de los PDFs/fichas.
   Ejemplo:
   - Excel: importes y pesos.
   - PDFs/fichas: ISIN, gestora, comisiones visibles, riesgo, categoría, liquidez, rentabilidad, benchmark/referencia si aparece.

3. Cartera detectada:
   Lista compacta con fondo, importe y peso.

4. Coste visible estimado:
   Calcula coste anual aproximado en euros si hay importe y porcentaje de coste visible.
   Aclara si es coste mínimo visible y qué costes pueden faltar: TER/OCF completo, costes de subyacentes, custodia, retrocesiones, informe ex-post.

5. Ficha por fondo:
   Máximo 4-5 líneas por fondo.
   Para cada fondo:
   **Nombre del fondo**
   - ISIN:
   - Importe / peso:
   - Coste visible:
   - Riesgo / categoría:
   - Liquidez / horizonte:
   - Observación clave:
   - Datos pendientes:

6. Estimación inicial de exposición:
   Debes estimar, aunque sea en rangos, la exposición aproximada de la cartera a:
   - renta variable;
   - renta fija / monetario / liquidez / otros.

   Usa la política de inversión de cada fondo:
   - Si un fondo invierte 30-50% en renta variable, usa ese rango y un punto medio solo como aproximación.
   - Si un fondo tiene máximo 25% en renta variable, usa rango 0-25% y una estimación prudente.
   - Si un fondo tiene máximo 15% en renta variable, usa rango 0-15% y una estimación prudente.
   - Si un fondo invierte >75% en renta variable, trátalo como mayoritariamente renta variable.

   Reglas de coherencia:
   - No des dos rangos contradictorios.
   - Si haces un cálculo ponderado y da ~20%, no escribas después "30-40%".
   - Para esta cartera de ejemplo, si no hay composición exacta, una buena salida sería:
     "Con los datos disponibles, la cartera parece tener aproximadamente 18-28% en renta variable, con un punto medio cercano al 20-22%. El resto estaría en renta fija, monetario, liquidez u otros activos. Es una estimación inicial basada en las fichas; la refinaré si aparecen datos exactos de composición."

7. Referencia / benchmark:
   Si aparece un benchmark oficial en los documentos, menciónalo.
   Si no aparece, NO inventes benchmark oficial.
   Usa "mezcla de referencia" o "exposición de referencia".
   Ejemplo:
   "No es el benchmark oficial de la cartera; es una mezcla de referencia para comparar riesgo: aprox. 25% renta variable global + 75% renta fija/monetario EUR."

8. Lectura inicial:
   Máximo 5 bullets.
   Resume concentración, gestora, perfil de riesgo, fondos caros y puntos de atención.
   No digas que la cartera es buena o mala todavía.

9. Gestión emocional:
   Breve: 3-4 frases.
   Explica que el miedo baja cuando los documentos se convierten en estructura.
   No empujes a decidir ni a mover dinero.

10. Datos que puedo intentar completar con fuentes públicas o documentos adicionales:
   Esta sección es obligatoria.
   Incluye solo datos públicos o potencialmente públicos:
   - KID / DFI actualizado;
   - ficha actualizada;
   - TER/OCF público;
   - benchmark oficial si está publicado;
   - composición actual por renta variable / renta fija;
   - categoría oficial;
   - datos públicos de la gestora/fondo;
   - alternativas y proveedores si el cliente quiere explorar.

   IMPORTANTE:
   No pongas KID/DFI, ficha pública, benchmark público o categoría oficial dentro del mensaje al banco como si fueran datos privados. Primero deben aparecer aquí, como datos que el eCoach puede intentar completar con fuentes públicas.

11. Datos privados que debe aportar el banco/cliente:
   Esta sección es obligatoria.
   Incluye solo datos privados o específicos de la cuenta:
   - fecha/precio de compra;
   - plusvalías/minusvalías;
   - informe personalizado de costes y gastos ex-post;
   - custodia aplicada a esa cuenta;
   - retrocesiones efectivamente aplicadas a esa relación;
   - situación fiscal;
   - otros productos contratados por el cliente.

   No mezcles datos públicos con datos privados.

12. Siguiente decisión del cliente:
   No prepares todavía ningún mensaje para el banco.
   No escribas "Mensaje para el banco".
   No escribas "Próxima acción sugerida".
   No ofrezcas una lista larga de caminos.
   Después del diagnóstico, ofrece solo tres posibilidades de inversión, con este contenido:

   "Dada tu cartera y lo que quieres —entender, bajar costes y no decidir a ciegas— hay tres caminos principales:

   1. **Cartera autogestionada**: tú eliges y ejecutas directamente los fondos.
   2. **Gestión delegada**: delegas la cartera en un proveedor.
   3. **Cartera guiada por eCoach**: tú mantienes el control, pero eCoach te ayuda a entender, comparar y avanzar paso a paso.

   ¿Qué camino quieres explorar primero?"

13. Independencia:
   Si se mencionan alternativas, proveedores o siguientes caminos, debes decir:
   "Soy independiente del proveedor que elijas. Puedo trabajar con tu banco actual, MyInvestor, Renta 4, Indexa, Openbank u otro proveedor. El proveedor no necesita saber que te estoy ayudando. Yo no tomo decisiones por ti: te ayudo a entender, comparar y ejecutar solo si tú decides."

16. Fallback:
   Termina diciendo que si el banco responde de forma incompleta, el cliente puede subir aquí la respuesta y el eCoach preparará una segunda pregunta.

Contenido extraído:

<<<DOCUMENTOS_EXTRAIDOS
{extracted_context}
DOCUMENTOS_EXTRAIDOS>>>
""".strip()


async def reply_message_text_in_chunks(message, text: str, chunk_size: int = 3800) -> None:
    if message is None:
        return

    if not text:
        return

    paragraphs = text.split("\n\n")
    chunks: list[str] = []
    current = ""

    for paragraph in paragraphs:
        candidate = paragraph if not current else current + "\n\n" + paragraph

        if len(candidate) <= chunk_size:
            current = candidate
            continue

        if current:
            chunks.append(current)
            current = ""

        if len(paragraph) > chunk_size:
            lines = paragraph.split("\n")
            line_current = ""

            for line in lines:
                line_candidate = line if not line_current else line_current + "\n" + line

                if len(line_candidate) <= chunk_size:
                    line_current = line_candidate
                else:
                    if line_current:
                        chunks.append(line_current)
                    line_current = line

            if line_current:
                current = line_current
        else:
            current = paragraph

    if current:
        chunks.append(current)

    for chunk in chunks:
        await message.reply_text(chunk)


async def reply_text_in_chunks(update: Update, text: str, chunk_size: int = 3800) -> None:
    message = update.effective_message
    await reply_message_text_in_chunks(message, text, chunk_size=chunk_size)



def upload_session_file_count(session_dir: Path) -> int:
    raw_dir = session_dir / "raw"
    if not raw_dir.exists():
        return 0
    return len([p for p in raw_dir.iterdir() if p.is_file()])


async def delayed_upload_summary_notice(
    bot,
    chat_id: int,
    client_name: str,
    session_dir: Path,
    delay_seconds: float = 4.0,
) -> None:
    try:
        await asyncio.sleep(delay_seconds)

        set_current_client_name(client_name)
        file_count = upload_session_file_count(session_dir)

        upload_inline_keyboard = InlineKeyboardMarkup(
            [[InlineKeyboardButton("📄 Analizar documentos", callback_data="analyze_uploaded_documents")]]
        )

        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"Documentos recibidos en esta sesión: {file_count}.\n"
                "Cuando haya subido todos, pulse el botón de abajo."
            ),
            reply_markup=upload_inline_keyboard,
        )

    except asyncio.CancelledError:
        return
    except Exception as exc:
        print(f"[UPLOAD_NOTICE_ERROR] {type(exc).__name__}: {exc}", flush=True)


def schedule_upload_summary_notice(update: Update, context: ContextTypes.DEFAULT_TYPE, session_dir: Path) -> None:
    if update.effective_chat is None:
        return

    client_name = current_client_name()
    chat_id = update.effective_chat.id
    key = f"{client_name}:{chat_id}"

    previous = UPLOAD_NOTICE_TASKS.get(key)
    if previous is not None and not previous.done():
        previous.cancel()

    UPLOAD_NOTICE_TASKS[key] = asyncio.create_task(
        delayed_upload_summary_notice(
            bot=context.bot,
            chat_id=chat_id,
            client_name=client_name,
            session_dir=session_dir,
        )
    )


async def handle_uploaded_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    client_name = client_name_from_update(update)
    set_current_client_name(client_name)
    ensure_client_files()

    message = update.effective_message
    if message is None or message.document is None:
        return

    document = message.document
    original_name = document.file_name or f"telegram_file_{document.file_unique_id}"
    safe_name = safe_uploaded_filename(original_name)
    suffix = Path(safe_name).suffix.lower()

    if suffix not in SUPPORTED_UPLOAD_EXTENSIONS:
        await message.reply_text(
            "He recibido el archivo, pero por ahora solo puedo procesar PDF, XLSX, CSV, TXT y MD."
        )
        return

    session_dir = get_or_create_upload_session_dir()
    raw_dir = session_dir / "raw"
    destination = unique_path(raw_dir / safe_name)

    telegram_file = await context.bot.get_file(document.file_id)
    await telegram_file.download_to_drive(custom_path=str(destination))

    # eCoach_Relaciones v3: document receipt is project state, not only a raw upload.
    ecoach_mark_documents_received(active_client_dir(), filename=safe_name)

    # Do not reply after every file. Telegram sends albums/batches as separate updates.
    # We debounce and send one contextual inline button after the upload burst ends.
    schedule_upload_summary_notice(update, context, session_dir)


async def analyze_uploaded_documents_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    client_name = client_name_from_update(update)
    set_current_client_name(client_name)
    ensure_client_files()

    message = update.effective_message
    if message is None:
        return

    await message.reply_text("Voy a leer los documentos subidos y convertirlos en una primera estructura.")

    session_dir, extracted_context = await asyncio.to_thread(extract_latest_upload_session)

    if session_dir is None or not extracted_context.strip() or extracted_context.startswith("No hay"):
        await message.reply_text(extracted_context)
        return

    ecoach_mark_documents_analyzed(active_client_dir())

    ecoach_mark_documents_analyzed(active_client_dir())

    orchestrator_message = build_uploaded_docs_orchestrator_message(extracted_context)

    answer = await asyncio.to_thread(
        run_orchestrator_for_client,
        client_name,
        orchestrator_message,
        "uploaded_documents",
    )

    # End current upload batch after analysis so the next upload creates a fresh session.
    marker = active_upload_session_marker_file()
    if marker.exists():
        marker.unlink()

    answer = clean_document_diagnosis_for_demo(answer)
    await send_long_message(update, answer, reply_markup=alternatives_path_keyboard())
    ecoach_mark_diagnosis_complete(active_client_dir(), source="uploaded_documents_message")


async def analyze_uploaded_documents_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query is not None:
        await query.answer()

    client_name = client_name_from_update(update)
    set_current_client_name(client_name)
    ensure_client_files()

    message = query.message if query is not None else update.effective_message
    if message is None:
        return

    await message.reply_text("Voy a leer los documentos subidos y convertirlos en una primera estructura.")

    session_dir, extracted_context = await asyncio.to_thread(extract_latest_upload_session)

    if session_dir is None or not extracted_context.strip() or extracted_context.startswith("No hay"):
        await message.reply_text(extracted_context)
        return

    orchestrator_message = build_uploaded_docs_orchestrator_message(extracted_context)

    answer = await asyncio.to_thread(
        run_orchestrator_for_client,
        client_name,
        orchestrator_message,
        "uploaded_documents",
    )

    marker = active_upload_session_marker_file()
    if marker.exists():
        marker.unlink()

    answer = clean_document_diagnosis_for_demo(answer)
    await reply_message_text_in_chunks(message, answer)
    await message.reply_text(
        "Elige una de estas tres opciones para continuar:",
        reply_markup=alternatives_path_keyboard(),
    )
    ecoach_mark_diagnosis_complete(active_client_dir(), source="uploaded_documents_callback")

# --- End uploaded document ingestion v0.1 ---


def session_buffer_file() -> Path:
    return client_files()["quien_soy"].parent / "session_buffer.md"


def session_archive_dir() -> Path:
    return client_files()["quien_soy"].parent / "session_archives"



def should_skip_session_buffer(user_text: str, assistant_answer: str = "") -> bool:
    """Skip operational commands that are already handled immediately elsewhere."""
    text = f"{user_text}\n{assistant_answer}".lower()

    explicit_followup_markers = [
        "recuérdame",
        "recuerdame",
        "recordarme",
        "acuérdame",
        "acuerdame",
        "hazme un seguimiento",
        "crea un seguimiento",
        "crear un seguimiento",
        "seguimiento para",
        "followup para",
        "recuérdalo",
        "recuerdalo",
        "te recordaré",
        "te recordare",
    ]

    return any(marker in text for marker in explicit_followup_markers)



def append_session_buffer(user_text: str, assistant_answer: str) -> None:
    """Append one user/assistant exchange to the current session buffer."""
    if should_skip_session_buffer(user_text, assistant_answer):
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Session buffer append skipped",
                    "",
                    "Reason: explicit seguimiento/reminder command already handled immediately.",
                    f"Input preview: {compact_log_preview(user_text, 500)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )
        return

    buffer_path = session_buffer_file()

    if not buffer_path.exists():
        write_text_file(
            buffer_path,
            "\n".join(
                [
                    "# Session buffer",
                    "",
                    "This file contains conversation since the last confirmed Second Brain consolidation.",
                    "",
                    "---",
                    "",
                ]
            ),
        )

    append_text_file(
        buffer_path,
        "\n".join(
            [
                f"## {format_app_minute()}",
                "",
                "### User",
                "",
                user_text.strip(),
                "",
                "### eCoach",
                "",
                assistant_answer.strip(),
                "",
                "---",
                "",
            ]
        ),
    )


def clear_session_buffer_to_archive() -> Path:
    """Archive the current session buffer and reset it."""
    buffer_path = session_buffer_file()
    archive_dir = session_archive_dir()
    archive_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(ZoneInfo(APP_TIMEZONE)).strftime("%Y%m%d_%H%M%S")
    archive_path = archive_dir / f"session_buffer_{timestamp}.md"

    if buffer_path.exists():
        shutil.copy2(buffer_path, archive_path)

    write_text_file(
        buffer_path,
        "\n".join(
            [
                "# Session buffer",
                "",
                "This file contains conversation since the last confirmed Second Brain consolidation.",
                "",
                "---",
                "",
            ]
        ),
    )

    return archive_path


# ---------------------------------------------------------------------
# Memory capture detection helpers
# ---------------------------------------------------------------------

def detect_memory_capture_candidates(user_text: str, assistant_answer: str = "") -> dict:
    """Detect possible memory/seguimiento candidates without saving anything."""
    system_prompt = load_required_app_text("memory_capture_system_prompt.md")

    prompt = f"""USER MESSAGE:
{user_text}

ECOACH ANSWER:
{assistant_answer}

Return only valid JSON.
"""

    raw_answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=system_prompt,
        provider_order=ROUTER_PROVIDER_ORDER,
        deepseek_model=ROUTER_DEEPSEEK_MODEL,
        gemini_model=ROUTER_GEMINI_MODEL,
    )

    data = extract_json_from_text(raw_answer)

    if not isinstance(data, dict):
        raise RuntimeError("Memory capture detector did not return a JSON object.")

    data["_usage"] = usage
    return data


def log_memory_capture_candidates(user_text: str, assistant_answer: str = "") -> None:
    """Run memory capture detection and log candidates only. Does not modify memory."""
    try:
        data = detect_memory_capture_candidates(
            user_text=user_text,
            assistant_answer=assistant_answer,
        )

        memory_candidates = data.get("memory_candidates", [])
        followup_candidates = data.get("followup_candidates", [])

        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Memory capture candidates",
                    "",
                    f"Has candidates: {data.get('has_candidates', False)}",
                    f"Memory candidates: {memory_candidates}",
                    f"Seguimiento candidates: {followup_candidates}",
                    f"Detector provider: {data.get('_usage', {}).get('provider', 'unknown')}",
                    f"Detector model: {data.get('_usage', {}).get('model', 'unknown')}",
                    f"Detector tokens: {data.get('_usage', {}).get('total_tokens', 0)}",
                    f"Detector cost USD: {data.get('_usage', {}).get('estimated_cost_usd', 0):.6f}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

    except Exception as error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Memory capture detector failed",
                    "",
                    f"Error: {str(error)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )


async def show_session_buffer_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    buffer_path = session_buffer_file()

    if not buffer_path.exists():
        await update.message.reply_text(
            "No hay sesión pendiente de guardar todavía.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    content = read_text_file(buffer_path).strip()

    if not content or content.count("### User") == 0:
        await update.message.reply_text(
            "No hay conversación nueva pendiente de guardar en Second Brain.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    thinking_message = await update.message.reply_text(
        "Estoy resumiendo esta sesión para proponerte cambios en tu Second Brain…",
        reply_markup=MAIN_KEYBOARD,
    )

    try:
        data = await asyncio.wait_for(
            asyncio.to_thread(generate_session_consolidation),
            timeout=120,
        )
    except asyncio.TimeoutError:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Session consolidation timeout",
                    "",
                    f"Client: {current_client_name()}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        await thinking_message.edit_text(
            "La consolidación de la sesión está tardando demasiado y he cancelado este intento. "
            "La sesión sigue guardada en el buffer; puedes intentarlo de nuevo.",
            reply_markup=MAIN_KEYBOARD,
        )
        return
    except Exception as error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Session consolidation failed",
                    "",
                    f"Client: {current_client_name()}",
                    f"Error: {str(error)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        await thinking_message.edit_text(
            f"No he podido consolidar la sesión: {error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    preview = format_session_consolidation_preview(data)

    has_updates = any(
        [
            data.get("quien_soy_updates"),
            data.get("que_quiero_updates"),
            data.get("que_tengo_que_hacer_updates"),
            data.get("followup_candidates"),
        ]
    )

    try:
        await thinking_message.delete()
    except Exception:
        pass

    if not has_updates:
        await send_long_message(
            update,
            preview,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    chat_key = current_chat_key(update)
    PENDING_SESSION_CONSOLIDATION[chat_key] = data

    await send_long_message(
        update,
        preview,
        reply_markup=SESSION_CONSOLIDATION_CONFIRM_KEYBOARD,
    )



# ---------------------------------------------------------------------
# Session consolidation helpers
# ---------------------------------------------------------------------

def build_session_consolidation_prompt() -> tuple[str, str]:
    system_prompt = load_required_app_text("session_consolidation_system_prompt.md")

    buffer_path = session_buffer_file()
    buffer_content = read_text_file(buffer_path) if buffer_path.exists() else ""

    quien_soy = read_text_file(client_files()["quien_soy"])
    que_quiero = read_text_file(client_files()["que_quiero"])
    que_tengo_que_hacer = read_text_file(client_files()["que_tengo_que_hacer"])

    followups_path = client_files().get("followup_triggers")
    followups_content = read_text_file(followups_path) if followups_path and followups_path.exists() else ""

    prompt = f"""SESSION BUFFER:
{buffer_content}

CURRENT QUIEN_SOY:
{quien_soy}

CURRENT QUE_QUIERO:
{que_quiero}

CURRENT QUE_TENGO_QUE_HACER:
{que_tengo_que_hacer}

CURRENT FOLLOWUPS:
{followups_content}

Return only valid JSON.
"""
    return system_prompt, prompt


def generate_session_consolidation() -> dict:
    system_prompt, prompt = build_session_consolidation_prompt()

    raw_answer, usage = llm_generate(
        prompt=prompt,
        system_prompt=system_prompt,
        provider_order=ROUTER_PROVIDER_ORDER,
        deepseek_model=ROUTER_DEEPSEEK_MODEL,
        gemini_model=ROUTER_GEMINI_MODEL,
    )

    data = extract_json_from_text(raw_answer)

    if not isinstance(data, dict):
        raise RuntimeError("Session consolidation did not return a JSON object.")

    data["_usage"] = usage
    return data


def format_session_consolidation_preview(data: dict) -> str:
    lines = []

    summary = str(data.get("summary", "")).strip()
    if summary:
        lines.extend(["Resumen de la sesión:", summary, ""])

    sections = [
        ("Quién soy", data.get("quien_soy_updates", [])),
        ("Qué quiero", data.get("que_quiero_updates", [])),
        ("Plan de acción", data.get("que_tengo_que_hacer_updates", [])),
    ]

    for title, items in sections:
        clean_items = [str(item).strip() for item in items if str(item).strip()]
        if clean_items:
            lines.append(f"{title}:")
            for item in clean_items:
                lines.append(f"- {item}")
            lines.append("")

    followups = data.get("followup_candidates", [])
    if followups:
        lines.append("Seguimientos propuestos:")
        for item in followups:
            title = str(item.get("title", "")).strip()
            waiting_for = str(item.get("waiting_for", "")).strip()
            if title or waiting_for:
                lines.append(f"- {title}: {waiting_for}")
        lines.append("")

    if not lines:
        return "No he detectado cambios claros para guardar en Second Brain."

    lines.append("¿Quieres guardar esta consolidación en tu Second Brain?")
    return "\n".join(lines)




def apply_session_updates_to_memory_file(target_key: str, old_content: str, updates: list[str]) -> str:
    """Integrate confirmed session updates into the existing memory markdown."""
    clean_updates = [str(item).strip() for item in updates if str(item).strip()]

    if not clean_updates:
        return old_content

    system_prompt = """Eres un editor de memoria patrimonial del usuario.

Tu tarea:
- Recibir un archivo markdown actual del Second Brain.
- Recibir una lista de datos confirmados por el usuario.
- Integrar esos datos en las secciones adecuadas del archivo.
- Mantener la estructura existente todo lo posible.
- No inventar datos.
- No borrar datos salvo que sean claramente reemplazados por información nueva.
- Está prohibido devolver una sección llamada 'Consolidaciones de sesión'.
- Si el archivo actual ya contiene secciones 'Consolidaciones de sesión', elimina esas secciones después de integrar sus datos útiles en las secciones adecuadas.
- Integra edad, residencia, gastos, hipoteca, liquidez, ingresos, objetivos y tareas en las secciones existentes correspondientes.
- Si un dato nuevo reemplaza uno anterior, conserva solo el dato más reciente.
- Si un dato ya existe, no lo dupliques; actualízalo o mantenlo.
- Devolver SOLO el markdown completo actualizado, sin JSON, sin comentarios, sin explicación.
"""

    updates_text = "\n".join(f"- {item}" for item in clean_updates)

    prompt = f"""Archivo objetivo: {target_key}

CONTENIDO ACTUAL:
{old_content}

DATOS CONFIRMADOS PARA INTEGRAR:
{updates_text}

Devuelve SOLO el markdown completo actualizado.
"""

    updated_content, usage = llm_generate(
        prompt=prompt,
        system_prompt=system_prompt,
        provider_order=ROUTER_PROVIDER_ORDER,
        deepseek_model=ROUTER_DEEPSEEK_MODEL,
        gemini_model=ROUTER_GEMINI_MODEL,
    )

    updated_content = updated_content.strip()

    if not updated_content:
        raise RuntimeError(f"No updated content returned for {target_key}.")

    if "Consolidaciones de sesión" in updated_content:
        raise RuntimeError(
            f"Integration output for {target_key} still contains forbidden 'Consolidaciones de sesión' section."
        )

    # Defensive cleanup in case the model wraps markdown in code fences.
    if updated_content.startswith("```markdown"):
        updated_content = updated_content.removeprefix("```markdown").strip()
    if updated_content.startswith("```"):
        updated_content = updated_content.removeprefix("```").strip()
    if updated_content.endswith("```"):
        updated_content = updated_content.removesuffix("```").strip()

    return updated_content.strip() + "\n"


async def confirm_session_consolidation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    chat_key = current_chat_key(update)

    if chat_key not in PENDING_SESSION_CONSOLIDATION:
        await update.message.reply_text(
            "No hay ninguna consolidación de sesión pendiente de confirmar.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    data = PENDING_SESSION_CONSOLIDATION.pop(chat_key)

    timestamp = datetime.now(ZoneInfo(APP_TIMEZONE)).strftime("%Y%m%d_%H%M%S")
    backup_dir = client_files()["quien_soy"].parent / "backups" / f"session_consolidation_{timestamp}"
    backup_dir.mkdir(parents=True, exist_ok=True)

    target_map = {
        "quien_soy": data.get("quien_soy_updates", []),
        "que_quiero": data.get("que_quiero_updates", []),
        "que_tengo_que_hacer": data.get("que_tengo_que_hacer_updates", []),
    }

    changed_targets = []

    for target_key, updates in target_map.items():
        clean_updates = [str(item).strip() for item in updates if str(item).strip()]
        if not clean_updates:
            continue

        target_path = client_files()[target_key]

        if target_path.exists():
            shutil.copy2(target_path, backup_dir / target_path.name)

        old_content = read_text_file(target_path)

        try:
            new_content = apply_session_updates_to_memory_file(
                target_key=target_key,
                old_content=old_content,
                updates=clean_updates,
            )
        except Exception as error:
            append_text_file(
                wealth_log_file(),
                "\n".join(
                    [
                        f"## {format_app_minute()}",
                        "",
                        "### Session memory integration failed",
                        "",
                        f"Target: {target_key}",
                        f"Error: {str(error)}",
                        "Fallback: append updates under session consolidation section.",
                        "",
                        "---",
                        "",
                    ]
                ),
            )

            addition = [
                "",
                "## Consolidaciones de sesión",
                f"- Fecha: {format_app_minute()}",
            ]

            for item in clean_updates:
                addition.append(f"- {item}")

            new_content = old_content.rstrip() + "\n" + "\n".join(addition) + "\n"

        write_text_file(target_path, new_content)
        changed_targets.append(target_key)

    followups = data.get("followup_candidates", [])

    archive_path = clear_session_buffer_to_archive()

    append_text_file(
        wealth_log_file(),
        "\n".join(
            [
                f"## {format_app_minute()}",
                "",
                "### Session consolidation confirmed",
                "",
                f"Changed targets: {changed_targets}",
                f"Seguimiento candidates not yet applied: {followups}",
                f"Backup folder: {backup_dir}",
                f"Archived session buffer: {archive_path}",
                "",
                "---",
                "",
            ]
        ),
    )

    if changed_targets:
        labels = ", ".join(memory_file_label(target) for target in changed_targets)
        await update.message.reply_text(
            f"Sesión guardada correctamente en: {labels}.\n\nHe archivado el buffer de sesión y guardado copia de seguridad.",
            reply_markup=MAIN_KEYBOARD,
        )
    else:
        await update.message.reply_text(
            "No había cambios claros que guardar. He archivado el buffer de sesión.",
            reply_markup=MAIN_KEYBOARD,
        )


async def cancel_session_consolidation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_key = current_chat_key(update)
    PENDING_SESSION_CONSOLIDATION.pop(chat_key, None)

    await update.message.reply_text(
        "Guardado de sesión cancelado. La sesión sigue en el buffer por si quieres guardarla más tarde.",
        reply_markup=MAIN_KEYBOARD,
    )





async def version_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)

    repo_dir = Path(__file__).resolve().parent
    git_dir = repo_dir / ".git"

    branch = "unknown"
    commit = "unknown"

    try:
        head_path = git_dir / "HEAD"
        head_text = head_path.read_text(encoding="utf-8").strip()

        if head_text.startswith("ref: "):
            ref_name = head_text.replace("ref: ", "", 1).strip()
            branch = ref_name.split("/")[-1] or "unknown"

            ref_path = git_dir / ref_name
            if ref_path.exists():
                commit_hash = ref_path.read_text(encoding="utf-8").strip()
                if commit_hash:
                    commit = commit_hash[:12]
            else:
                packed_refs = git_dir / "packed-refs"
                if packed_refs.exists():
                    for line in packed_refs.read_text(encoding="utf-8").splitlines():
                        if not line or line.startswith("#") or line.startswith("^"):
                            continue
                        parts = line.split()
                        if len(parts) == 2 and parts[1] == ref_name:
                            commit = parts[0][:12]
                            break
        else:
            commit = head_text[:12]
            branch = "detached"

    except Exception:
        pass

    message = "\n".join(
        [
            "Versión eCoach:",
            "",
            f"- Rama Git: {branch}",
            f"- Commit: {commit}",
            f"- Zona horaria: {APP_TIMEZONE}",
            f"- Scheduler date-only: {PROACTIVE_SCHEDULER_HOUR:02d}:{PROACTIVE_SCHEDULER_MINUTE:02d}",
            "- Scheduler con hora: cada minuto",
        ]
    )

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )


async def diagnostics(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()
    ensure_followup_triggers_file()

    client_name = current_client_name()
    chat_id = get_remembered_chat_id()

    active_followups = load_followup_triggers()
    archived_followups = load_followup_archive()

    pending_followups = [
        item for item in active_followups
        if str(item.get("status", "pending")).lower() == "pending"
    ]

    date_only_pending = [
        item for item in pending_followups
        if not followup_has_time(item)
    ]

    timed_pending = [
        item for item in pending_followups
        if followup_has_time(item)
    ]

    buffer_path = session_buffer_file()
    buffer_size = buffer_path.stat().st_size if buffer_path.exists() else 0

    pending_preview_lines: list[str] = []

    for index, item in enumerate(pending_followups[:5], start=1):
        date_text = str(item.get("date", "")).strip()
        time_text = str(item.get("time", "")).strip()
        time_display = f" {time_text}" if time_text else ""
        message_template = str(item.get("message_template", "")).strip()

        if len(message_template) > 90:
            message_template = message_template[:87] + "..."

        pending_preview_lines.append(
            f"{index}. {date_text}{time_display} — {message_template}"
        )

    if not pending_preview_lines:
        pending_preview_lines.append("No hay seguimientos pendientes.")

    message = "\n".join(
        [
            "Diagnóstico eCoach:",
            "",
            f"- Cliente activo: {client_name}",
            f"- Chat ID recordado: {'sí' if chat_id is not None else 'no'}",
            "",
            "Seguimientos:",
            f"- Activos totales: {len(active_followups)}",
            f"- Pendientes totales: {len(pending_followups)}",
            f"- Pendientes date-only: {len(date_only_pending)}",
            f"- Pendientes con hora: {len(timed_pending)}",
            f"- Archivados: {len(archived_followups)}",
            "",
            "Pendientes próximos:",
            *pending_preview_lines,
            "",
            "Schedulers:",
            f"- Date-only diario: {PROACTIVE_SCHEDULER_HOUR:02d}:{PROACTIVE_SCHEDULER_MINUTE:02d} {APP_TIMEZONE}",
            "- Con hora: cada minuto",
            "",
            "Second Brain:",
            f"- Session buffer existe: {'sí' if buffer_path.exists() else 'no'}",
            f"- Session buffer tamaño: {buffer_size} bytes",
        ]
    )

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    message = """Ayuda rápida de eCoach

Puedes usarme de tres formas:

1. Preguntar
Ejemplo:
“¿Conviene amortizar hipoteca o invertir?”

2. Guardar memoria estable
Ejemplo:
“Mis gastos son 2.500€/mes.”
“Quiero mantener 6 meses de colchón.”

Eso no se guarda automáticamente en Second Brain.
Para guardarlo, pulsa:
💾 Guardar sesión

3. Crear seguimientos
Ejemplo:
“Recuérdame mañana revisar la respuesta del banco.”
“Recuérdame hoy a las 17:30 enviar el documento.”

Los seguimientos explícitos sí se guardan inmediatamente.

Botones:
👤 Quién soy — ver perfil.
🎯 Qué quiero — ver objetivos.
✅ Plan de acción — ver próximas acciones.
⏰ Seguimientos — ver recordatorios pendientes.
💾 Guardar sesión — consolidar memoria.

Comandos útiles:
- /version
- /diagnostics
- /followup_help
- /scheduler_status
- /followup_archive
- /followup_archive_stats
"""

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    welcome = """Hola. Soy eCoach Relaciones.

Puedo ayudarte a practicar agencia relacional entre sesiones de terapia.

No sustituyo a tu psicologa.
No decido por ti.
Te ayudo a pausar, separar hechos de historias, recordar tus valores y elegir el siguiente paso con mas claridad.

Idea central:

No persecucion ansiosa.
No juicio delegado.
Agencia relacional guiada.

Botones principales:

Quien soy
Ver el resumen de tu situacion y patron relacional.

Que quiero
Ver tus valores, criterios y objetivo relacional actual.

Plan de accion
Ver tu Mi Plan y el siguiente paso claro.

Seguimientos
Ver recordatorios y check-ins pendientes.

Guardar sesion
Revisar lo hablado y decidir que guardar.

Para probar la demo, puedes escribir:

Estoy saliendo con un hombre que me gusta.
Cuando estamos juntos, todo parece bien.
Pero entre citas es muy ambiguo.
A veces escribe con carino, pero otras desaparece.
Mi psicologa me dice que no persiga y que observe si hay coherencia, pero cuando me activo se me olvida todo.
"""
    await update.message.reply_text(welcome, reply_markup=MAIN_KEYBOARD)

async def show_quien_soy(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    await send_long_message(
        update,
        "Aquí está tu resumen actual de **Quién soy**:\n\n"
        + read_text_file(client_files()["quien_soy"]),
        reply_markup=MAIN_KEYBOARD,
    )


async def show_que_quiero(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    await send_long_message(
        update,
        "Aquí está tu resumen actual de **Qué quiero**:\n\n"
        + read_text_file(client_files()["que_quiero"]),
        reply_markup=MAIN_KEYBOARD,
    )


async def show_que_tengo_que_hacer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    await send_long_message(
        update,
        "Aquí está tu resumen actual de **Plan de acción**:\n\n"
        + read_text_file(client_files()["que_tengo_que_hacer"]),
        reply_markup=MAIN_KEYBOARD,
    )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    context.user_data.clear()

    message = load_required_app_text("cancel_message.md")

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )

async def reset_client(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()
    remember_chat_id(update)

    client_name = current_client_name()
    timestamp = datetime.now(ZoneInfo(APP_TIMEZONE)).strftime("%Y%m%d_%H%M%S")

    async with get_client_lock(client_name):
        client_dir = client_files()["quien_soy"].parent
        reset_backup_dir = client_dir / "backups" / f"reset_{timestamp}"
        reset_backup_dir.mkdir(parents=True, exist_ok=True)

        files_to_backup = [
            "quien_soy",
            "que_quiero",
            "que_tengo_que_hacer",
            "estilo_respuesta",
            "agent_observations",
            "historial_interacciones",
            "followup_triggers",
            "proactivity_log",
        ]

        for key in files_to_backup:
            source_path = client_files().get(key)
            if source_path and source_path.exists():
                shutil.copy2(source_path, reset_backup_dir / source_path.name)

        starter_files = {
            "quien_soy": INITIAL_CLIENT_FILES_DIR / "quien_soy.md",
            "que_quiero": INITIAL_CLIENT_FILES_DIR / "que_quiero.md",
            "que_tengo_que_hacer": INITIAL_CLIENT_FILES_DIR / "que_tengo_que_hacer.md",
            "estilo_respuesta": INITIAL_CLIENT_FILES_DIR / "estilo_respuesta.md",
        }

        for key, source_path in starter_files.items():
            if source_path.exists():
                shutil.copy2(source_path, client_files()[key])

        write_text_file(client_files()["agent_observations"], "# Agent observations\n\n")
        write_text_file(client_files()["historial_interacciones"], "# Historial de interacciones\n\n")
        write_text_file(client_files()["proactivity_log"], "# Proactivity log\n\n")
        save_followup_triggers([])

        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Client reset",
                    "",
                    f"Client: {client_name}",
                    f"Backup folder: {reset_backup_dir}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

    await update.message.reply_text(
        "Cliente reiniciado. He guardado una copia de seguridad y he restaurado los archivos iniciales.",
        reply_markup=MAIN_KEYBOARD,
    )



# --- SECOND BRAIN DESTRUCTIVE RESET GUARD v3 ---
# Product rule:
# Destructive Second Brain commands are state-management commands.
# They must be intercepted before router/orchestrator/LLM calls.

import unicodedata as _sb_guard_unicodedata
import unicodedata

PENDING_SECOND_BRAIN_RESET: dict[str, str] = {}


def _sb_guard_pending_file() -> Path:
    return client_files()["quien_soy"].parent / "pending_second_brain_reset.txt"


def _sb_guard_load_pending_target() -> str | None:
    path = _sb_guard_pending_file()
    if not path.exists():
        return None

    value = path.read_text(encoding="utf-8").strip()
    if value in {"quien_soy", "que_quiero", "que_tengo_que_hacer"}:
        return value

    return None


def _sb_guard_save_pending_target(target_key: str) -> None:
    path = _sb_guard_pending_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(target_key, encoding="utf-8")


def _sb_guard_clear_pending_target() -> None:
    path = _sb_guard_pending_file()
    if path.exists():
        path.unlink()


def _sb_guard_normalize(text: str) -> str:
    value = (text or "").strip().lower()
    value = _sb_guard_unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not _sb_guard_unicodedata.combining(ch))
    value = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in value)
    return " ".join(value.split())


def _sb_guard_target_from_text(user_text: str) -> str | None:
    t = _sb_guard_normalize(user_text)

    destructive_words = (
        "borra",
        "borrar",
        "elimina",
        "eliminar",
        "vacia",
        "vaciar",
        "resetea",
        "reset",
    )

    if not any(word in t for word in destructive_words):
        return None

    if "quien soy" in t or "perfil" in t:
        return "quien_soy"

    if "que quiero" in t or "objetivos" in t or "preferencias" in t:
        return "que_quiero"

    if "plan de accion" in t or "que tengo que hacer" in t:
        return "que_tengo_que_hacer"

    return None


def _sb_guard_is_confirmation(user_text: str) -> bool:
    t = _sb_guard_normalize(user_text)
    return t in {
        "si",
        "si confirmo",
        "confirmo",
        "adelante",
        "ok",
        "vale",
        "de acuerdo",
        "hazlo",
    }


def _sb_guard_is_cancel(user_text: str) -> bool:
    t = _sb_guard_normalize(user_text)
    return t in {
        "no",
        "cancelar",
        "cancela",
        "no confirmo",
        "dejalo",
        "deja",
    }


def _sb_guard_label(target_key: str) -> str:
    labels = {
        "quien_soy": "Quién soy",
        "que_quiero": "Qué quiero",
        "que_tengo_que_hacer": "Plan de acción",
    }
    return labels.get(target_key, target_key)


def _sb_guard_template(target_key: str) -> str:
    if target_key == "quien_soy":
        return """# Quién soy

## Perfil personal

* Resto de datos pendientes.

## Situación familiar

* Resto de datos pendientes.

## Ingresos y trabajo

* Resto de datos pendientes.

## Gastos y liquidez

* Resto de datos pendientes.

## Relaciones actual

* Resto de datos pendientes.

## Deudas e hipotecas

* Resto de datos pendientes.

## Perfil de riesgo

* Resto de datos pendientes.

## Datos pendientes

* Resto de datos pendientes.
"""

    if target_key == "que_quiero":
        return """# Qué quiero

## Seguridad financiera

* Resto de objetivos pendientes.

## Objetivos patrimoniales

* Resto de objetivos pendientes.

## Criterios de decisión

* Resto de criterios pendientes.

## Preferencias

* Resto de preferencias pendientes.

## Límites / cosas que evitar

* Resto de límites pendientes.
"""

    if target_key == "que_tengo_que_hacer":
        return """# Plan de acción

## Próxima acción activa

* Resto de tareas pendientes.

## Pendiente de respuesta externa

* Resto de respuestas pendientes.

## Documentos que faltan

* Resto de documentos pendientes.

## Decisiones pendientes

* Resto de decisiones pendientes.

## Backlog

* Resto de tareas pendientes.
"""

    raise ValueError(f"Unknown Second Brain target: {target_key}")


def second_brain_file_for_target(target_key: str) -> Path:
    files = client_files()

    if target_key == "quien_soy":
        return files["quien_soy"]

    if target_key == "que_quiero":
        return files["que_quiero"]

    if target_key == "que_tengo_que_hacer":
        return files["que_tengo_que_hacer"]

    raise ValueError(f"Unknown Second Brain target: {target_key}")


def repair_second_brain_content(target_key: str, content: str) -> str:
    """
    Conservative repair:
    - If file is empty, broken, or explicitly marked as empty by the client, reset template.
    - Otherwise preserve existing content for now.
    """
    raw = content or ""
    normalized = _sb_guard_normalize(raw)

    broken_markers = [
        "vacio por peticion del cliente",
        "pendiente de definir cuando el cliente lo indique",
        "qui?n soy",
        "situaci?n familiar",
        "qu? quiero",
        "plan de acci?n",
    ]

    if not raw.strip() or any(marker in normalized for marker in broken_markers):
        return _sb_guard_template(target_key)

    required_titles = {
        "quien_soy": "# Quién soy",
        "que_quiero": "# Qué quiero",
        "que_tengo_que_hacer": "# Plan de acción",
    }

    if required_titles[target_key] not in raw:
        return _sb_guard_template(target_key)

    return raw


def repair_second_brain_file(target_key: str) -> None:
    path = second_brain_file_for_target(target_key)
    path.parent.mkdir(parents=True, exist_ok=True)

    old = path.read_text(encoding="utf-8") if path.exists() else ""
    new = repair_second_brain_content(target_key, old)

    if new != old:
        path.write_text(new, encoding="utf-8")


def repair_all_second_brain_files() -> None:
    for target_key in ("quien_soy", "que_quiero", "que_tengo_que_hacer"):
        repair_second_brain_file(target_key)


def reset_second_brain_file(target_key: str) -> None:
    path = second_brain_file_for_target(target_key)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_sb_guard_template(target_key), encoding="utf-8")


async def handle_second_brain_reset_guard(update: Update, user_text: str) -> bool:
    """
    Returns True if the message was fully handled and should NOT go to router/orchestrator.
    """
    chat_key = current_chat_key(update)

    target_key = _sb_guard_load_pending_target()

    if target_key:
        label = _sb_guard_label(target_key)

        if _sb_guard_is_cancel(user_text):
            _sb_guard_clear_pending_target()
            PENDING_SECOND_BRAIN_RESET.pop(chat_key, None)
            await update.message.reply_text(
                "Perfecto. No he borrado nada.",
                reply_markup=MAIN_KEYBOARD,
            )
            return True

        if _sb_guard_is_confirmation(user_text):
            _sb_guard_clear_pending_target()
            PENDING_SECOND_BRAIN_RESET.pop(chat_key, None)

            if target_key == "que_tengo_que_hacer":
                await update.message.reply_text(
                    "El Plan de acción no se borra como una nota libre. "
                    "Lo gestiona el algoritmo. Puedes decirme qué ha cambiado "
                    "y te propondré cómo actualizarlo.",
                    reply_markup=MAIN_KEYBOARD,
                )
                return True

            reset_second_brain_file(target_key)

            await update.message.reply_text(
                f"He borrado los datos de ‘{label}’, manteniendo la estructura del archivo.",
                reply_markup=MAIN_KEYBOARD,
            )
            return True

    target_key = _sb_guard_target_from_text(user_text)

    if target_key is None:
        return False

    label = _sb_guard_label(target_key)

    if target_key == "que_tengo_que_hacer":
        await update.message.reply_text(
            "El Plan de acción no se edita ni se borra como una nota libre. "
            "Lo gestiona el algoritmo. Puedes decirme qué ha cambiado y te propondré "
            "cómo actualizarlo.",
            reply_markup=MAIN_KEYBOARD,
        )
        return True

    PENDING_SECOND_BRAIN_RESET[chat_key] = target_key
    _sb_guard_save_pending_target(target_key)

    await update.message.reply_text(
        f"Puedo borrar los datos de ‘{label}’, pero mantendré la estructura del archivo. "
        f"¿Confirmas que quieres dejar ‘{label}’ vacío?",
        reply_markup=MAIN_KEYBOARD,
    )
    return True

# --- END SECOND BRAIN DESTRUCTIVE RESET GUARD v3 ---




PUBLIC_ENRICHMENT_CALLBACK = "public_enrichment:last_portfolio"
PRIVATE_BANK_DATA_CALLBACK = "private_bank_data:last_portfolio"
ALTERNATIVES_CALLBACK = "alternatives:last_portfolio"
SELF_MANAGED_PATH_CALLBACK = "alternatives:self_managed"
DELEGATED_PATH_CALLBACK = "alternatives:delegated"
GUIDED_PATH_CALLBACK = "alternatives:guided"
COMPARE_CURRENT_PORTFOLIO_CALLBACK = "alternatives:compare_current_portfolio"
KEEP_SIMILAR_RISK_CALLBACK = "alternatives:keep_similar_risk"
REDUCE_RISK_CALLBACK = "alternatives:reduce_risk"
DEFINE_CONCRETE_ALTERNATIVE_CALLBACK = "alternatives:define_concrete_alternative"
PREPARE_PROVISIONAL_ALTERNATIVE_CALLBACK = "alternatives:prepare_provisional_alternative"
DEFINE_INITIAL_ALTERNATIVE_CALLBACK = "alternatives:define_initial_alternative"

REVIEW_BEFORE_EXECUTE_CALLBACK = "portfolio:review_before_execute"
SEARCH_PROVIDERS_CALLBACK = "providers:search"
PROVIDER_DIAPHANUM_CALLBACK = "provider_diaphanum"
PROVIDER_EBN_CALLBACK = "provider_ebn_banco"
CREATE_MI_PLAN_FOLLOWUP_CALLBACK = "create_mi_plan_followup_2026_06_17_1000"
DESIGN_MI_PLAN_CALLBACK = "design_mi_plan"
PROVIDER_MYINVESTOR_CALLBACK = "providers:myinvestor"
PROVIDER_RENTA4_CALLBACK = "providers:renta4"
PROVIDER_OPENBANK_CALLBACK = "providers:openbank"
PROVIDER_INDEXA_CALLBACK = "providers:indexa"
PROVIDER_INBESTME_CALLBACK = "providers:inbestme"
SEARCH_FUNDS_CALLBACK = "funds:search_available"
MISSING_DATA_CHECKLIST_CALLBACK = "data:missing_checklist"





def with_inline_transition_separator(text: str) -> str:
    clean = str(text).strip()
    if clean.startswith("\u2500\u2500\u2500"):
        return clean
    return "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n" + clean

async def clear_clicked_inline_keyboard(query) -> None:
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass

def public_enrichment_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f310 Completar con fuentes p\u00fablicas",
                    callback_data=PUBLIC_ENRICHMENT_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f3e6 Pedir datos privados al banco",
                    callback_data=PRIVATE_BANK_DATA_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f50e Explorar alternativas",
                    callback_data=ALTERNATIVES_CALLBACK,
                )
            ],
        ]
    )

PUBLIC_ENRICHMENT_RE = re.compile(
    r"^\s*(completa|completar|busca|buscar|actualiza|actualizar).*(fuentes\s+p[?u]blicas|informaci[o?]n\s+p[?u]blica|dfi|kid|ter|ocf|datos\s+de\s+los\s+fondos|fondos)\s*[\.\!\???]*$",
    re.IGNORECASE,
)


def normalize_public_enrichment_text(text: str) -> str:
    """
    Normalize Telegram text robustly:
    - lowercase
    - remove accents, including decomposed Unicode forms
    - keep plain searchable text
    """
    raw = (text or "").lower()
    decomposed = unicodedata.normalize("NFKD", raw)
    without_accents = "".join(
        char for char in decomposed
        if not unicodedata.combining(char)
    )
    return without_accents


def is_public_enrichment_intent(text: str) -> bool:
    normalized = normalize_public_enrichment_text(text)

    return (
        ("fuentes publicas" in normalized)
        or ("informacion publica" in normalized)
        or ("completa" in normalized and "public" in normalized)
        or ("completar" in normalized and "public" in normalized)
        or ("busca" in normalized and "public" in normalized)
        or ("buscar" in normalized and "public" in normalized)
        or ("actualiza" in normalized and "fondos" in normalized)
        or ("actualizar" in normalized and "fondos" in normalized)
        or ("kid" in normalized)
        or ("dfi" in normalized)
        or ("ter" in normalized)
        or ("ocf" in normalized)
    )


def save_last_portfolio_isins_for_public_enrichment(client_dir: Path, isins: list[str]) -> None:
    if not isins:
        return

    try:
        client_dir.mkdir(parents=True, exist_ok=True)
        out_path = client_dir / "last_portfolio_isins.json"
        payload = {
            "isins": isins,
            "source": "public_enrichment_detection",
        }
        out_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        return


def extract_recent_isins_for_public_enrichment(client_dir: Path) -> list[str]:
    candidates: list[str] = []
    pattern = re.compile(r"\b[A-Z]{2}[A-Z0-9]{10}\b")

    explicit_file = client_dir / "last_portfolio_isins.json"
    if explicit_file.exists():
        try:
            explicit_text = explicit_file.read_text(encoding="utf-8", errors="ignore")
            candidates.extend(pattern.findall(explicit_text))
        except Exception:
            pass

    likely_files = [
        client_dir / "last_portfolio_analysis.md",
        client_dir / "last_document_analysis.md",
        client_dir / "portfolio_analysis.md",
    ]

    uploaded_docs_dir = client_dir / "UploadedDocs"
    if uploaded_docs_dir.exists():
        likely_files.extend(uploaded_docs_dir.glob("**/*.txt"))
        likely_files.extend(uploaded_docs_dir.glob("**/*.md"))
        likely_files.extend(uploaded_docs_dir.glob("**/*.json"))
        likely_files.extend(uploaded_docs_dir.glob("**/*.csv"))

    for file_path in likely_files:
        try:
            if file_path.exists() and file_path.is_file():
                file_text = file_path.read_text(encoding="utf-8", errors="ignore")
                candidates.extend(pattern.findall(file_text))
        except Exception:
            continue

    seen = set()
    out = []
    for isin in candidates:
        if isin not in seen:
            seen.add(isin)
            out.append(isin)

    return out





def public_data_search(query: str) -> list[dict]:
    """
    Future public-source search interface.

    Current status:
    - Stub only.
    - Does not call the web.
    - Does not send private client data anywhere.
    - Later this can be connected to controlled public sources such as CNMV, fund manager pages, KID/DFI documents, etc.
    """
    return [
        {
            "query": query,
            "status": "stub",
            "source_name": "Public search not connected yet",
            "source_url": None,
            "confidence": "pending",
        }
    ]


def fund_lookup_by_isin(isin: str) -> dict:
    """
    Future fund lookup by public identifier.

    Privacy rule:
    - Input should be an ISIN or other public product identifier.
    - No client-specific amounts, positions, tax data, or personal information.
    """
    searches = [
        public_data_search(f"{isin} KID DFI"),
        public_data_search(f"{isin} TER OCF"),
        public_data_search(f"{isin} benchmark category"),
        public_data_search(f"{isin} equity fixed income exposure"),
    ]

    return {
        "isin": isin,
        "status": "stub",
        "fields": {
            "kid_dfi": "Pendiente de conectar fuente p?blica",
            "ter_ocf": "Pendiente de conectar fuente p?blica",
            "official_category": "Pendiente de conectar fuente p?blica",
            "benchmark": "Pendiente de conectar fuente p?blica",
            "composition": "Pendiente de conectar fuente p?blica",
            "manager_or_provider": "Pendiente de conectar fuente p?blica",
        },
        "sources": searches,
    }




def get_demo_public_fund_name_by_isin(isin: str) -> str | None:
    """
    Temporary verified demo mapping.

    Later this should be replaced by real public-source lookup.
    """
    names = {
        "ES0174398034": "Rural Mixto Internacional 30/50 FI",
        "ES0141986002": "Rural Futuro ISR FI Est\u00e1ndar",
        "ES0174406035": "Rural Mixto Internacional 25 FI",
        "ES0156832000": "Rural Mixto Internacional 15 FI",
    }
    return names.get(isin)


def build_public_enrichment_field_result(
    isin: str,
    field: str,
    label: str,
    value: str | None = None,
    status: str = "pending",
    source_name: str | None = None,
    source_url: str | None = None,
    confidence: str = "pending",
) -> dict:
    """
    Standard public enrichment result record.

    This schema is the bridge between future public-source search and the Telegram output.
    """
    return {
        "isin": isin,
        "field": field,
        "label": label,
        "value": value,
        "status": status,
        "source_name": source_name,
        "source_url": source_url,
        "confidence": confidence,
    }


def build_public_enrichment_results_for_isin(isin: str) -> list[dict]:
    """
    Current placeholder enrichment results for one ISIN.

    The fund name is filled from a small verified demo mapping.
    Later each field can be filled by public_data_search() / fund_lookup_by_isin().
    """
    fund_name = get_demo_public_fund_name_by_isin(isin)

    return [
        build_public_enrichment_field_result(
            isin,
            "official_fund_name",
            "Fondo",
            value=fund_name,
            status="found_demo_mapping" if fund_name else "pending",
            source_name="Demo public lookup",
            confidence="medium" if fund_name else "pending",
        ),
        build_public_enrichment_field_result(isin, "ter_ocf", "TER/OCF"),
        build_public_enrichment_field_result(isin, "official_category", "Categor\u00eda"),
        build_public_enrichment_field_result(isin, "benchmark", "Benchmark"),
        build_public_enrichment_field_result(isin, "composition", "Composici\u00f3n"),
    ]


def build_public_enrichment_results(isins: list[str]) -> list[dict]:
    results: list[dict] = []
    for isin in isins:
        results.extend(build_public_enrichment_results_for_isin(isin))
    return results


def build_public_enrichment_stub_table(isins: list[str]) -> str:
    """
    Build a source-aware placeholder table from structured enrichment records.
    """
    results = build_public_enrichment_results(isins)

    by_isin: dict[str, dict[str, dict]] = {}
    for item in results:
        by_isin.setdefault(item["isin"], {})[item["field"]] = item

    rows = []
    for isin in isins:
        fields = by_isin.get(isin, {})

        fund_name = fields.get("official_fund_name", {}).get("value") or "Pendiente"
        ter = fields.get("ter_ocf", {}).get("value") or "Pendiente"
        category = fields.get("official_category", {}).get("value") or "Pendiente"
        benchmark = fields.get("benchmark", {}).get("value") or "Pendiente"
        composition = fields.get("composition", {}).get("value") or "Pendiente"

        if fund_name != "Pendiente":
            source = "Identificaci\u00f3n inicial por ISIN"
            status = "Nombre identificado; resto pendiente"
        else:
            source = "Pendiente"
            status = "Fuentes p\u00fablicas a\u00fan no conectadas"

        rows.append(
            f"| {isin} | {fund_name} | {ter} | {category} | {benchmark} | {composition} | {source} | {status} |"
        )

    table = (
        "| ISIN | Fondo | TER/OCF | Categor\u00eda | Benchmark | Composici\u00f3n | Fuente | Estado |\n"
        "|---|---|---:|---|---|---|---|---|\n"
        + "\n".join(rows)
    )

    return table

def store_public_enrichment_sources(client_dir: Path, enrichment_results: list[dict]) -> None:
    """
    Future storage point for public-source results.

    Current status:
    - Stub only.
    - Intentionally does not store anything yet.
    """
    return

def build_public_enrichment_message(isins: list[str]) -> str:
    isin_lines = "\n".join(f"- {isin}" for isin in isins)

    return (
        "Voy a intentar completar informaci\u00f3n p\u00fablica de los fondos.\n\n"
        "Fondos / ISINs detectados:\n"
        f"{isin_lines}\n\n"
        "Buscar\u00e9 o intentar\u00e9 completar:\n"
        "- KID / DFI actualizado.\n"
        "- Ficha actualizada.\n"
        "- TER / OCF p\u00fablico.\n"
        "- Categor\u00eda oficial.\n"
        "- Benchmark oficial si est\u00e1 publicado.\n"
        "- Composici\u00f3n actual: renta variable, renta fija, liquidez u otros.\n"
        "- Datos p\u00fablicos de la gestora/fondo.\n\n"
        "Importante:\n"
        "- No buscar\u00e9 datos privados en la web.\n"
        "- No enviar\u00e9 informaci\u00f3n personal del cliente a fuentes p\u00fablicas.\n"
        "- Solo usar\u00e9 identificadores p\u00fablicos del fondo, como el ISIN.\n\n"
        "De momento he preparado la lista de b\u00fasqueda p\u00fablica.\n"
        "En la siguiente versi\u00f3n consultar\u00e9 fuentes p\u00fablicas controladas y le mostrar\u00e9 qu\u00e9 datos vienen de cada fuente.\n\n"
        "Vista previa de enriquecimiento p\u00fablico:\n\n"
        f"{build_public_enrichment_stub_table(isins)}\n\n"
        "Puede elegir ahora:\n"
        "1. Completar datos p\u00fablicos de estos fondos.\n"
        "2. Comparar con alternativas de menor coste.\n"
        "3. Preparar un mensaje para el banco pidiendo los datos privados que faltan."
    )

async def handle_public_enrichment_command(update, context):
    user_text = ""
    try:
        user_text = (update.message.text or "").strip()
    except Exception:
        user_text = ""

    if user_text and await try_handle_ecoach_control_message(update, context, user_text):
        return

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    isins = extract_recent_isins_for_public_enrichment(client_dir)
    save_last_portfolio_isins_for_public_enrichment(client_dir, isins)

    if not isins:
        await update.message.reply_text(
            "Puedo intentar completar informaci\u00f3n p\u00fablica, pero ahora mismo no encuentro los ISINs de la \u00faltima cartera analizada.\n\n"
            "Por favor, sube de nuevo los documentos o pulsa primero \u201c\U0001f4c4 Analizar documentos\u201d.",
            reply_markup=public_enrichment_keyboard(),
        )
        return

    msg = build_public_enrichment_message(isins)


    await update.message.reply_text(msg, reply_markup=public_enrichment_keyboard())



def build_private_bank_data_request_message(isins: list[str]) -> str:
    fund_lines = []

    for isin in isins:
        fund_name = None
        try:
            fund_name = get_demo_public_fund_name_by_isin(isin)
        except Exception:
            fund_name = None

        if fund_name:
            fund_lines.append(f"- {fund_name} ({isin})")
        else:
            fund_lines.append(f"- {isin}")

    funds_text = "\n".join(fund_lines) if fund_lines else "- [indicar fondos / ISINs]"

    return (
        "Mensaje preparado para enviar al banco:\n\n"
        "Buenos d\u00edas,\n\n"
        "Estoy revisando mi cartera y necesito que me faciliten, para cada uno de los siguientes fondos/productos:\n\n"
        f"{funds_text}\n\n"
        "La siguiente informaci\u00f3n:\n\n"
        "1. Fecha de compra o suscripci\u00f3n de cada posici\u00f3n.\n"
        "2. Precio o valor de adquisici\u00f3n.\n"
        "3. Valor actual de mercado.\n"
        "4. Plusval\u00eda o minusval\u00eda latente.\n"
        "5. Comisiones reales aplicadas a mi cuenta, incluyendo custodia, gesti\u00f3n, asesoramiento, intermediaci\u00f3n o cualquier otro coste.\n"
        "6. Retrocesiones o incentivos percibidos por la entidad, si existen.\n"
        "7. Informe ex-post de costes y gastos del \u00faltimo ejercicio disponible.\n"
        "8. Fiscalidad estimada en caso de venta o traspaso, si pueden facilitarla.\n\n"
        "Les agradecer\u00eda que la informaci\u00f3n viniera desglosada por producto y en formato claro, preferiblemente PDF o Excel.\n\n"
        "Muchas gracias.\n\n"
        "---\n\n"
        "Nota del eCoach:\n"
        "Si el banco no responde, responde de forma incompleta o no entrega los documentos, puedes subir aqu\u00ed lo que tengas. "
        "Revisar\u00e9 los archivos con confidencialidad e intentar\u00e9 extraer de ellos las respuestas: costes, fechas, valores, "
        "plusval\u00edas, minusval\u00edas, retrocesiones e informe de costes si aparece en la documentaci\u00f3n."
    )








def define_alternative_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\u27a1\ufe0f Continuar con datos disponibles",
                    callback_data=PREPARE_PROVISIONAL_ALTERNATIVE_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f4cc Ver datos que faltan",
                    callback_data=MISSING_DATA_CHECKLIST_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f4ca Comparar con mi cartera actual",
                    callback_data=COMPARE_CURRENT_PORTFOLIO_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f50e Explorar alternativas",
                    callback_data=ALTERNATIVES_CALLBACK,
                )
            ],
        ]
    )

def risk_intention_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\u2248 Mantener riesgo parecido",
                    callback_data=KEEP_SIMILAR_RISK_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\u2193 Reducir riesgo",
                    callback_data=REDUCE_RISK_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f9f1 Definir alternativa concreta",
                    callback_data=DEFINE_CONCRETE_ALTERNATIVE_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f50e Explorar alternativas",
                    callback_data=ALTERNATIVES_CALLBACK,
                )
            ],
        ]
    )

def compare_current_portfolio_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f4ca Comparar con mi cartera actual",
                    callback_data=COMPARE_CURRENT_PORTFOLIO_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f310 Completar con fuentes p\u00fablicas",
                    callback_data=PUBLIC_ENRICHMENT_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f50e Explorar alternativas",
                    callback_data=ALTERNATIVES_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f3e6 Pedir datos privados al banco",
                    callback_data=PRIVATE_BANK_DATA_CALLBACK,
                )
            ],
        ]
    )



def clean_document_diagnosis_for_demo(message: str) -> str:
    """Clean uploaded-document diagnosis for Saturday demo."""
    if not message:
        return message

    cleaned = message.strip()

    hard_cut_markers = [
        "**Datos que puedo intentar completar con fuentes públicas",
        "**Datos privados que debe aportar",
        "**Mensaje para el banco:**",
        "Mensaje para el banco:",
        "**Siguiente decisión:**",
        "Siguiente decisión:",
        "Ahora te toca decidir",
        "Dada tu cartera y lo que",
        "Soy independiente del proveedor",
        "Próxima acción sugerida:",
        "**Próxima acción sugerida:**",
        "**¿Qué camino quieres explorar ahora?**",
        "Dime cuál de estos caminos te interesa más:",
        "- 1) Mantener la cartera actual",
        "1. Mantener la cartera actual pero entenderla mejor.",
    ]

    positions = [cleaned.find(m) for m in hard_cut_markers if cleaned.find(m) != -1]
    if positions:
        cleaned = cleaned[:min(positions)].rstrip()

    cleaned = cleaned.replace(
        "Los fondos mixtos 15, 25 y 30/50 tienen costes entre 1,0% y 1,5% de gestión, no extremos pero tampoco los más baratos del mercado.",
        "Los fondos mixtos 15, 25 y 30/50 tienen costes entre 1,0% y 1,5% de gestión. Para una cartera conservadora, la cartera parece cara y claramente mejorable si se simplifica."
    )

    cleaned = cleaned.replace(
        "Coste ponderado cercano al 1,4%, que no es excesivo pero puede ser mejorable si se simplifica.",
        "Coste ponderado cercano al 1,4%. Para una cartera conservadora, parece un coste alto y claramente mejorable si se simplifica."
    )

    final_choice = """Ahora te toca decidir qué camino quieres explorar. Hay tres opciones principales:

1. **Cartera autogestionada**: tú eliges y ejecutas directamente los fondos.
2. **Gestión delegada**: delegas la cartera en un proveedor.
3. **Cartera guiada por eCoach**: tú mantienes el control, pero yo te ayudo a entender, comparar y avanzar paso a paso.

¿Qué camino quieres explorar primero?"""

    return (cleaned.rstrip() + chr(10) + chr(10) + final_choice).strip()



def alternatives_path_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f9ed Cartera autogestionada",
                    callback_data=SELF_MANAGED_PATH_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f6e1\ufe0f Gesti\u00f3n delegada",
                    callback_data=DELEGATED_PATH_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f9e9 Cartera guiada",
                    callback_data=GUIDED_PATH_CALLBACK,
                )
            ],
        ]
    )

def build_alternatives_path_message() -> str:
    return (
        "Podemos explorar alternativas, pero primero conviene elegir el tipo de camino.\n\n"
        "1. Cartera autogestionada\n"
        "Menor coste, mayor soberan\u00eda y m\u00e1s control. Requiere m\u00e1s implicaci\u00f3n, "
        "pero el eCoach te acompa\u00f1a paso a paso: comparaci\u00f3n, decisiones, instrucciones, "
        "revisiones y rebalanceos.\n\n"
        "2. Gesti\u00f3n delegada\n"
        "Mayor simplicidad y menos trabajo operativo. Suele tener m\u00e1s coste y menos soberan\u00eda, "
        "pero el eCoach te ayuda a revisar si compensa, qu\u00e9 preguntas hacer y qu\u00e9 costes vigilar.\n\n"
        "3. Cartera guiada\n"
        "Punto intermedio. Mantienes la decisi\u00f3n final, pero seguimos una estructura clara: "
        "perfil de riesgo, cartera modelo, reglas de rebalanceo y revisiones peri\u00f3dicas.\n\n"
        "Pr\u00f3xima acci\u00f3n sugerida:\n"
        "Elige uno de estos tres caminos. Despu\u00e9s podremos comparar alternativas manteniendo el riesgo parecido, "
        "reduciendo costes o buscando m\u00e1s simplicidad."
    )




def save_selected_alternative_path(client_dir: Path, path_key: str, path_label: str) -> None:
    try:
        client_dir.mkdir(parents=True, exist_ok=True)
        out_path = client_dir / "selected_alternative_path.json"
        payload = {
            "path_key": path_key,
            "path_label": path_label,
        }
        out_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        return


def load_selected_alternative_path(client_dir: Path) -> dict | None:
    try:
        in_path = client_dir / "selected_alternative_path.json"
        if not in_path.exists():
            return None
        return json.loads(in_path.read_text(encoding="utf-8"))
    except Exception:
        return None

def build_self_managed_path_message() -> str:
    return (
        "Camino elegido: Cartera autogestionada\n\n"
        "Objetivo:\n"
        "Reducir costes y aumentar soberan\u00eda, manteniendo el control de las decisiones. "
        "T\u00fa decides y ejecutas; el eCoach te ayuda a comparar, entender riesgos, preparar instrucciones, "
        "revisar desviaciones y hacer rebalanceos.\n\n"
        "Ventaja principal:\n"
        "- Menor coste y mayor control.\n\n"
        "Coste emocional/pr\u00e1ctico:\n"
        "- Requiere m\u00e1s implicaci\u00f3n y seguir un proceso ordenado.\n\n"
        "Pr\u00f3xima acci\u00f3n sugerida:\n"
        "Confirmar si quieres mantener un riesgo parecido al actual o reducirlo. Despu\u00e9s puedo construir una comparaci\u00f3n inicial."
    )


def build_delegated_path_message() -> str:
    return (
        "Camino elegido: Gesti\u00f3n delegada\n\n"
        "Objetivo:\n"
        "Reducir trabajo operativo y delegar la gesti\u00f3n, aceptando normalmente m\u00e1s coste y menos control directo. "
        "El eCoach no decide por ti, pero te ayuda a revisar si la delegaci\u00f3n compensa: costes, riesgos, transparencia, "
        "incentivos y calidad de la explicaci\u00f3n.\n\n"
        "Ventaja principal:\n"
        "- M\u00e1s simplicidad y menos carga operativa.\n\n"
        "Coste emocional/pr\u00e1ctico:\n"
        "- Menos soberan\u00eda y posible mayor coste.\n\n"
        "Pr\u00f3xima acci\u00f3n sugerida:\n"
        "Pedir o subir la propuesta de gesti\u00f3n delegada para revisar costes, perfil de riesgo, incentivos y condiciones."
    )


def build_guided_path_message() -> str:
    return """Camino elegido: cartera guiada

Esto significa: no autogestionada sola, no delegada a ciegas. Soberanía acompañada.

Ahora elegimos una plataforma o proveedor desde el que construir una propuesta concreta.

Hay tres opciones razonables para estudiar primero:

1. **MyInvestor**: banco/plataforma con muchos fondos y costes bajos. Ventaja: es práctico y amplio. Riesgo: al ser banco, también puede ofrecer otros productos o tener una lógica comercial más activa.

2. **Diaphanum**: asesor financiero tradicional e independiente, más premium. Ventaja: más acompañamiento humano. Riesgo: puede ser menos simple o más caro para una cartera pequeña.

3. **EBN Banco**: banco especializado con arquitectura abierta y comisión explícita. Ventaja: más transparencia de costes. Riesgo: hay que entender bien la comisión explícita y compararla con el ahorro en fondos.

¿Qué proveedor quieres estudiar primero?"""


async def handle_guided_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"
    save_selected_alternative_path(client_dir, "guided", "Cartera guiada")

    await query.message.reply_text(build_guided_path_message(), reply_markup=provider_comparison_keyboard())

async def handle_alternatives_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    msg = build_alternatives_path_message()

    await query.message.reply_text(with_inline_transition_separator(msg), reply_markup=alternatives_path_keyboard())

async def handle_private_bank_data_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    isins = extract_recent_isins_for_public_enrichment(client_dir)

    msg = build_private_bank_data_request_message(isins)

    await query.message.reply_text(with_inline_transition_separator(msg), reply_markup=public_enrichment_keyboard())

async def handle_public_enrichment_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    isins = extract_recent_isins_for_public_enrichment(client_dir)
    save_last_portfolio_isins_for_public_enrichment(client_dir, isins)

    if not isins:
        await query.message.reply_text(
            "Puedo intentar completar informaci\u00f3n p\u00fablica, pero ahora mismo no encuentro los ISINs de la \u00faltima cartera analizada.\n\n"
            "Por favor, sube de nuevo los documentos o pulsa primero \u201c\U0001f4c4 Analizar documentos\u201d.",
            reply_markup=public_enrichment_keyboard(),
        )
        return

    msg = build_public_enrichment_message(isins)


    await query.message.reply_text(with_inline_transition_separator(msg), reply_markup=public_enrichment_keyboard())


def detect_relationship_initial_message(user_text: str) -> bool:
    text = user_text.lower()

    relationship_markers = [
        "hombre que me gusta",
        "psicóloga",
        "psicologa",
        "no persiga",
        "no perseguir",
        "entre citas",
        "ambiguo",
        "desaparece",
        "se me olvida todo",
        "cuando me activo",
        "relación romántica",
        "relacion romantica",
    ]

    hits = sum(1 for marker in relationship_markers if marker in text)
    return hits >= 2


async def handle_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    remember_chat_id(update)

    client_name = current_client_name()
    user_text = update.message.text.strip()

    if detect_relationship_initial_message(user_text):
        await reply_initial_discovery_with_llm(update, user_text)
        return


    # ECOACH PROJECT CONTROL GUARD:
    # Structural project-board messages must win before old portfolio/public-enrichment flows.
    if await try_handle_ecoach_control_message(update, context, user_text):
        return

    # HARD GUARD: public enrichment intent must never reach the generic orchestrator.
    if is_public_enrichment_intent(user_text):
        await handle_public_enrichment_command(update, context)
        return

    repair_all_second_brain_files()

    if await handle_second_brain_reset_guard(update, user_text):
        return

    thinking_text = load_required_app_text("thinking_message.md")

    thinking_message = await update.message.reply_text(
        thinking_text,
        reply_markup=MAIN_KEYBOARD,
    )

    try:
        source_label = load_required_app_text("free_text_source_label.md")

        async with get_client_lock(client_name):
            orchestrator_result = await asyncio.wait_for(
                asyncio.to_thread(
                    run_orchestrator_for_client,
                    client_name,
                    user_text,
                    source_label,
                ),
                timeout=240,
            )

    except asyncio.TimeoutError:
        timeout_message = (
            "La respuesta está tardando demasiado y he cancelado este intento. "
            "Por favor, pruebe de nuevo con una pregunta más corta o inténtelo otra vez."
        )

        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Orchestrator timeout",
                    "",
                    f"Client: {client_name}",
                    f"Input preview: {compact_log_preview(user_text, 600)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        try:
            await thinking_message.edit_text(
                timeout_message,
                reply_markup=MAIN_KEYBOARD,
            )
        except Exception as edit_error:
            append_text_file(
                wealth_log_file(),
                "\n".join(
                    [
                        f"## {format_app_minute()}",
                        "",
                        "### Timeout message edit failed",
                        "",
                        f"Client: {client_name}",
                        f"Edit error: {edit_error}",
                        "",
                        "---",
                        "",
                    ]
                ),
            )

            await update.message.reply_text(
                timeout_message,
                reply_markup=MAIN_KEYBOARD,
            )

        return

    except Exception as error:
        try:
            await thinking_message.delete()
        except Exception:
            pass

        error_template = load_required_app_text("processing_error_message.md")

        error_message = render_template(
            error_template,
            {
                "error": str(error),
            },
        )

        await update.message.reply_text(
            error_message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        await thinking_message.delete()
    except Exception as delete_error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Thinking message delete failed",
                    "",
                    f"Client: {client_name}",
                    f"Delete error: {delete_error}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

    try:
        await send_long_message(
            update,
            orchestrator_result,
            reply_markup=MAIN_KEYBOARD,
        )

        append_session_buffer(user_text, orchestrator_result)

    except Exception as send_error:
        append_text_file(
            wealth_log_file(),
            "\n".join(
                [
                    f"## {format_app_minute()}",
                    "",
                    "### Final answer send failed",
                    "",
                    f"Client: {client_name}",
                    f"Send error: {send_error}",
                    f"Answer preview: {compact_log_preview(orchestrator_result, 1200)}",
                    "",
                    "---",
                    "",
                ]
            ),
        )

        try:
            await update.message.reply_text(
                "He preparado la respuesta, pero Telegram ha fallado al enviarla. "
                "La respuesta ha quedado guardada en el historial interno. "
                "Por favor, inténtelo de nuevo o pida una respuesta más corta.",
                reply_markup=MAIN_KEYBOARD,
            )
        except Exception as fallback_error:
            append_text_file(
                wealth_log_file(),
                "\n".join(
                    [
                        f"## {format_app_minute()}",
                        "",
                        "### Final answer fallback send failed",
                        "",
                        f"Client: {client_name}",
                        f"Fallback error: {fallback_error}",
                        "",
                        "---",
                        "",
                    ]
                ),
            )


async def followup_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)

    message = """Ayuda de seguimientos

Crear recordatorios:
Puedes escribir en lenguaje natural:

“Recuérdame mañana revisar la respuesta del banco.”
“Recuérdame el viernes llamar al asesor.”
“Recuérdame hoy a las 17:30 enviar el documento.”

Regla importante:
- Si das fecha sin hora, el recordatorio queda como date-only y se enviará por la mañana.
- Si das fecha y hora, se enviará aproximadamente a esa hora.
- Si solo mencionas que algo pasará en el futuro, pero no pides recordatorio, te preguntaré antes de guardarlo.

Ver:
- /followups — ver seguimientos pendientes.
- /due_followups — ver seguimientos vencidos date-only.
- /all_followups — ver todos los seguimientos activos.
- /followup_archive — ver últimos seguimientos archivados.
- /followup_archive_stats — estadísticas del archivo.
- /scheduler_status — estado del scheduler.

Gestionar:
- /followup_done 1 — marcar el seguimiento pendiente número 1 como completado.
- /delete_followup 1 — borrar el seguimiento pendiente número 1.
- /followup_snooze 1 mañana — posponer el seguimiento número 1.
- /run_followups — enviar manualmente los seguimientos date-only vencidos.
- /followup_cleanup — archivar seguimientos no pendientes que hubieran quedado en activos.
"""

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )


async def show_followups(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    pending = get_pending_followup_triggers()

    if not pending:
        await update.message.reply_text(
            load_required_app_text("followups_empty_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    header = load_required_app_text("followups_list_header.md")
    item_template = load_required_app_text("followups_item_template.md")

    default_date = load_required_app_text("followups_default_date.md")
    default_type = load_required_app_text("followups_default_type.md")
    default_message = load_required_app_text("followups_default_message.md")
    default_status = load_required_app_text("followups_default_status.md")

    items: list[str] = []

    for index, trigger in enumerate(pending, start=1):
        item = render_template(
            item_template,
            {
                "index": str(index),
                "date": str(trigger.get("date", default_date)),
                "time_display": f" {str(trigger.get('time', '')).strip()}" if str(trigger.get("time", "")).strip() else "",
                "type": str(trigger.get("type", default_type)),
                "message_template": str(trigger.get("message_template", default_message)),
                "status": str(trigger.get("status", default_status)),
            },
        )

        items.append(item)

    await send_long_message(
        update,
        header + "\n\n" + "\n\n".join(items),
        reply_markup=MAIN_KEYBOARD,
    )

async def show_due_followups(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    triggers = load_followup_triggers()
    today = today_app()

    due: list[dict] = []

    remaining_triggers: list[dict] = []

    for trigger in triggers:
        status = str(trigger.get("status", "pending")).lower()

        if status != "pending":
            remaining_triggers.append(trigger)
            continue

        trigger_date = parse_followup_date(str(trigger.get("date", "")))

        if trigger_date is None:
            continue

        if trigger_date <= today:
            due.append(trigger)

    if not due:
        await update.message.reply_text(
            "No hay seguimientos vencidos para enviar ahora.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    header = "Seguimientos vencidos:"
    item_template = load_required_app_text("followups_item_template.md")

    default_date = load_required_app_text("followups_default_date.md")
    default_type = load_required_app_text("followups_default_type.md")
    default_message = load_required_app_text("followups_default_message.md")
    default_status = load_required_app_text("followups_default_status.md")

    items: list[str] = []

    for index, trigger in enumerate(due, start=1):
        item = render_template(
            item_template,
            {
                "index": str(index),
                "date": str(trigger.get("date", default_date)),
                "time_display": f" {str(trigger.get('time', '')).strip()}" if str(trigger.get("time", "")).strip() else "",
                "type": str(trigger.get("type", default_type)),
                "message_template": str(trigger.get("message_template", default_message)),
                "status": str(trigger.get("status", default_status)),
            },
        )

        items.append(item)

    await send_long_message(
        update,
        header + "\n\n" + "\n\n".join(items),
        reply_markup=MAIN_KEYBOARD,
    )


async def scheduler_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    chat_id = get_remembered_chat_id()
    triggers = load_followup_triggers()

    today = today_app()
    now = now_app()

    pending_count = 0
    date_only_pending_count = 0
    timed_pending_count = 0
    date_only_due_count = 0
    timed_due_count = 0
    invalid_date_count = 0
    invalid_time_count = 0

    for trigger in triggers:
        status = str(trigger.get("status", "pending")).lower()

        if status != "pending":
            continue

        pending_count += 1

        trigger_date = parse_followup_date(str(trigger.get("date", "")))
        has_time = followup_has_time(trigger)

        if trigger_date is None:
            invalid_date_count += 1
            continue

        if has_time:
            timed_pending_count += 1
            due_at = followup_due_datetime(trigger)

            if due_at is None:
                invalid_time_count += 1
                continue

            if due_at <= now:
                timed_due_count += 1

        else:
            date_only_pending_count += 1

            if trigger_date <= today:
                date_only_due_count += 1

    chat_status = "sí" if chat_id is not None else "no"

    message = "\n".join(
        [
            "Estado del scheduler de seguimientos:",
            "",
            "Dos tipos de recordatorio:",
            "- Date-only: tiene fecha, pero no hora. Se envía por la mañana.",
            "- Con hora: tiene fecha y hora. Se comprueba cada minuto.",
            "",
            f"- Scheduler date-only activo: sí",
            f"- Hora diaria date-only: {PROACTIVE_SCHEDULER_HOUR:02d}:{PROACTIVE_SCHEDULER_MINUTE:02d} {APP_TIMEZONE}",
            f"- Scheduler con hora activo: sí, cada minuto",
            f"- Chat ID recordado: {chat_status}",
            "",
            f"- Seguimientos pendientes totales: {pending_count}",
            f"- Pendientes date-only: {date_only_pending_count}",
            f"- Pendientes con hora: {timed_pending_count}",
            f"- Vencidos date-only: {date_only_due_count}",
            f"- Vencidos con hora: {timed_due_count}",
            f"- Fecha inválida: {invalid_date_count}",
            f"- Hora inválida: {invalid_time_count}",
        ]
    )

    await update.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )


async def show_all_followups(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    triggers = load_followup_triggers()

    if not triggers:
        await update.message.reply_text(
            load_required_app_text("all_followups_empty_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    header = load_required_app_text("all_followups_list_header.md")
    item_template = load_required_app_text("all_followups_item_template.md")

    default_date = load_required_app_text("followups_default_date.md")
    default_type = load_required_app_text("followups_default_type.md")
    default_message = load_required_app_text("followups_default_message.md")
    default_status = load_required_app_text("followups_default_status.md")
    default_sent_at = load_required_app_text("all_followups_default_sent_at.md")

    items: list[str] = []

    for index, trigger in enumerate(triggers, start=1):
        item = render_template(
            item_template,
            {
                "index": str(index),
                "date": str(trigger.get("date", default_date)),
                "time_display": f" {str(trigger.get('time', '')).strip()}" if str(trigger.get("time", "")).strip() else "",
                "type": str(trigger.get("type", default_type)),
                "status": str(trigger.get("status", default_status)),
                "message_template": str(trigger.get("message_template", default_message)),
                "created_at": str(trigger.get("created_at", "")),
                "sent_at": str(trigger.get("sent_at", default_sent_at)),
                "reason": str(trigger.get("reason", "")),
            },
        )

        items.append(item)

    await send_long_message(
        update,
        header + "\n\n" + "\n\n".join(items),
        reply_markup=MAIN_KEYBOARD,
    )

async def mark_followup_done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    if not context.args:
        usage_template = load_required_app_text("followup_command_usage_message.md")
        message = render_template(
            usage_template,
            {
                "command": "/followup_done",
            },
        )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        selected_index = int(context.args[0])
    except ValueError:
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    triggers = load_followup_triggers()

    pending_positions = [
        index for index, trigger in enumerate(triggers)
        if str(trigger.get("status", "pending")).lower() == "pending"
    ]

    if selected_index < 1 or selected_index > len(pending_positions):
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    trigger_position = pending_positions[selected_index - 1]

    completed_trigger = triggers[trigger_position]
    completed_trigger["status"] = "completed"
    completed_trigger["completed_at"] = format_app_datetime()

    append_proactivity_log(
        event_type="followup_completed",
        trigger=completed_trigger,
        notes="Marked completed manually via /followup_done.",
    )

    archive_followup_trigger(completed_trigger, "completed")
    del triggers[trigger_position]
    save_followup_triggers(triggers)

    await update.message.reply_text(
        load_required_app_text("followup_done_success_message.md"),
        reply_markup=MAIN_KEYBOARD,
    )

async def delete_followup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    if not context.args:
        usage_template = load_required_app_text("followup_command_usage_message.md")
        message = render_template(
            usage_template,
            {
                "command": "/delete_followup",
            },
        )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        selected_index = int(context.args[0])
    except ValueError:
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    triggers = load_followup_triggers()

    pending_positions = [
        index for index, trigger in enumerate(triggers)
        if str(trigger.get("status", "pending")).lower() == "pending"
    ]

    if selected_index < 1 or selected_index > len(pending_positions):
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    trigger_position = pending_positions[selected_index - 1]
    deleted_trigger = triggers[trigger_position]

    append_proactivity_log(
        event_type="followup_deleted",
        trigger=deleted_trigger,
        notes="Deleted manually via /delete_followup.",
    )

    deleted_trigger["status"] = "deleted"
    deleted_trigger["deleted_at"] = format_app_datetime()
    archive_followup_trigger(deleted_trigger, "deleted")

    del triggers[trigger_position]

    save_followup_triggers(triggers)

    await update.message.reply_text(
        load_required_app_text("followup_delete_success_message.md"),
        reply_markup=MAIN_KEYBOARD,
    )

async def snooze_followup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_followup_triggers_file()

    if len(context.args) < 2:
        await update.message.reply_text(
            "Uso: /followup_snooze 1 <nueva fecha o descripción de fecha>.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        selected_index = int(context.args[0])
    except ValueError:
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    new_date = parse_snooze_date(" ".join(context.args[1:]))

    if new_date is None:
        await update.message.reply_text(
            "No he entendido la nueva fecha. Escríbela con más contexto.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    triggers = load_followup_triggers()

    pending_positions = [
        index for index, trigger in enumerate(triggers)
        if str(trigger.get("status", "pending")).lower() == "pending"
    ]

    if selected_index < 1 or selected_index > len(pending_positions):
        await update.message.reply_text(
            load_required_app_text("followup_invalid_index_message.md"),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    trigger_position = pending_positions[selected_index - 1]
    snoozed_trigger = triggers[trigger_position]

    old_date = str(snoozed_trigger.get("date", ""))
    new_date_text = new_date.strftime("%Y-%m-%d")

    snoozed_trigger["date"] = new_date_text
    snoozed_trigger["status"] = "pending"
    snoozed_trigger["snoozed_at"] = format_app_datetime()
    snoozed_trigger["previous_date"] = old_date

    append_proactivity_log(
        event_type="followup_snoozed",
        trigger=snoozed_trigger,
        notes=f"Snoozed manually via /followup_snooze from {old_date} to {new_date_text}.",
    )

    save_followup_triggers(triggers)

    await update.message.reply_text(
        f"Seguimiento pospuesto a {new_date_text}.",
        reply_markup=MAIN_KEYBOARD,
    )


async def send_due_followups(reply_text_func, notes: str) -> dict:
    """
    Send all pending seguimientos whose date is today or earlier.

    This is reusable:
    - /run_followups can call it manually.
    - a future automatic scheduler can call it daily.
    """
    ensure_followup_triggers_file()

    triggers = load_followup_triggers()
    today = today_app()

    sent_count = 0
    has_invalid_dates = False
    remaining_triggers: list[dict] = []

    for trigger in triggers:
        status = str(trigger.get("status", "pending")).lower()

        if status != "pending":
            remaining_triggers.append(trigger)
            continue

        if followup_has_time(trigger):
            remaining_triggers.append(trigger)
            continue

        trigger_date = parse_followup_date(str(trigger.get("date", "")))

        if trigger_date is None:
            has_invalid_dates = True
            remaining_triggers.append(trigger)
            continue

        if trigger_date > today:
            remaining_triggers.append(trigger)
            continue

        message_template = str(trigger.get("message_template", "")).strip()

        if not message_template:
            remaining_triggers.append(trigger)
            continue

        await reply_text_func(message_template)

        trigger["status"] = "sent"
        trigger["sent_at"] = format_app_datetime()

        append_proactivity_log(
            event_type="followup_sent",
            trigger=trigger,
            notes=notes,
        )

        archive_followup_trigger(trigger, "sent")
        sent_count += 1

    save_followup_triggers(remaining_triggers)

    return {
        "sent_count": sent_count,
        "has_invalid_dates": has_invalid_dates,
    }


async def send_due_timed_followups(reply_text_func, notes: str) -> dict:
    """Send pending seguimientos with date+time whose datetime is due."""
    ensure_followup_triggers_file()

    triggers = load_followup_triggers()
    now = now_app()

    sent_count = 0
    has_invalid_datetimes = False
    remaining_triggers: list[dict] = []

    for trigger in triggers:
        status = str(trigger.get("status", "pending")).lower()

        if status != "pending":
            remaining_triggers.append(trigger)
            continue

        if not followup_has_time(trigger):
            remaining_triggers.append(trigger)
            continue

        due_at = followup_due_datetime(trigger)

        if due_at is None:
            has_invalid_datetimes = True
            remaining_triggers.append(trigger)
            continue

        if due_at > now:
            remaining_triggers.append(trigger)
            continue

        message_template = str(trigger.get("message_template", "")).strip()

        if not message_template:
            remaining_triggers.append(trigger)
            continue

        await reply_text_func(message_template)

        trigger["status"] = "sent"
        trigger["sent_at"] = format_app_datetime()

        append_proactivity_log(
            event_type="timed_followup_sent",
            trigger=trigger,
            notes=notes,
        )

        archive_followup_trigger(trigger, "timed_sent")
        sent_count += 1

    save_followup_triggers(remaining_triggers)

    return {
        "sent_count": sent_count,
        "has_invalid_datetimes": has_invalid_datetimes,
    }



async def run_followups(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    async def reply_text(text: str) -> None:
        await update.message.reply_text(
            text,
            reply_markup=MAIN_KEYBOARD,
        )
    
    activate_client_from_update(update)

    result = await send_due_followups(
        reply_text_func=reply_text,
        notes="Sent manually via /run_followups.",
    )

    sent_count = result["sent_count"]
    has_invalid_dates = result["has_invalid_dates"]

    if sent_count == 0:
        message = load_required_app_text("run_followups_no_due_message.md")

        if has_invalid_dates:
            message += "\n\n" + load_required_app_text("run_followups_invalid_date_message.md")

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    summary_template = load_required_app_text("run_followups_sent_summary.md")

    summary_message = render_template(
        summary_template,
        {
            "count": str(sent_count),
        },
    )

    await update.message.reply_text(
        summary_message,
        reply_markup=MAIN_KEYBOARD,
    )


async def scheduled_run_followups(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Daily automatic seguimiento check for all Telegram client folders."""
    client_names = iter_telegram_client_names()

    try:
        if not client_names:
            set_current_client_name(DEFAULT_CLIENT_NAME)

            append_text_file(
                wealth_log_file(),
                "\n".join(
                    [
                        f"## {format_app_minute()}",
                        "",
                        "### Scheduler",
                        "",
                        "No Telegram client folders found. Automatic seguimientos were not sent.",
                        "",
                        "---",
                        "",
                    ]
                ),
            )
            return

        for client_name in client_names:
            async with get_client_lock(client_name):
                set_current_client_name(client_name)
                ensure_client_files()
                ensure_followup_triggers_file()

                chat_id = get_remembered_chat_id()

                if chat_id is None:
                    append_text_file(
                        wealth_log_file(),
                        "\n".join(
                            [
                                f"## {format_app_minute()}",
                                "",
                                "### Scheduler",
                                "",
                                "No Telegram chat ID remembered for this client. Automatic seguimientos were not sent.",
                                f"Client: {client_name}",
                                "",
                                "---",
                                "",
                            ]
                        ),
                    )
                    continue

                async def send_text(text: str, target_chat_id: int = chat_id) -> None:
                    await context.bot.send_message(
                        chat_id=target_chat_id,
                        text=text,
                        reply_markup=MAIN_KEYBOARD,
                    )

                result = await send_due_followups(
                    reply_text_func=send_text,
                    notes="Sent automatically by daily scheduler.",
                )

                append_text_file(
                    wealth_log_file(),
                    "\n".join(
                        [
                            f"## {format_app_minute()}",
                            "",
                            "### Scheduler",
                            "",
                            "Automatic seguimiento check completed for client.",
                            f"Client: {client_name}",
                            f"Sent count: {result['sent_count']}",
                            f"Invalid dates present: {result['has_invalid_dates']}",
                            "",
                            "---",
                            "",
                        ]
                    ),
                )

    finally:
        set_current_client_name(DEFAULT_CLIENT_NAME)


def append_proactivity_log(
    event_type: str,
    trigger: dict,
    notes: str = "",
) -> None:
    """Append a proactive-agent event to proactivity_log.md."""
    ensure_client_files()

    timestamp = format_app_datetime()

    template = load_required_app_text("proactivity_log_entry.md")

    entry = render_template(
        template,
        {
            "timestamp": timestamp,
            "event_type": event_type,
            "followup_id": str(trigger.get("id", "")),
            "followup_type": str(trigger.get("type", "")),
            "followup_date": str(trigger.get("date", "")),
            "status": str(trigger.get("status", "")),
            "message_template": str(trigger.get("message_template", "")),
            "notes": notes,
        },
    )

    append_text_file(client_files()["proactivity_log"], entry)

async def show_proactivity_log(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    log_text = read_text_file(client_files()["proactivity_log"])

    if not log_text:
        log_text = load_required_app_text("proactivity_log_empty_message.md")

    await send_long_message(
        update,
        log_text,
        reply_markup=MAIN_KEYBOARD,
    )

async def show_agent_observations(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    observations = read_text_file(client_files()["agent_observations"])

    if not observations:
        observations = load_required_app_text("agent_observations_empty_message.md")

    await send_long_message(
        update,
        observations,
        reply_markup=MAIN_KEYBOARD,
    )


async def scheduled_run_timed_followups(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Frequent automatic check for timed seguimientos."""
    client_names = iter_telegram_client_names()

    for client_name in client_names:
        async with get_client_lock(client_name):
            set_current_client_name(client_name)
            ensure_client_files()
            ensure_followup_triggers_file()

            chat_id = get_remembered_chat_id()

            if chat_id is None:
                continue

            async def reply_text(text: str, chat_id=chat_id) -> None:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=text,
                    reply_markup=MAIN_KEYBOARD,
                )

            await send_due_timed_followups(
                reply_text_func=reply_text,
                notes="Sent automatically by timed seguimiento scheduler.",
            )


# ---------------------------------------------------------------------
# App
# ---------------------------------------------------------------------

# ---------------------------------------------------------------------
# Short alternatives flow overrides
# These definitions intentionally override earlier longer versions.
# ---------------------------------------------------------------------

def build_alternatives_intro_message() -> str:
    return (
        "Ahora no estamos eligiendo un producto. Estamos eligiendo una postura de control.\n\n"
        "1. Cartera autogestionada\n"
        "- Máxima soberanía.\n"
        "- Más carga mental.\n\n"
        "2. Gestión delegada\n"
        "- Más simple y tranquila.\n"
        "- Pero sigue siendo delegar.\n\n"
        "3. Cartera guiada\n"
        "- Punto intermedio.\n"
        "- La clienta decide, eCoach estructura y acompaña.\n\n"
        "La historia central:\n"
        "Banco = dependencia opaca.\n"
        "Delegada = dependencia más limpia.\n"
        "Guiada = soberanía acompañada."
    )


def build_selected_alternative_path_message(path_label: str) -> str:
    if "deleg" in path_label.lower():
        return (
            "Camino elegido provisionalmente: gestión delegada\n\n"
            "Tiene sentido si ahora pesa el miedo o la carga mental. Puede ser más limpio que el banco.\n\n"
            "Pero hay que nombrarlo bien: sigue siendo delegación.\n\n"
            "Siguiente paso: comparar proveedores delegados y ver si esta postura encaja con lo que la clienta quiere."
        )

    if "guiada" in path_label.lower():
        return (
            "Camino elegido: cartera guiada\n\n"
            "Esto significa: no autogestionada sola, no delegada a ciegas. Soberanía acompañada.\n\n"
            "Siguiente paso: elegir una plataforma/proveedor a estudiar y construir un Mi Plan actualizado."
        )

    return (
        f"Camino elegido: {path_label}\n\n"
        "Siguiente paso: comparar con la cartera actual y comprobar si este camino encaja con coste, riesgo, carga mental y soberanía."
    )


def build_risk_intention_confirmation_message(risk_label: str) -> str:
    return (
        f"Riesgo fijado: {risk_label}\n\n"
        "Siguiente paso: compara con tu cartera actual o define una alternativa concreta."
    )


async def handle_alternatives_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        with_inline_transition_separator(build_alternatives_intro_message()),
        reply_markup=alternatives_path_keyboard(),
    )



async def handle_compare_current_portfolio_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        "Para esta demo, seguimos por el camino de cartera guiada: primero elegimos proveedor y después construimos una propuesta concreta.",
        reply_markup=provider_comparison_keyboard(),
    )



async def handle_missing_data_checklist_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        "Para esta demo no abrimos todavía la rama de datos faltantes. Seguimos por el camino principal: elegir proveedor y construir una propuesta guiada.",
        reply_markup=provider_comparison_keyboard(),
    )


async def handle_define_concrete_alternative_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        "Para esta demo construiremos la alternativa concreta después de elegir proveedor. El siguiente paso es seleccionar MyInvestor, Diaphanum o EBN Banco.",
        reply_markup=provider_comparison_keyboard(),
    )


async def handle_self_managed_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    save_selected_alternative_path(
        client_dir,
        "self_managed",
        "Cartera autogestionada",
    )
    ecoach_select_path("self_managed", "Cartera autogestionada", client_dir)

    await query.message.reply_text(
        with_inline_transition_separator(
            build_selected_alternative_path_message("Cartera autogestionada")
        ),
        reply_markup=compare_current_portfolio_keyboard(),
    )


async def handle_delegated_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    save_selected_alternative_path(
        client_dir,
        "delegated",
        "Gestión delegada",
    )
    ecoach_select_path("delegated", "Gestión delegada", client_dir)

    await query.message.reply_text(
        with_inline_transition_separator(
            build_selected_alternative_path_message("Gesti?n delegada")
        ),
        reply_markup=compare_current_portfolio_keyboard(),
    )



async def handle_guided_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    save_selected_alternative_path(
        client_dir,
        "guided",
        "Cartera guiada",
    )
    ecoach_select_path("guided", "Cartera guiada", client_dir)

    facts = """Camino elegido: cartera guiada.

Concepto:
- No es cartera autogestionada sola.
- No es gestión delegada a ciegas.
- Es soberanía acompañada: la usuaria mantiene el control, pero eCoach ayuda a entender, comparar y avanzar paso a paso.

Siguiente decisión:
Elegir proveedor/plataforma desde donde construir una propuesta concreta.

Opciones:
1. MyInvestor:
   - Banco/plataforma con muchos fondos y costes bajos.
   - Ventaja: práctico, amplio y útil para construir una cartera concreta de bajo coste.
   - Límite: sigue siendo un banco/plataforma con lógica comercial.

2. Diaphanum:
   - Asesor financiero más tradicional, humano y premium.
   - Ventaja: más acompañamiento personal.
   - Límite: puede ser menos simple o más caro para una cartera pequeña.

3. EBN Banco:
   - Banco con arquitectura abierta y comisión explícita.
   - Ventaja: transparencia de costes.
   - Límite: hay que entender la comisión explícita y compararla con el ahorro en fondos.

Finalidad de la respuesta:
Explicar el camino de cartera guiada y terminar preguntando qué proveedor quiere estudiar primero.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_fund_portfolio",
        task="Write the client-facing answer after the user chooses cartera guiada.",
        facts=facts,
        reply_markup=provider_comparison_keyboard(),
    )

async def handle_keep_similar_risk_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    save_selected_risk_intention(
        client_dir,
        "keep_similar",
        "Mantener riesgo parecido",
    )

    await query.message.reply_text(
        with_inline_transition_separator(
            build_risk_intention_confirmation_message("Mantener riesgo parecido")
        ),
        reply_markup=compare_current_portfolio_keyboard(),
    )


async def handle_reduce_risk_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    save_selected_risk_intention(
        client_dir,
        "reduce_risk",
        "Reducir riesgo",
    )

    await query.message.reply_text(
        with_inline_transition_separator(
            build_risk_intention_confirmation_message("Reducir riesgo")
        ),
        reply_markup=compare_current_portfolio_keyboard(),
    )

# ---------------------------------------------------------------------
# Available-data orientation handler override
# Forces the "Continuar con datos disponibles" screen to show the next-step keyboard.
# ---------------------------------------------------------------------

async def handle_prepare_provisional_alternative_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    selected_path = load_selected_alternative_path(client_dir)
    selected_risk = load_selected_risk_intention(client_dir)

    try:
        isins = extract_recent_isins_for_public_enrichment(client_dir)
    except Exception:
        isins = []

    msg = build_provisional_alternative_message(selected_path, selected_risk, client_dir, isins)

    await query.message.reply_text(
        with_inline_transition_separator(msg),
        reply_markup=available_data_orientation_keyboard(),
    )










def portfolio_result_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Revisar antes de ejecutar",
                    callback_data=REVIEW_BEFORE_EXECUTE_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "Cambiar proveedor",
                    callback_data=SEARCH_PROVIDERS_CALLBACK,
                )
            ],
        ]
    )

def fund_search_result_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Crear cartera coste menor",
                    callback_data="build_same_risk_lower_cost_portfolio",
                )
            ],
            [
                InlineKeyboardButton(
                    "Cambiar proveedor",
                    callback_data=SEARCH_PROVIDERS_CALLBACK,
                )
            ],
        ]
    )

def provider_selected_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f50e Buscar fondos disponibles",
                    callback_data=SEARCH_FUNDS_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\u21a9\ufe0f Cambiar proveedor",
                    callback_data=SEARCH_PROVIDERS_CALLBACK,
                )
            ],
        ]
    )



def create_mi_plan_followup_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Crear seguimiento — 17/06 10:00",
                    callback_data=CREATE_MI_PLAN_FOLLOWUP_CALLBACK,
                )
            ]
        ]
    )


def design_mi_plan_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Practicar caso real — mensaje 23:40",
                    callback_data=DESIGN_MI_PLAN_CALLBACK,
                )
            ]
        ]
    )

def provider_comparison_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "MyInvestor",
                    callback_data=PROVIDER_MYINVESTOR_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "Diaphanum",
                    callback_data=PROVIDER_DIAPHANUM_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "EBN Banco",
                    callback_data=PROVIDER_EBN_CALLBACK,
                )
            ],
        ]
    )


def comparison_table_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f50e Buscar proveedores",
                    callback_data=SEARCH_PROVIDERS_CALLBACK,
                )
            ],
        ]
    )

def available_data_orientation_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f50e Buscar proveedores",
                    callback_data=SEARCH_PROVIDERS_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f4ca Ver tabla comparativa",
                    callback_data=DEFINE_INITIAL_ALTERNATIVE_CALLBACK,
                )
            ],
        ]
    )



def build_initial_alternative_message(
    selected_path: dict | None,
    selected_risk: dict | None,
    client_dir: Path | None = None,
    isins: list[str] | None = None,
) -> str:
    return (
        "Tabla comparativa inicial\n\n"
        "| Ruta | Objetivo | Coste esperado | Riesgo | Simplicidad |\n"
        "|---|---|---:|---|---|\n"
        "| Actual | Mantener cartera | 1,41% visible | Actual | Baja/media |\n"
        "| A. Simple | Riesgo parecido, menor coste | Menor | Parecido | Alta |\n"
        "| B. Defensiva | Menor riesgo, mas tranquilidad | Menor o similar | Menor | Alta |\n\n"
        "Lectura:\n"
        "- La ruta A parece la mas natural si quieres mantener una logica parecida y reducir coste/solapamiento.\n"
        "- La ruta B tiene sentido si priorizas tranquilidad sobre rentabilidad.\n"
        "- No es recomendacion final: faltan fiscalidad, plusvalias/minusvalias y costes reales de cuenta.\n\n"
        "Siguiente paso:\n"
        "Buscar proveedores/plataformas para ejecutar A o B."
    )

async def handle_define_initial_alternative_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    user_id = update.effective_user.id
    client_dir = Path("ClientData") / f"telegram_{user_id}"

    selected_path = load_selected_alternative_path(client_dir)
    selected_risk = load_selected_risk_intention(client_dir)

    try:
        isins = extract_recent_isins_for_public_enrichment(client_dir)
    except Exception:
        isins = []

    msg = build_initial_alternative_message(selected_path, selected_risk, client_dir, isins)

    await query.message.reply_text(
        with_inline_transition_separator(msg),
        reply_markup=comparison_table_keyboard(),
    )



def build_provider_search_placeholder_message() -> str:
    return (
        "Proveedores/plataformas a comparar\n\n"
        "Comparación inicial, no búsqueda web en vivo todavía.\n\n"
        "| Proveedor | Tipo | Encaje inicial |\n"
        "|---|---|---|\n"
        "| MyInvestor | Plataforma/RTO de fondos | Candidato para cartera guiada si encajan costes y catálogo |\n"
        "| Renta 4 | Broker/plataforma de fondos | Amplio universo, más tradicional |\n"
        "| Openbank | Banco/plataforma | Entorno bancario más simple |\n"
        "| Indexa | Gestión delegada indexada | Encaja con delegada: simple, pero menos soberanía |\n"
        "| InbestMe | Gestión delegada indexada | Encaja con delegada: simple, pero menos soberanía |\n\n"
        "Criterio rápido:\n"
        "- Para gestión delegada: mirar Indexa o InbestMe.\n"
        "- Para cartera guiada: estudiar MyInvestor, Renta 4 u Openbank según fondos disponibles y facilidad operativa.\n\n"
        "Siguiente paso:\n"
        "Elige proveedor para preparar la siguiente capa del plan."
    )

async def handle_search_providers_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        with_inline_transition_separator(build_provider_search_placeholder_message()),
        reply_markup=provider_comparison_keyboard(),
    )

PROVIDER_LABELS = {
    PROVIDER_MYINVESTOR_CALLBACK: "MyInvestor",
    PROVIDER_DIAPHANUM_CALLBACK: "Diaphanum",
    PROVIDER_EBN_CALLBACK: "EBN Banco",
}




def save_selected_provider(client_dir: Path, provider_name: str) -> None:
    try:
        client_dir.mkdir(parents=True, exist_ok=True)
        (client_dir / "selected_provider.json").write_text(
            json.dumps({"provider_name": provider_name}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        return


def load_selected_provider(client_dir: Path) -> str:
    try:
        in_path = client_dir / "selected_provider.json"
        if not in_path.exists():
            return "proveedor elegido"
        data = json.loads(in_path.read_text(encoding="utf-8"))
        return data.get("provider_name") or "proveedor elegido"
    except Exception:
        return "proveedor elegido"

def build_provider_selected_message(provider_name: str, provider_key: str | None = None) -> str:
    provider_key = provider_key or provider_name.lower().replace(" ", "_")

    if provider_key == "indexa":
        return (
            "Proveedor elegido para estudiar: Indexa\n\n"
            "Encaje: gestión delegada. Es una forma más clara y normalmente más estructurada que una cartera bancaria opaca, "
            "pero sigue siendo delegación.\n\n"
            "Mi Plan provisional:\n"
            "1. simular perfil de riesgo;\n"
            "2. comparar costes con la cartera bancaria;\n"
            "3. revisar si el riesgo se parece al actual;\n"
            "4. decidir si delegar encaja con lo que la clienta quiere;\n"
            "5. programar seguimiento.\n\n"
            "Nota importante: esto no es una orden de traspasar. Es una opción estructurada."
        )

    if provider_key == "myinvestor":
        return (
            "Proveedor elegido para estudiar: MyInvestor\n\n"
            "Encaje: cartera guiada / RTO, si la clienta quiere más soberanía sin quedarse sola.\n\n"
            "No lo recomiendo como primera opción automática. Lo tratamos como candidato razonable si encaja con los criterios: "
            "coste, fondos disponibles, facilidad operativa y posibilidad de seguimiento.\n\n"
            "Mi Plan guiado:\n"
            "1. definir estructura simple;\n"
            "2. buscar fondos disponibles;\n"
            "3. revisar costes y traspasabilidad;\n"
            "4. preparar checklist de apertura;\n"
            "5. acompañar el proceso sin ejecutar por la clienta."
        )

    return (
        f"Proveedor elegido para estudiar: {provider_name}\n\n"
        "Siguiente paso:\n"
        "Buscar fondos disponibles y comprobar si encaja con el camino elegido, los costes, la simplicidad operativa y el nivel de soberanía deseado."
    )




async def handle_create_mi_plan_followup_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    facts = """Follow-up created.

Exact date:
- miércoles 17 de junio de 2026

Exact time:
- 10:00

Topic:
- revisar el avance del Mi Plan con MyInvestor

In that follow-up, eCoach will review:
- whether the MyInvestor account is operational;
- whether the two funds have been found;
- whether any doubt appeared during the transfer;
- whether the 80/20 target portfolio is still maintained;
- what the next concrete step is.

Finality:
Confirm the follow-up has been created.
Keep the message calm, concrete and motivating.
"""

    followup_text = await asyncio.to_thread(
        generate_skill_client_reply,
        "manage_fund_portfolio",
        "Write the follow-up creation confirmation.",
        facts,
    )

    try:
        client_dir = active_client_dir()
        followups_dir = client_dir / "followups"
        followups_dir.mkdir(parents=True, exist_ok=True)
        followup_file = followups_dir / "mi_plan_followup_2026-06-17_10-00.md"
        followup_file.write_text(followup_text, encoding="utf-8")
    except Exception:
        pass

    await query.message.reply_text(followup_text)

async def handle_design_mi_plan_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    facts = """Mi Plan.

Objective:
- Move from an expensive and unclear portfolio to a simpler, understandable and cheaper portfolio.
- Keep a similar profile to the current portfolio.

Profile to preserve:
- Current approximate portfolio: 20% equity / 80% defensive part
- Profile: conservative-moderate
- Main objective: reduce costs without increasing risk much

Target portfolio in MyInvestor:
1. 80% — DWS EURO ULTRA SHORT "NC" (EUR) A
   - Function: defensive / money-market / very short fixed income part
   - Approximate amount: 16.710 €
   - Estimated cost: 0,15%

2. 20% — AMUNDI INDEX MSCI WORLD "AE" (EUR) ACC
   - Function: global indexed equity
   - Approximate amount: 4.178 €
   - Estimated cost: 0,30%

Expected improvement:
- Current visible cost: ~293 €/year
- Estimated new portfolio cost: ~38 €/year
- Estimated saving: ~255 €/year
- Approximate reduction: ~87% less visible cost

Practical steps:
1. Open or review MyInvestor account.
2. Search for the two funds:
   - DWS EURO ULTRA SHORT "NC" (EUR) A
   - AMUNDI INDEX MSCI WORLD "AE" (EUR) ACC
3. Prepare transfer while keeping target proportion:
   - 80% defensive part
   - 20% global equity
4. Save confirmations, dates and transferred funds.

Follow-up:
- Proposed first follow-up: miércoles 17 de junio de 2026 a las 10:00
- In the follow-up we will review:
  - whether the MyInvestor account is operational;
  - whether the two funds have been found;
  - whether doubts appeared during the transfer process;
  - whether the target portfolio remains 80/20;
  - and the next concrete step.

Finality:
Write Mi Plan clearly, warmly and practically.
End by making the follow-up feel like the next clear step.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_fund_portfolio",
        task="Write Mi Plan after the user accepts the MyInvestor proposal.",
        facts=facts,
        reply_markup=create_mi_plan_followup_keyboard(),
    )

async def handle_provider_selected_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    callback_data = query.data or ""
    provider_name = PROVIDER_LABELS.get(callback_data, "proveedor")

    if provider_name == "MyInvestor":
        facts = """Proveedor elegido: MyInvestor.

Cartera original:
- Importe total analizado: 20.888 €
- Exposición estimada a renta variable: aproximadamente 20%
- Parte defensiva / renta fija / monetaria: aproximadamente 80%
- Coste visible actual estimado: ~1,40% anual
- Coste visible actual en euros: ~293 €/año

Objetivo:
- Mantener un perfil parecido al de la cartera actual.
- No cambiar radicalmente el riesgo.
- Simplificar la cartera.
- Reducir costes.

Propuesta provisional en MyInvestor:
1. 80% — DWS EURO ULTRA SHORT "NC" (EUR) A
   - Función: parte defensiva / monetaria / renta fija muy corta
   - Coste estimado: 0,15%
   - Importe aproximado: 16.710 €

2. 20% — AMUNDI INDEX MSCI WORLD "AE" (EUR) ACC
   - Función: renta variable global indexada
   - Coste estimado: 0,30%
   - Importe aproximado: 4.178 €

Coste estimado nueva cartera:
- 80% × 0,15% = 0,12%
- 20% × 0,30% = 0,06%
- Coste ponderado total: ~0,18% anual
- 20.888 € × 0,18% = ~38 €/año

Diferencia positiva de coste:
- Coste actual visible: ~293 €/año
- Coste estimado nueva cartera: ~38 €/año
- Ahorro estimado: ~255 €/año
- Reducción aproximada: ~87% menos coste visible

Finalidad:
Explain the proposed MyInvestor portfolio clearly, preserving all numbers exactly.
End by saying that if it seems reasonable, the next step is to convert it into Mi Plan: what to check before transferring, what to ask if a doubt appears in MyInvestor, and how follow-up will work.
Do not add extra caution/legal paragraphs.
"""

        await answer_callback_with_skill(
            query=query,
            skill_name="manage_fund_portfolio",
            task="Write the MyInvestor concrete portfolio proposal.",
            facts=facts,
            reply_markup=design_mi_plan_keyboard(),
        )
        return

    if provider_name == "Diaphanum":
        facts = """Proveedor elegido: Diaphanum.

Diaphanum:
- More traditional and human financial adviser.
- More premium.
- Can make sense if the user wants human accompaniment.
- For a small portfolio of 20.888 €, total cost or complexity may not compensate versus a simpler low-cost solution.
- For this demo, the most direct provider for a concrete low-cost portfolio is MyInvestor.

Finalidad:
Explain this fairly and invite the user to choose a provider again.
"""

        await answer_callback_with_skill(
            query=query,
            skill_name="manage_fund_portfolio",
            task="Write the provider explanation for Diaphanum.",
            facts=facts,
            reply_markup=provider_comparison_keyboard(),
        )
        return

    if provider_name == "EBN Banco":
        facts = """Proveedor elegido: EBN Banco.

EBN Banco:
- Open architecture logic.
- Explicit fee.
- Can be more transparent than hidden costs inside expensive bank funds.
- Key comparison: fund cost savings versus EBN's explicit fee.
- For this demo, the most direct provider for a concrete low-cost portfolio is MyInvestor.

Finalidad:
Explain this fairly and invite the user to choose a provider again.
"""

        await answer_callback_with_skill(
            query=query,
            skill_name="manage_fund_portfolio",
            task="Write the provider explanation for EBN Banco.",
            facts=facts,
            reply_markup=provider_comparison_keyboard(),
        )
        return

    facts = """Unknown provider selected.

For this demo, the provider developed with a concrete portfolio is MyInvestor.
Invite the user to choose MyInvestor, Diaphanum or EBN Banco.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_fund_portfolio",
        task="Write fallback provider-choice answer.",
        facts=facts,
        reply_markup=provider_comparison_keyboard(),
    )

def get_provider_public_sources(provider_name: str) -> list[dict]:
    return PROVIDER_PUBLIC_SOURCES.get(provider_name, [])


def build_provider_sources_section(provider_name: str) -> str:
    sources = get_provider_public_sources(provider_name)

    if not sources:
        return (
            "Fuentes publicas controladas:\n"
            "- Pendiente de definir fuentes oficiales para este proveedor.\n"
        )

    lines = ["Fuentes publicas controladas:"]
    for source in sources:
        label = source.get("label", "Fuente")
        purpose = source.get("purpose", "consulta publica")
        url = source.get("url", "")
        lines.append(f"- {label}: {purpose}. {url}")

    return "\n".join(lines) + "\n"



def fetch_public_source_status(url: str, timeout_seconds: int = 8) -> dict:
    """
    Controlled public-source fetch.

    Safety:
    - Fetches only predefined public provider URLs.
    - Sends no client private data.
    - Stores only status/title-level metadata.
    """
    try:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 eCoachRelaciones/0.1 public-source-check"
            },
        )

        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            status_code = getattr(response, "status", None) or response.getcode()
            content_type = response.headers.get("Content-Type", "")
            raw = response.read(120000)

        title = ""
        try:
            html = raw.decode("utf-8", errors="ignore")
            match = re.search(r"<title[^>]*>(.*?)</title>", html, flags=re.I | re.S)
            if match:
                title = re.sub(r"\s+", " ", match.group(1)).strip()
        except Exception:
            title = ""

        return {
            "ok": True,
            "status_code": status_code,
            "content_type": content_type,
            "title": title,
            "error": "",
        }

    except Exception as error:
        return {
            "ok": False,
            "status_code": "",
            "content_type": "",
            "title": "",
            "error": str(error)[:160],
        }


def build_provider_source_status_section(provider_name: str) -> str:
    sources = get_provider_public_sources(provider_name)

    if not sources:
        return (
            "Fuentes publicas controladas:\n"
            "- Pendiente de definir fuentes oficiales para este proveedor.\n"
        )

    lines = ["Fuentes publicas controladas:"]
    for source in sources:
        label = source.get("label", "Fuente")
        url = source.get("url", "")
        status = fetch_public_source_status(url)

        if status.get("ok"):
            title = status.get("title") or "sin titulo detectado"
            lines.append(f"- {label}: OK ({status.get('status_code')}) - {title}")
        else:
            lines.append(f"- {label}: no verificada ahora - {status.get('error')}")

    return "\n".join(lines) + "\n"



def fetch_public_source_text(url: str, timeout_seconds: int = 10, max_bytes: int = 400000) -> dict:
    """
    Controlled public-source text fetch.
    Fetches only predefined provider URLs and sends no client private data.
    """
    try:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 eCoachRelaciones/0.1 public-source-read"
            },
        )

        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            status_code = getattr(response, "status", None) or response.getcode()
            content_type = response.headers.get("Content-Type", "")
            raw = response.read(max_bytes)

        text_value = raw.decode("utf-8", errors="ignore")

        return {
            "ok": True,
            "status_code": status_code,
            "content_type": content_type,
            "text": text_value,
            "error": "",
        }

    except Exception as error:
        return {
            "ok": False,
            "status_code": "",
            "content_type": "",
            "text": "",
            "error": str(error)[:160],
        }


def strip_html_for_public_fund_scan(html: str) -> str:
    cleaned = re.sub(r"(?is)<script.*?</script>", " ", html or "")
    cleaned = re.sub(r"(?is)<style.*?</style>", " ", cleaned)
    cleaned = re.sub(r"(?is)<[^>]+>", " ", cleaned)
    cleaned = re.sub(r"&nbsp;|&#160;", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def extract_isin_contexts_from_text(text_value: str, max_items: int = 12) -> list[dict]:
    plain = strip_html_for_public_fund_scan(text_value)
    results = []
    seen = set()

    for match in re.finditer(r"\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b", plain):
        isin = match.group(0)
        if isin in seen:
            continue
        seen.add(isin)

        start = max(0, match.start() - 110)
        end = min(len(plain), match.end() + 110)
        context = plain[start:end].strip()

        results.append(
            {
                "isin": isin,
                "context": context,
            }
        )

        if len(results) >= max_items:
            break

    return results


def extract_provider_fund_candidates(provider_name: str, max_items: int = 12) -> list[dict]:
    sources = get_provider_public_sources(provider_name)
    candidates = []

    for source in sources:
        purpose = (source.get("purpose") or "").lower()
        label = source.get("label", "Fuente")
        url = source.get("url", "")

        # Prefer fund pages, but allow all sources as fallback.
        if "fondo" not in purpose and "fondos" not in label.lower():
            continue

        fetched = fetch_public_source_text(url)
        if not fetched.get("ok"):
            continue

        for item in extract_isin_contexts_from_text(fetched.get("text", ""), max_items=max_items):
            item["source_label"] = label
            item["source_url"] = url
            candidates.append(item)

            if len(candidates) >= max_items:
                return candidates

    return candidates



def build_provider_fund_candidates_section(provider_name: str) -> str:
    candidates = extract_provider_fund_candidates(provider_name)

    if not candidates:
        return (
            "Fondos detectados automaticamente:\n"
            "- No he detectado ISINs en las paginas publicas controladas usadas ahora.\n"
            "- Esto no significa que no existan fondos disponibles; significa que esta fuente no los expone de forma facil en HTML.\n"
            "- Siguiente mejora: anadir una fuente/catalogo mas directa del proveedor o una busqueda web controlada.\n"
        )

    lines = ["Fondos detectados automaticamente:"]

    seen = set()
    for item in candidates:
        isin = item.get("isin", "")
        source_label = item.get("source_label", "Fuente")

        if not isin or isin in seen:
            continue

        seen.add(isin)
        lines.append(f"- {isin} ? fuente: {source_label}")

    lines.append("")
    lines.append("Nota:")
    lines.append("- Detectados en fuente publica controlada.")
    lines.append("- Falta enriquecer nombre, categoria, TER y tipo de activo.")
    lines.append("- No son todavia una cartera recomendada.")

    return "\n".join(lines) + "\n"


MYINVESTOR_RICH_CATALOG_JS_URL = "https://myinvestor.es/_next/static/chunks/pages/_app-bfe99f422a71afdc.js"


def parse_float_or_none(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        value = str(value).strip().replace(",", ".")
        if not value:
            return None
        return float(value)
    except Exception:
        return None


def clean_js_string_value(value: str | None) -> str:
    if not value:
        return ""
    try:
        return value.encode("utf-8").decode("unicode_escape", errors="ignore")
    except Exception:
        return value


def extract_myinvestor_rich_catalog(max_bytes: int = 3500000) -> list[dict]:
    """
    Extracts a public enriched MyInvestor catalogue subset from the Next.js bundle.

    Important:
    - This is a public source.
    - It is a subset, not necessarily the full 911-fund universe.
    - It includes TER and asset metadata for many funds.
    """
    try:
        fetched = fetch_public_source_text(
            MYINVESTOR_RICH_CATALOG_JS_URL,
            timeout_seconds=15,
            max_bytes=max_bytes,
        )
        js = fetched.get("text", "") if fetched.get("ok") else ""
    except Exception:
        js = ""

    if not js:
        return []

    num = r"(-?(?:\d+(?:\.\d+)?|\.\d+))"

    def grab(pattern: str, source: str) -> str:
        match = re.search(pattern, source, flags=re.S)
        return match.group(1) if match else ""

    starts = [
        m.start()
        for m in re.finditer(
            r'\{idFondo:\d+,codigoIsin:"[A-Z]{2}[A-Z0-9]{9}[0-9]"',
            js,
        )
    ]

    rows = []
    seen = set()

    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else min(len(js), start + 30000)
        obj = js[start:end]

        isin = grab(r'codigoIsin:"([A-Z]{2}[A-Z0-9]{9}[0-9])"', obj)
        if not isin or isin in seen:
            continue
        seen.add(isin)

        id_fondo = grab(r"idFondo:(\d+)", obj)
        name = clean_js_string_value(grab(r'nombre:"(.*?)"', obj))
        category = clean_js_string_value(grab(r'categoria:"(.*?)"', obj))
        manager = clean_js_string_value(
            grab(r'entidadGestora:"(.*?)"', obj)
            or grab(r'categoriaMyInvestor:"(.*?)"', obj)
        )
        asset = clean_js_string_value(grab(r'tipoActivo:"(.*?)"', obj))
        geo = clean_js_string_value(grab(r'zonaGeografica:"(.*?)"', obj))

        row = {
            "idFondo": int(id_fondo) if id_fondo else None,
            "isin": isin,
            "name": name,
            "category": category,
            "manager": manager,
            "ter": parse_float_or_none(grab(r"ter:" + num, obj)),
            "asset": asset,
            "geo": geo,
            "equity_pct": parse_float_or_none(grab(r"activosAcciones:" + num, obj)),
            "bonds_pct": parse_float_or_none(grab(r"activosObligaciones:" + num, obj)),
            "cash_pct": parse_float_or_none(grab(r"activosEfectivo:" + num, obj)),
            "urlKiid": grab(
                r'urlKiid:"(https://api\.fundinfo\.com/document/.*?\.pdf\?apiKey=.*?)"',
                obj,
            ).strip(),
        }

        rows.append(row)

    return rows


def estimate_current_portfolio_equity_pct() -> float:
    """
    v1 estimate based on known current Rural portfolio categories.

    Current detected portfolio:
    - 22.5% RV Mixta 30/50 -> midpoint 40%
    - 5.0% RV Internacional >75% -> midpoint 87.5%
    - 29.6% RF Mixta max 25% RV -> midpoint 12.5%
    - 43.0% RF Mixta max 15% RV -> midpoint 7.5%
    """
    return (22.5 * 40.0 + 5.0 * 87.5 + 29.6 * 12.5 + 43.0 * 7.5) / 100.0


def choose_best_global_equity_candidate(catalog: list[dict]) -> dict | None:
    candidates = []
    for row in catalog:
        ter = row.get("ter")
        if ter is None or ter > 0.50:
            continue

        text_value = " ".join(
            [
                row.get("name", ""),
                row.get("category", ""),
                row.get("asset", ""),
                row.get("geo", ""),
                row.get("manager", ""),
            ]
        ).lower()

        is_equity = (
            row.get("asset") == "Renta Variable"
            or "equity" in text_value
            or "renta variable" in text_value
        )
        is_global = any(
            token in text_value
            for token in ["developed world", "global equity", "msci world", "global stock", "world index"]
        )
        is_not_sector = not any(
            token in text_value
            for token in ["real estate", "sector", "emerging", "small-cap", "small cap", "japan", "europe", "us equity", "s&p 500", "nasdaq"]
        )

        if is_equity and is_global and is_not_sector:
            candidates.append(row)

    if not candidates:
        return None

    return sorted(candidates, key=lambda r: (r.get("ter") or 999.0, r.get("name", "")))[0]


def choose_best_fixed_income_candidate(catalog: list[dict]) -> dict | None:
    candidates = []
    for row in catalog:
        ter = row.get("ter")
        if ter is None or ter > 0.50:
            continue

        text_value = " ".join(
            [
                row.get("name", ""),
                row.get("category", ""),
                row.get("asset", ""),
                row.get("geo", ""),
                row.get("manager", ""),
            ]
        ).lower()

        is_fixed_income = (
            row.get("asset") == "Renta Fija"
            or "fixed income" in text_value
            or "bond" in text_value
            or "renta fija" in text_value
            or "rf " in text_value
        )

        # Prefer EUR/global/eurozone diversified bond exposure, avoid USD/GBP/CHF when possible.
        has_reasonable_currency_fit = not any(
            token in text_value
            for token in [" usd ", " gbp ", " chf ", "dolar", "libra"]
        )

        if is_fixed_income and has_reasonable_currency_fit:
            candidates.append(row)

    if not candidates:
        return None

    def score(row: dict) -> tuple:
        text_value = (row.get("name", "") + " " + row.get("category", "") + " " + row.get("geo", "")).lower()

        preference = 0
        if "short-term" in text_value or "short term" in text_value or "corto plazo" in text_value:
            preference -= 2
        if "global" in text_value:
            preference -= 1
        if "euro" in text_value or "eurozone" in text_value or "eur" in text_value:
            preference -= 1
        if "inflation-linked" in text_value or "inflation" in text_value:
            preference += 2

        return (row.get("ter") or 999.0, preference, row.get("name", ""))

    return sorted(candidates, key=score)[0]


def build_same_risk_lower_cost_portfolio_message() -> str:
    total_amount = 208_882.44
    current_cost_pct = 1.41
    current_annual_cost = 2935.03

    equity_pct = estimate_current_portfolio_equity_pct()
    fixed_income_pct = 100.0 - equity_pct

    catalog = extract_myinvestor_rich_catalog()
    cheap_count = len([r for r in catalog if r.get("ter") is not None and r.get("ter") <= 0.50])

    equity_fund = choose_best_global_equity_candidate(catalog)
    fixed_income_fund = choose_best_fixed_income_candidate(catalog)

    if not equity_fund or not fixed_income_fund:
        return (
            "Cartera alternativa provisional\n\n"
            "No he podido construir una propuesta automatica suficientemente limpia con el catalogo publico enriquecido.\n\n"
            f"Catalogo leido: {len(catalog)} fondos.\n"
            f"Fondos con TER <= 0.50%: {cheap_count}.\n\n"
            "Siguiente paso:\n"
            "Revisar candidatos manualmente o usar el catalogo completo del buscador si esta disponible."
        )

    equity_ter = equity_fund.get("ter") or 0.0
    fixed_income_ter = fixed_income_fund.get("ter") or 0.0

    new_weighted_ter = (equity_pct * equity_ter + fixed_income_pct * fixed_income_ter) / 100.0
    new_annual_cost = total_amount * new_weighted_ter / 100.0
    annual_saving = current_annual_cost - new_annual_cost

    return (
        "Cartera alternativa provisional\n\n"
        "Objetivo:\n"
        "Mantener una exposicion RV/RF parecida y reducir coste visible.\n\n"
        "Cartera actual:\n"
        f"- RV estimada: {equity_pct:.1f}%.\n"
        f"- RF/defensivo estimado: {fixed_income_pct:.1f}%.\n"
        f"- Coste visible actual: {current_cost_pct:.2f}%.\n"
        f"- Coste anual actual: {current_annual_cost:,.0f} EUR/ano.\n\n"
        "Propuesta inicial MyInvestor:\n\n"
        "| Bloque | Fondo | ISIN | Peso | TER |\n"
        "|---|---|---|---:|---:|\n"
        f"| RV global | {equity_fund.get('name')} | {equity_fund.get('isin')} | {equity_pct:.1f}% | {equity_ter:.2f}% |\n"
        f"| RF / bonos | {fixed_income_fund.get('name')} | {fixed_income_fund.get('isin')} | {fixed_income_pct:.1f}% | {fixed_income_ter:.2f}% |\n\n"
        "Coste nuevo estimado:\n"
        f"- TER medio nuevo: {new_weighted_ter:.2f}%.\n"
        f"- Coste anual nuevo: {new_annual_cost:,.0f} EUR/ano.\n"
        f"- Ahorro anual estimado: {annual_saving:,.0f} EUR/ano.\n\n"
        "Fuente usada:\n"
        f"- Catalogo publico enriquecido MyInvestor: {len(catalog)} fondos leidos.\n"
        f"- Fondos con TER <= 0.50%: {cheap_count}.\n\n"
        "Aviso:\n"
        "Misma proporcion RV/RF aproximada no significa mismo riesgo exacto. Falta revisar duracion, credito, divisa, fiscalidad, traspasabilidad y plusvalias antes de ejecutar."
    )

def build_fund_search_v0_message(
    provider_name: str,
    source_status_section: str | None = None,
    fund_candidates_section: str | None = None,
) -> str:
    sources_section = source_status_section or build_provider_sources_section(provider_name)
    candidates_section = fund_candidates_section or build_provider_fund_candidates_section(provider_name)

    return (
        f"Fondos a buscar en {provider_name}\n\n"
        "Modulo v1: consulta de fuentes publicas controladas. No datos privados.\n\n"
        f"{sources_section}\n"
        f"{candidates_section}\n"
        "Rutas a construir:\n"
        "A. Riesgo parecido: 25-35% RV global + 65-75% RF/monetario/bonos.\n"
        "B. Defensiva: 15-20% RV global + 80-85% RF/monetario/bonos.\n\n"
        "Filtros:\n"
        "- Disponible en el proveedor elegido.\n"
        "- Traspasable si es fondo.\n"
        "- Bajo coste.\n"
        "- UCITS / diversificado.\n"
        "- Fiscalidad y plusvalias pendientes antes de mover dinero.\n\n"
        "Siguiente modulo:\n"
        "Enriquecer estos ISINs con nombre, categoria, TER y tipo de activo."
    )


async def handle_build_same_risk_lower_cost_portfolio_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    msg = await asyncio.to_thread(build_same_risk_lower_cost_portfolio_message)

    await query.message.reply_text(
        with_inline_transition_separator(msg),
        reply_markup=portfolio_result_keyboard(),
    )



def build_review_before_execute_message() -> str:
    return (
        "Revision antes de ejecutar\n\n"
        "La cartera propuesta reduce mucho el coste visible, pero antes de mover dinero hay que revisar 5 cosas:\n\n"
        "1. Duracion de la renta fija\n"
        "- Confirmar que el fondo de bonos no tenga mucha mas sensibilidad a tipos que la cartera actual.\n\n"
        "2. Riesgo de credito\n"
        "- Confirmar si la renta fija es gobierno, corporativa, investment grade o mezcla.\n\n"
        "3. Divisa\n"
        "- Priorizar clase EUR o EUR hedged si el objetivo es evitar riesgo divisa adicional.\n\n"
        "4. Fiscalidad y traspasabilidad\n"
        "- Confirmar que los fondos son traspasables y que el cambio no genera venta fiscal no deseada.\n\n"
        "5. Plusvalias/minusvalias\n"
        "- Pedir al banco el detalle fiscal antes de ejecutar.\n\n"
        "Decision provisional:\n"
        "- La propuesta es valida como comparacion de coste.\n"
        "- Todavia no es una orden de ejecucion.\n\n"
        "Siguiente paso:\n"
        "Pedir o cargar datos fiscales de la cartera actual."
    )


async def handle_review_before_execute_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    await query.message.reply_text(
        with_inline_transition_separator(build_review_before_execute_message()),
        reply_markup=portfolio_result_keyboard(),
    )

async def handle_search_funds_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    msg = await asyncio.to_thread(build_same_risk_lower_cost_portfolio_message)

    await query.message.reply_text(
        with_inline_transition_separator(msg),
        reply_markup=portfolio_result_keyboard(),
    )


# ---------------------------------------------------------------------
# eCoach Relaciones — Saturday demo relationship flow overrides
# ---------------------------------------------------------------------

def alternatives_path_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Gestionarlo sola",
                    callback_data=SELF_MANAGED_PATH_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "Delegar el juicio",
                    callback_data=DELEGATED_PATH_CALLBACK,
                )
            ],
            [
                InlineKeyboardButton(
                    "Agencia relacional guiada",
                    callback_data=GUIDED_PATH_CALLBACK,
                )
            ],
        ]
    )


def design_mi_plan_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Diseña Mi Plan",
                    callback_data=DESIGN_MI_PLAN_CALLBACK,
                )
            ]
        ]
    )


def create_mi_plan_followup_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Crear seguimiento — mañana 10:00",
                    callback_data=CREATE_MI_PLAN_FOLLOWUP_CALLBACK,
                )
            ]
        ]
    )


async def handle_self_managed_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    facts = """Camino elegido: gestionarlo sola.

Meaning:
- The user tries to solve the romantic ambiguity alone.
- She may read about attachment, watch dating advice, ask friends, analyze messages, write and delete texts, wait, block, doubt.
- Pros: independence and learning.
- Cons: overanalysis, activation, impulsivity, confusion between limits and punishment, possible self-abandonment.

Finality:
Explain this path warmly and without shaming.
Make clear it can teach things, but when the nervous system is activated it can become chaotic.
Invite her to consider Agencia relacional guiada.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_relationship_pattern",
        task="Write the answer after the user chooses the self-managed relationship path.",
        facts=facts,
        reply_markup=alternatives_path_keyboard(),
    )


async def handle_delegated_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    facts = """Camino elegido: delegar el juicio.

Meaning:
- The user asks friends, rigid dating rules, therapy, the man, or eCoach to decide for her.
- She wants someone to say: write, wait, block, continue, leave.
- Pros: short-term relief.
- Cons: loss of internal authority, dependence, disconnection from values, relief without maturity.

Finality:
Explain this path warmly and without shaming.
Make clear that eCoach will not become another dependency.
Invite her to consider Agencia relacional guiada.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_relationship_pattern",
        task="Write the answer after the user chooses delegated judgement in relationship context.",
        facts=facts,
        reply_markup=alternatives_path_keyboard(),
    )





async def handle_guided_path_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    facts = """The user has chosen: Agencia relacional guiada.

Write ONE combined answer. Do not ask whether she wants Mi Plan. She chose guided agency, so give her the plan directly.

Context:
- Laura is seeing a man she likes.
- When they are together, everything seems good.
- Between dates there is ambiguity.
- Sometimes he writes warmly.
- Sometimes he disappears.
- Sometimes he proposes vague plans.
- Her psychologist has told her not to pursue and to observe coherence.
- When Laura gets activated, she forgets the guidance.

Core product idea:
The value is not that eCoach invents a hypothetical scenario.
The value is that Laura writes to eCoach when the real activation happens, before replying to him.
Then eCoach applies Mi Plan in real time.

Core idea:
- No persecución ansiosa.
- No juicio delegado.
- Agencia relacional guiada.

Boundary:
- eCoach does not replace the psychologist.
- The psychologist keeps therapeutic authority.
- Laura keeps the decision.
- eCoach helps between sessions, especially when the pattern activates.

The answer should include:

1. Very short explanation of agencia relacional guiada:
- Laura decides.
- eCoach helps her not lose herself when activated.
- eCoach helps her practice between sessions.

2. Mi Plan:
- Current relational question:
  ¿Esta conexión se está volviendo mutua y clara, o vivo de esperanza intermitente?

- Facts vs stories:
  Facts:
  - When they are together, connection feels good.
  - Sometimes he writes warmly.
  - Sometimes he disappears.
  - Sometimes he proposes vague plans.
  - He says he wants to see her but does not always concretize.

  Stories:
  - I did something wrong.
  - I am too intense.
  - If I ask for clarity, I will lose him.
  - I must adapt to his rhythm.
  - I should become colder.
  - I should block him.

- Values:
  clarity, warmth, reciprocity, not pursuing, not punishing, not self-abandoning, expressing a need calmly, observing facts not fantasies.

- Boundary experiment:
  Me gusta verte y lo paso bien contigo.
  A la vez, los planes muy vagos no me van bien.
  Si te apetece vernos, prefiero concretar día y hora.

- Observe response:
  Does he become clearer?
  Does he respect the boundary?
  Does he move toward her with facts?
  Does he punish clarity?
  Does the relationship make her more herself or less herself?

- Therapy preparation:
  what happened, what she felt, what she thought, what she did, where she abandoned herself, where she held herself, what to bring to therapy.

3. End with a real-time instruction:
Tell Laura:
When the real activation happens, write here before replying.
She can write something imperfect like:
"Me acaba de escribir esto: 'Quizá nos vemos mañana, te digo algo 😘'. Me estoy activando y quiero contestar ya."

Then eCoach will apply Mi Plan in real time:
pause, facts vs stories, values, possible reply, observe consistency, therapy material.

Do NOT offer to create a follow-up yet.
Do NOT show a follow-up button yet.
The follow-up comes only after a real activation has been processed.

Tone:
- Clear.
- Warm.
- Compact.
- Not too therapeutic.
- No "Con calidez".
- No repeated intro.
"""

    await answer_callback_with_skill(
        query=query,
        skill_name="manage_relationship_pattern",
        task="Write the combined Agencia relacional guiada + Mi Plan answer and invite real-time activation use.",
        facts=facts,
        reply_markup=None,
    )

async def handle_design_mi_plan_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    message = """La práctica real ocurre cuando aparezca la activación de verdad.

Cuando recibas un mensaje ambiguo, notes ansiedad y tengas impulso de responder rápido, escríbeme antes de contestar.

Puedes escribir algo tan simple como:

“Me acaba de escribir esto: ‘Quizá nos vemos mañana, te digo algo 😘’. Me estoy activando y quiero contestar ya.”

Entonces aplicaremos tu Mi Plan en tiempo real:
- pausa;
- hechos vs historias;
- valores;
- posible respuesta;
- observación de consistencia;
- material para tu psicóloga."""

    await query.message.reply_text(
        message,
        reply_markup=MAIN_KEYBOARD,
    )

async def handle_create_mi_plan_followup_button(update, context):
    query = update.callback_query
    await query.answer()
    await clear_clicked_inline_keyboard(query)

    activate_client_from_update(update)
    ensure_client_files()
    ensure_followup_triggers_file()

    tomorrow = today_app() + timedelta(days=1)
    followup_date = tomorrow.strftime("%Y-%m-%d")
    followup_time = "10:00"

    trigger = {
        "date": followup_date,
        "time": followup_time,
        "message_template": (
            "Seguimiento de agencia relacional guiada: revisar si apareció activación, "
            "separar hechos de historias, recordar valores relacionales y elegir el siguiente paso claro. "
            "Si hay algo importante, preparar material para la psicóloga."
        ),
        "reason": "Mi Plan de eCoach Relaciones — Laura",
        "source": "relaciones_mi_plan_button",
        "status": "pending",
    }

    saved_followups = save_immediate_followup_triggers(
        [trigger],
        source="relaciones_mi_plan_button",
    )

    if not saved_followups:
        await query.message.reply_text(
            "No he podido crear el seguimiento. Inténtalo de nuevo o revisa /scheduler_status.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    confirmation = """Listo, Laura. He creado el seguimiento para mañana a las 10:00.

Revisaremos:
- si apareció activación;
- si pudiste pausar antes de actuar;
- qué hechos había y qué historias construyó el miedo;
- qué valor relacional quieres cuidar ahora;
- y si conviene preparar algo para tu psicóloga.

No será para juzgarte. Si apareció ansiedad, no es un fracaso: es el momento exacto para practicar.

Mañana volvemos al siguiente paso claro."""

    await query.message.reply_text(
        confirmation,
        reply_markup=MAIN_KEYBOARD,
    )

async def handle_real_time_relationship_activation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    activate_client_from_update(update)
    ensure_client_files()

    user_text = update.message.text or ""

    facts = f"""The user is writing during a REAL relational activation.

This is not a hypothetical case created by eCoach.
The user has come to eCoach before replying, exactly as Mi Plan instructed.

User message:
{user_text}

Context from Mi Plan:
- Laura likes this man.
- When they are together, things feel good.
- Between dates, ambiguity activates her.
- Her psychologist told her not to pursue and to observe coherence.
- eCoach must help between therapy sessions without replacing the psychologist.

Task:
Apply Mi Plan in real time.

The answer should include:

1. Pause:
- This is activation, not failure.
- Do not answer in the first wave.
- Take 30 seconds.
- The goal is not to suppress emotion, but to avoid letting activation write the message.

2. Facts vs stories:
Use the user's actual message.
Facts may include:
- he wrote late;
- he used warm language or emoji;
- he used vague wording;
- he did not concretize day/time;
- Laura feels urgency to answer.

Stories may include:
- I am not important;
- I am too intense;
- I must accept vagueness;
- I must become cold;
- I will lose him if I ask clearly.

3. Values:
- clarity;
- warmth;
- reciprocity;
- not pursuing;
- not punishing;
- not self-abandoning.

4. Possible replies:
Give 3 options:
A. warm and clear;
B. short and calm;
C. more direct.

Recommended version should be warm-clear, close to:
"Me apetece verte. Para mí mañana funciona mejor si lo concretamos con algo de margen. Si te va bien, dime día/hora y lo organizamos."

Do not make it manipulative.
Do not make it cold.
Do not make it needy.
Do not decide for Laura.
Say she can choose the version that feels most aligned.

5. Observe consistency:
After sending a clear warm message, the task is not obsessive monitoring.
The task is to observe whether he brings more clarity or keeps ambiguity.

6. Material for psychologist:
Suggest saving:
- trigger;
- body reaction;
- facts;
- story;
- action chosen;
- how she felt after choosing from values.

7. End by offering the follow-up:
Say we can create tomorrow's 10:00 follow-up to review what happened.

Tone:
- Practical.
- Warm.
- Product-like.
- Not patrimonial.
- Not financial.
- Not generic.
- No sentence saying "soy guía patrimonial-financiero".
"""

    await answer_message_with_skill(
        update=update,
        context=context,
        skill_name="manage_relationship_pattern",
        task="Apply Mi Plan in real time to this relationship activation.",
        facts=facts,
        reply_markup=create_mi_plan_followup_keyboard(),
    )

def main() -> None:
    app = (
        ApplicationBuilder()
        .token(TOKEN)
        .concurrent_updates(8)
        .connection_pool_size(16)
        .pool_timeout(30)
        .build()
    )

    if app.job_queue is None:
        raise RuntimeError(
            "JobQueue is not available. Install it with: pip install \"python-telegram-bot[job-queue]\""
        )
    
    app.job_queue.run_daily(
        scheduled_run_followups,
        time=time(
            hour=PROACTIVE_SCHEDULER_HOUR,
            minute=PROACTIVE_SCHEDULER_MINUTE,
            tzinfo=ZoneInfo(APP_TIMEZONE),
        ),
        name="daily_followup_scheduler",
    )

    app.job_queue.run_repeating(
        scheduled_run_timed_followups,
        interval=60,
        first=10,
        name="timed_followup_scheduler",
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("diagnostics", diagnostics))
    app.add_handler(CommandHandler("version", version_command))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(CommandHandler("reset_client", reset_client))
    app.add_handler(CommandHandler("followups", show_followups))
    app.add_handler(CommandHandler("followup_help", followup_help))
    app.add_handler(CommandHandler("due_followups", show_due_followups))
    app.add_handler(CommandHandler("scheduler_status", scheduler_status))
    app.add_handler(CommandHandler("all_followups", show_all_followups))
    app.add_handler(CommandHandler("proactivity_log", show_proactivity_log))
    app.add_handler(CommandHandler("followup_done", client_locked_handler(mark_followup_done)))
    app.add_handler(CommandHandler("delete_followup", client_locked_handler(delete_followup)))
    app.add_handler(CommandHandler("followup_snooze", client_locked_handler(snooze_followup)))
    app.add_handler(CommandHandler("run_followups", client_locked_handler(run_followups)))
    app.add_handler(CommandHandler("agent_observations", show_agent_observations))
    app.add_handler(CommandHandler("reset_project", reset_ecoach_project))
    app.add_handler(CommandHandler("workflow_board", show_workflow_board))
    app.add_handler(CommandHandler("project_board", show_workflow_board))

    app.add_handler(MessageHandler(filters.Regex(r"^Cancelar$"), cancel))
    app.add_handler(MessageHandler(filters.Regex(r"^✅ Confirmar cambio memoria$"), confirm_memory_change))
    app.add_handler(MessageHandler(filters.Regex(r"^❌ Cancelar cambio memoria$"), cancel_memory_change))
    app.add_handler(MessageHandler(filters.Regex(r"^Quien soy$"), show_quien_soy))
    app.add_handler(MessageHandler(filters.Regex(r"^Que quiero$"), show_que_quiero))
    app.add_handler(MessageHandler(filters.Regex(r"^Plan de accion$"), show_que_tengo_que_hacer))
    app.add_handler(MessageHandler(filters.Regex(r"^Seguimientos$"), show_followups))
    app.add_handler(MessageHandler(filters.Regex(r"^Guardar sesion$"), show_session_buffer_status))
    app.add_handler(MessageHandler(filters.Regex(r"^✅ Confirmar guardado sesión$"), confirm_session_consolidation))
    app.add_handler(MessageHandler(filters.Regex(r"^❌ Cancelar guardado sesión$"), cancel_session_consolidation))
    app.add_handler(MessageHandler(filters.Regex(r"^🎯 Resumen de qué quiero$"), show_que_quiero))
    app.add_handler(MessageHandler(filters.Regex(r"^🧭 Resumen de qué tengo que hacer$"), show_que_tengo_que_hacer))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(analyze_uploaded_documents_callback_handler, pattern=r"^analyze_uploaded_documents$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(MessageHandler(filters.Document.ALL, handle_uploaded_document))
    # Disabled in eCoach Relaciones demo: app.add_handler(MessageHandler(filters.Regex(r"(?i)^\s*(analiza|analizar|lee|leer|revisa|revisar).*(documentos|pdf|excel|xlsx|cartera|fondos)[\s\.\!\?¡¿]*$"), analyze_uploaded_documents_handler))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_public_enrichment_button, pattern=f"^{PUBLIC_ENRICHMENT_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_private_bank_data_button, pattern=f"^{PRIVATE_BANK_DATA_CALLBACK}$"))
    app.add_handler(CallbackQueryHandler(handle_alternatives_button, pattern=f"^{ALTERNATIVES_CALLBACK}$"))
    app.add_handler(CallbackQueryHandler(handle_self_managed_path_button, pattern=f"^{SELF_MANAGED_PATH_CALLBACK}$"))
    app.add_handler(CallbackQueryHandler(handle_delegated_path_button, pattern=f"^{DELEGATED_PATH_CALLBACK}$"))
    app.add_handler(CallbackQueryHandler(handle_guided_path_button, pattern=f"^{GUIDED_PATH_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_compare_current_portfolio_button, pattern=f"^{COMPARE_CURRENT_PORTFOLIO_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_define_concrete_alternative_button, pattern=f"^{DEFINE_CONCRETE_ALTERNATIVE_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_prepare_provisional_alternative_button, pattern=f"^{PREPARE_PROVISIONAL_ALTERNATIVE_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_define_initial_alternative_button, pattern=f"^{DEFINE_INITIAL_ALTERNATIVE_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_search_providers_button, pattern=f"^{SEARCH_PROVIDERS_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_build_same_risk_lower_cost_portfolio_button, pattern="^build_same_risk_lower_cost_portfolio$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_review_before_execute_button, pattern=f"^{REVIEW_BEFORE_EXECUTE_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_search_funds_button, pattern=f"^{SEARCH_FUNDS_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_provider_selected_button, pattern=f"^{PROVIDER_MYINVESTOR_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_provider_selected_button, pattern=f"^{PROVIDER_RENTA4_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_provider_selected_button, pattern=f"^{PROVIDER_OPENBANK_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_provider_selected_button, pattern=f"^{PROVIDER_INDEXA_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_provider_selected_button, pattern=f"^{PROVIDER_INBESTME_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_missing_data_checklist_button, pattern=f"^{MISSING_DATA_CHECKLIST_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_keep_similar_risk_button, pattern=f"^{KEEP_SIMILAR_RISK_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(CallbackQueryHandler(handle_reduce_risk_button, pattern=f"^{REDUCE_RISK_CALLBACK}$"))
    # Disabled in eCoach Relaciones demo: app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(PUBLIC_ENRICHMENT_RE), handle_public_enrichment_command), group=0)

    app.add_handler(MessageHandler(
        filters.Regex(r"(?i)(me acaba de escribir|me estoy activando|estoy activada|quiero contestar ya|contestarle ya|hacerme la fría|hacerme la fria|son las 23:40|quizá nos vemos|quiza nos vemos)") & ~filters.COMMAND,
        client_locked_handler(handle_real_time_relationship_activation),
    ))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_free_text))

    app.add_handler(CallbackQueryHandler(handle_design_mi_plan_button, pattern=f"^{DESIGN_MI_PLAN_CALLBACK}$"))

    app.add_handler(CallbackQueryHandler(handle_create_mi_plan_followup_button, pattern=f"^{CREATE_MI_PLAN_FOLLOWUP_CALLBACK}$"))

    app.run_polling()


if __name__ == "__main__":
    main()







